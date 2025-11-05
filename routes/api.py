from flask import Blueprint, request, jsonify, g, Response
import uuid
import time
import json
from app import db
from models import Product, InventoryRecord, Alert, SalesRecord, ReorderRecommendation, KPIMetric
from services.alert_service import alert_system
from services.forecasting_service import demand_forecasting_service
from services.replenishment_service import replenishment_service
from services.kpi_service import kpi_service
from datetime import datetime, timedelta
import logging
from sqlalchemy import or_, asc, desc
from hashlib import md5
from hashlib import md5
import csv
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.auth import require_auth, require_permission, authenticate_user
from utils.notifications import get_notification_service
from utils.integrations import slack_notifier, webhook_publisher, telegram_notifier
from utils.rate_limit_redis import rate_limiter as redis_rate_limiter

# Crear blueprint para la API
api_bp = Blueprint('api', __name__)
# ================ MÉTRICAS PROMETHEUS (opcional) ================
_PROM_ENABLED = False
try:
    from prometheus_client import Counter, Histogram, generate_latest, CONTENT_TYPE_LATEST
    _PROM_ENABLED = True
    HTTP_REQUESTS_TOTAL = Counter('http_requests_total', 'Total de requests', ['method', 'endpoint', 'status'])
    HTTP_REQUEST_LATENCY = Histogram('http_request_latency_seconds', 'Latencia de requests', ['method', 'endpoint'])
except Exception:
    _PROM_ENABLED = False

# ==================== AUTENTICACIÓN ====================

# Request ID y logging por request
@api_bp.before_request
def _assign_request_id():
    g.request_id = str(uuid.uuid4())[:8]
    g._start_ts = time.time()
    # Rate limit persistente (si Redis configurado)
    try:
        if redis_rate_limiter.enabled:
            client_key = request.remote_addr or 'unknown'
            endpoint = request.endpoint or 'unknown'
            limit = int(os.getenv('RATE_LIMIT_PER_MINUTE', '60'))
            allowed, remaining = redis_rate_limiter.is_allowed(f"{client_key}:{endpoint}", limit)
            if not allowed:
                return jsonify({
                    'error': 'Rate limit exceeded',
                    'retry_after': 60
                }), 429
    except Exception:
        pass

@api_bp.after_request
def _add_correlation_headers(resp):
    try:
        if hasattr(g, 'request_id'):
            resp.headers['X-Request-ID'] = g.request_id
        duration_ms = 0
        if hasattr(g, '_start_ts'):
            duration_ms = int((time.time() - g._start_ts) * 1000)
        logger.info(f"{request.method} {request.path} -> {resp.status_code} ({duration_ms}ms)")
        # Métricas
        if _PROM_ENABLED:
            endpoint = request.endpoint or 'unknown'
            try:
                HTTP_REQUESTS_TOTAL.labels(request.method, endpoint, resp.status_code).inc()
                HTTP_REQUEST_LATENCY.labels(request.method, endpoint).observe(duration_ms / 1000.0)
            except Exception:
                pass
    except Exception:
        pass
    return resp

@api_bp.route('/metrics', methods=['GET'])
def metrics():
    if not _PROM_ENABLED:
        return jsonify({'error': 'Prometheus no habilitado'}), 503
    return Response(generate_latest(), mimetype=CONTENT_TYPE_LATEST)

@api_bp.route('/auth/login', methods=['POST'])
def login():
    """Endpoint de login para obtener token JWT"""
    try:
        data = request.get_json()
        if not data or 'username' not in data or 'password' not in data:
            return jsonify({'error': 'Username y password requeridos'}), 400
        
        username = data['username']
        password = data['password']
        
        result = authenticate_user(username, password)
        if not result:
            return jsonify({'error': 'Credenciales inválidas'}), 401
        
        return jsonify({
            'message': 'Login exitoso',
            'token': result['token'],
            'user': {
                'username': result['username'],
                'role': result['role'],
                'permissions': result['permissions']
            }
        })
    except Exception as e:
        logger.error(f'Error en login: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/auth/verify', methods=['GET'])
@require_auth
def verify_token():
    """Verifica si el token es válido"""
    return jsonify({
        'valid': True,
        'user': request.current_user
    })

# Configurar logging
logger = logging.getLogger(__name__)

# ==================== PRODUCTOS ====================

@api_bp.route('/products', methods=['GET'])
@require_auth
@require_permission('read')
def get_products():
    """Obtiene productos con filtros, orden y paginación opcional"""
    try:
        page = request.args.get('page', type=int)
        per_page = request.args.get('per_page', type=int)

        # Filtros
        q = request.args.get('q', type=str)
        sku = request.args.get('sku', type=str)
        category = request.args.get('category', type=str)
        min_price = request.args.get('min_price', type=float)
        max_price = request.args.get('max_price', type=float)

        # Orden
        sort = request.args.get('sort', default='id', type=str)
        order = request.args.get('order', default='asc', type=str)

        query = Product.query

        if q:
            like = f"%{q}%"
            query = query.filter(or_(
                Product.name.ilike(like),
                Product.sku.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like)
            ))

        if sku:
            query = query.filter(Product.sku.ilike(f"%{sku}%"))

        if category:
            query = query.filter(Product.category.ilike(f"%{category}%"))

        if min_price is not None:
            query = query.filter(Product.unit_price >= min_price)

        if max_price is not None:
            query = query.filter(Product.unit_price <= max_price)

        # Ordenamiento seguro
        sort_columns = {
            'id': Product.id,
            'name': Product.name,
            'sku': Product.sku,
            'category': Product.category,
            'unit_price': Product.unit_price,
            'created_at': Product.created_at
        }
        if sort in sort_columns:
            query = query.order_by(asc(sort_columns[sort]) if order == 'asc' else desc(sort_columns[sort]))

        if page and per_page:
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            items = pagination.items
            data = [{
                'id': p.id,
                'name': p.name,
                'sku': p.sku,
                'description': p.description,
                'category': p.category,
                'unit_price': p.unit_price,
                'cost_price': p.cost_price,
                'min_stock_level': p.min_stock_level,
                'max_stock_level': p.max_stock_level,
                'reorder_point': p.reorder_point,
                'supplier_id': p.supplier_id,
                'supplier_name': p.supplier.name if p.supplier else None,
                'current_stock': alert_system.get_current_stock(p.id)
            } for p in items]

            payload = {
                'items': data,
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages
            }
            etag = md5(str(payload).encode('utf-8')).hexdigest()
            resp = jsonify(payload)
            resp.headers['ETag'] = etag
            resp.headers['Cache-Control'] = 'public, max-age=30'
            return resp
        else:
            products = query.all()
            payload = [{
                'id': p.id,
                'name': p.name,
                'sku': p.sku,
                'description': p.description,
                'category': p.category,
                'unit_price': p.unit_price,
                'cost_price': p.cost_price,
                'min_stock_level': p.min_stock_level,
                'max_stock_level': p.max_stock_level,
                'reorder_point': p.reorder_point,
                'supplier_id': p.supplier_id,
                'supplier_name': p.supplier.name if p.supplier else None,
                'current_stock': alert_system.get_current_stock(p.id)
            } for p in products]
            etag = md5(str(payload).encode('utf-8')).hexdigest()
            resp = jsonify(payload)
            resp.headers['ETag'] = etag
            resp.headers['Cache-Control'] = 'public, max-age=60'
            return resp
    except Exception as e:
        logger.error(f'Error obteniendo productos: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/products/export', methods=['GET'])
def export_products_csv():
    """Exporta productos a CSV usando los mismos filtros de GET /products"""
    try:
        # Reutilizar lógica de filtros
        q = request.args.get('q', type=str)
        sku = request.args.get('sku', type=str)
        category = request.args.get('category', type=str)
        min_price = request.args.get('min_price', type=float)
        max_price = request.args.get('max_price', type=float)

        query = Product.query
        if q:
            like = f"%{q}%"
            query = query.filter(or_(
                Product.name.ilike(like),
                Product.sku.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like)
            ))
        if sku:
            query = query.filter(Product.sku.ilike(f"%{sku}%"))
        if category:
            query = query.filter(Product.category.ilike(f"%{category}%"))
        if min_price is not None:
            query = query.filter(Product.unit_price >= min_price)
        if max_price is not None:
            query = query.filter(Product.unit_price <= max_price)

        products = query.all()

        # Crear CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'name', 'sku', 'category', 'unit_price', 'cost_price', 'current_stock'])
        for p in products:
            writer.writerow([
                p.id, p.name, p.sku, p.category, p.unit_price, p.cost_price,
                alert_system.get_current_stock(p.id)
            ])
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=products.csv'
            }
        )
    except Exception as e:
        logger.error(f'Error exportando productos: {str(e)}')
        return jsonify({'error': 'Error exportando productos'}), 500

@api_bp.route('/products/export/excel', methods=['GET'])
def export_products_excel():
    """Exporta productos a Excel usando los mismos filtros de GET /products"""
    try:
        # Reutilizar lógica de filtros
        q = request.args.get('q', type=str)
        sku = request.args.get('sku', type=str)
        category = request.args.get('category', type=str)
        min_price = request.args.get('min_price', type=float)
        max_price = request.args.get('max_price', type=float)

        query = Product.query
        if q:
            like = f"%{q}%"
            query = query.filter(or_(
                Product.name.ilike(like),
                Product.sku.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like)
            ))
        if sku:
            query = query.filter(Product.sku.ilike(f"%{sku}%"))
        if category:
            query = query.filter(Product.category.ilike(f"%{category}%"))
        if min_price is not None:
            query = query.filter(Product.unit_price >= min_price)
        if max_price is not None:
            query = query.filter(Product.unit_price <= max_price)

        products = query.all()

        # Crear Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ['ID', 'Nombre', 'SKU', 'Categoría', 'Precio Unitario', 'Precio Costo', 'Stock Actual', 'Stock Mínimo', 'Stock Máximo', 'Proveedor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Datos
        for row, p in enumerate(products, 2):
            ws.cell(row=row, column=1, value=p.id)
            ws.cell(row=row, column=2, value=p.name)
            ws.cell(row=row, column=3, value=p.sku)
            ws.cell(row=row, column=4, value=p.category)
            ws.cell(row=row, column=5, value=p.unit_price)
            ws.cell(row=row, column=6, value=p.cost_price)
            ws.cell(row=row, column=7, value=alert_system.get_current_stock(p.id))
            ws.cell(row=row, column=8, value=p.min_stock_level)
            ws.cell(row=row, column=9, value=p.max_stock_level)
            ws.cell(row=row, column=10, value=p.supplier.name if p.supplier else '')

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=products.xlsx'
            }
        )
    except Exception as e:
        logger.error(f'Error exportando productos a Excel: {str(e)}')
        return jsonify({'error': 'Error exportando productos a Excel'}), 500

@api_bp.route('/products', methods=['POST'])
@require_auth
@require_permission('write')
def create_product():
    """Crea un nuevo producto"""
    try:
        data = request.get_json()
        
        product = Product(
            name=data['name'],
            sku=data['sku'],
            description=data.get('description', ''),
            category=data.get('category', ''),
            unit_price=data['unit_price'],
            cost_price=data['cost_price'],
            min_stock_level=data.get('min_stock_level', 10),
            max_stock_level=data.get('max_stock_level', 100),
            reorder_point=data.get('reorder_point', 20),
            supplier_id=data.get('supplier_id')
        )
        
        db.session.add(product)
        db.session.commit()
        
        return jsonify({'id': product.id, 'message': 'Producto creado exitosamente'}), 201
        
    except Exception as e:
        logger.error(f'Error creando producto: {str(e)}')
        db.session.rollback()
        return jsonify({'error': 'Error creando producto'}), 500

@api_bp.route('/products/<int:product_id>', methods=['GET'])
@require_auth
@require_permission('read')
def get_product(product_id):
    """Obtiene un producto específico"""
    try:
        product = Product.query.get_or_404(product_id)
        return jsonify({
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'description': product.description,
            'category': product.category,
            'unit_price': product.unit_price,
            'cost_price': product.cost_price,
            'min_stock_level': product.min_stock_level,
            'max_stock_level': product.max_stock_level,
            'reorder_point': product.reorder_point,
            'supplier_id': product.supplier_id,
            'supplier_name': product.supplier.name if product.supplier else None,
            'current_stock': alert_system.get_current_stock(product.id)
        })
    except Exception as e:
        logger.error(f'Error obteniendo producto {product_id}: {str(e)}')
        return jsonify({'error': 'Producto no encontrado'}), 404

# ==================== INVENTARIO ====================

@api_bp.route('/inventory', methods=['GET'])
def get_inventory():
    """Obtiene estado actual del inventario (con paginación opcional)"""
    try:
        page = request.args.get('page', type=int)
        per_page = request.args.get('per_page', type=int)

        query = Product.query
        if page and per_page:
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            products = pagination.items
        else:
            products = query.all()
        inventory = []
        
        for product in products:
            current_stock = alert_system.get_current_stock(product.id)
            inventory.append({
                'product_id': product.id,
                'product_name': product.name,
                'sku': product.sku,
                'current_stock': current_stock,
                'min_stock_level': product.min_stock_level,
                'max_stock_level': product.max_stock_level,
                'reorder_point': product.reorder_point,
                'status': 'low' if current_stock <= product.min_stock_level else 'normal',
                'value': current_stock * product.unit_price
            })
        
        if page and per_page:
            return jsonify({
                'items': inventory,
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages
            })
        else:
            return jsonify(inventory)
    except Exception as e:
        logger.error(f'Error obteniendo inventario: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/inventory/movements', methods=['POST'])
def record_inventory_movement():
    """Registra un movimiento de inventario"""
    try:
        data = request.get_json()
        
        movement = InventoryRecord(
            product_id=data['product_id'],
            quantity=data['quantity'],
            movement_type=data['movement_type'],  # 'in', 'out', 'adjustment'
            reference=data.get('reference', ''),
            notes=data.get('notes', '')
        )
        
        db.session.add(movement)
        db.session.commit()
        
        # Verificar alertas después del movimiento
        alert_system.check_low_stock_alerts()
        
        return jsonify({'id': movement.id, 'message': 'Movimiento registrado exitosamente'}), 201
        
    except Exception as e:
        logger.error(f'Error registrando movimiento: {str(e)}')
        db.session.rollback()
        return jsonify({'error': 'Error registrando movimiento'}), 500

# ==================== ALERTAS ====================

@api_bp.route('/alerts', methods=['GET'])
def get_alerts():
    """Obtiene todas las alertas"""
    try:
        severity = request.args.get('severity')
        alerts = alert_system.get_active_alerts(severity)
        
        return jsonify([{
            'id': alert.id,
            'product_id': alert.product_id,
            'product_name': alert.product.name,
            'alert_type': alert.alert_type,
            'message': alert.message,
            'severity': alert.severity,
            'is_read': alert.is_read,
            'is_resolved': alert.is_resolved,
            'created_at': alert.created_at.isoformat(),
            'resolved_at': alert.resolved_at.isoformat() if alert.resolved_at else None
        } for alert in alerts])
    except Exception as e:
        logger.error(f'Error obteniendo alertas: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/alerts/<int:alert_id>/resolve', methods=['POST'])
def resolve_alert(alert_id):
    """Marca una alerta como resuelta"""
    try:
        success = alert_system.resolve_alert(alert_id)
        if success:
            return jsonify({'message': 'Alerta resuelta exitosamente'})
        else:
            return jsonify({'error': 'Alerta no encontrada'}), 404
    except Exception as e:
        logger.error(f'Error resolviendo alerta {alert_id}: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/alerts/check', methods=['POST'])
def check_alerts():
    """Ejecuta verificación manual de alertas"""
    try:
        alerts_created = alert_system.check_low_stock_alerts()
        return jsonify({'message': f'Se crearon {alerts_created} alertas'})
    except Exception as e:
        logger.error(f'Error verificando alertas: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

# ==================== PREDICCIONES ====================

@api_bp.route('/forecasts/<int:product_id>', methods=['GET'])
def get_forecast(product_id):
    """Obtiene predicción de demanda para un producto"""
    try:
        days_ahead = request.args.get('days', 30, type=int)
        method = request.args.get('method', 'auto')
        
        forecast = demand_forecasting_service.forecast_demand(product_id, days_ahead, method)
        return jsonify(forecast)
    except Exception as e:
        logger.error(f'Error obteniendo predicción para producto {product_id}: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/forecasts', methods=['GET'])
def get_all_forecasts():
    """Obtiene resumen de todas las predicciones"""
    try:
        summary = demand_forecasting_service.get_forecast_summary()
        return jsonify(summary)
    except Exception as e:
        logger.error(f'Error obteniendo resumen de predicciones: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

# ==================== REPOSICIÓN ====================

@api_bp.route('/replenishment/recommendations', methods=['GET'])
def get_reorder_recommendations():
    """Obtiene recomendaciones de reposición"""
    try:
        recommendations = replenishment_service.generate_reorder_recommendations()
        return jsonify(recommendations)
    except Exception as e:
        logger.error(f'Error obteniendo recomendaciones: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/replenishment/recommendations/<int:recommendation_id>/process', methods=['POST'])
def process_recommendation(recommendation_id):
    """Procesa una recomendación de reposición"""
    try:
        success = replenishment_service.process_reorder_recommendation(recommendation_id)
        if success:
            return jsonify({'message': 'Recomendación procesada exitosamente'})
        else:
            return jsonify({'error': 'Recomendación no encontrada'}), 404
    except Exception as e:
        logger.error(f'Error procesando recomendación {recommendation_id}: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

# ==================== KPIs ====================

@api_bp.route('/kpis', methods=['GET'])
def get_kpis():
    """Obtiene todos los KPIs"""
    try:
        kpis = kpi_service.calculate_all_kpis()
        return jsonify(kpis)
    except Exception as e:
        logger.error(f'Error obteniendo KPIs: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/kpis/trends', methods=['GET'])
def get_kpi_trends():
    """Obtiene tendencias de KPIs"""
    try:
        days = request.args.get('days', 30, type=int)
        trends = kpi_service.get_kpi_trends(days)
        return jsonify(trends)
    except Exception as e:
        logger.error(f'Error obteniendo tendencias de KPIs: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

# ==================== VENTAS ====================

@api_bp.route('/sales', methods=['POST'])
def record_sale():
    """Registra una venta"""
    try:
        data = request.get_json()
        
        sale = SalesRecord(
            product_id=data['product_id'],
            quantity_sold=data['quantity_sold'],
            sale_date=datetime.fromisoformat(data['sale_date']),
            unit_price=data['unit_price'],
            total_amount=data['quantity_sold'] * data['unit_price'],
            customer_id=data.get('customer_id')
        )
        
        db.session.add(sale)
        
        # Registrar salida de inventario
        inventory_movement = InventoryRecord(
            product_id=data['product_id'],
            quantity=data['quantity_sold'],
            movement_type='out',
            reference=f'Sale-{sale.id}',
            notes='Venta registrada'
        )
        
        db.session.add(inventory_movement)
        db.session.commit()

        # Publicar eventos y notificar
        try:
            webhook_publisher.publish('sale.created', {
                'id': sale.id,
                'product_id': sale.product_id,
                'quantity_sold': sale.quantity_sold,
                'total_amount': sale.total_amount,
                'sale_date': sale.sale_date.isoformat(),
                'customer_id': sale.customer_id
            })
            slack_notifier.send_message(f":moneybag: Venta registrada ID {sale.id} | Prod {sale.product_id} | Cant {sale.quantity_sold} | Total ${sale.total_amount:,.2f}")
        except Exception as notify_err:
            logger.warning(f"No se pudo notificar evento de venta: {notify_err}")

        # Verificar alertas y notificar stock bajo
        try:
            alert_system.check_low_stock_alerts()
            current_stock = alert_system.get_current_stock(sale.product_id)
            product = Product.query.get(sale.product_id)
            if product and current_stock is not None and product.min_stock_level is not None:
                if current_stock <= product.min_stock_level:
                    slack_notifier.send_message(
                        f":warning: Stock bajo {product.name} (SKU {product.sku}) | Stock {current_stock} / Min {product.min_stock_level}")
                    webhook_publisher.publish('inventory.low_stock', {
                        'product_id': product.id,
                        'sku': product.sku,
                        'name': product.name,
                        'current_stock': current_stock,
                        'min_stock_level': product.min_stock_level
                    })
        except Exception as e_alert:
            logger.warning(f"Error verificando/notificando stock bajo: {e_alert}")

        return jsonify({'id': sale.id, 'message': 'Venta registrada exitosamente'}), 201
        
    except Exception as e:
        logger.error(f'Error registrando venta: {str(e)}')
        db.session.rollback()
        return jsonify({'error': 'Error registrando venta'}), 500

@api_bp.route('/sales', methods=['GET'])
def get_sales():
    """Obtiene registros de ventas (con paginación y filtros opcionales)"""
    try:
        # Filtros de fecha
        days = request.args.get('days', type=int)
        start_date_param = request.args.get('start_date')  # ISO: YYYY-MM-DD
        end_date_param = request.args.get('end_date')      # ISO: YYYY-MM-DD

        if start_date_param:
            start_date = datetime.fromisoformat(start_date_param)
        elif days is not None:
            start_date = datetime.utcnow() - timedelta(days=days)
        else:
            start_date = datetime.utcnow() - timedelta(days=30)

        end_date = None
        if end_date_param:
            # Incluir fin del día
            end_date = datetime.fromisoformat(end_date_param) + timedelta(days=1)

        page = request.args.get('page', type=int)
        per_page = request.args.get('per_page', type=int)
        customer_id = request.args.get('customer_id', type=int)

        query = SalesRecord.query
        query = query.filter(SalesRecord.sale_date >= start_date)
        if end_date:
            query = query.filter(SalesRecord.sale_date < end_date)
        if customer_id:
            query = query.filter(SalesRecord.customer_id == customer_id)
        query = query.order_by(SalesRecord.sale_date.desc())

        if page and per_page:
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            sales = pagination.items
        else:
            sales = query.all()

        items = [{
            'id': sale.id,
            'product_id': sale.product_id,
            'product_name': sale.product.name if getattr(sale, 'product', None) else None,
            'quantity_sold': sale.quantity_sold,
            'sale_date': sale.sale_date.isoformat(),
            'unit_price': sale.unit_price,
            'total_amount': sale.total_amount,
            'customer_id': sale.customer_id
        } for sale in sales]

        if page and per_page:
            payload = {
                'items': items,
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages
            }
            etag = md5(str(payload).encode('utf-8')).hexdigest()
            resp = jsonify(payload)
            resp.headers['ETag'] = etag
            resp.headers['Cache-Control'] = 'public, max-age=30'
            return resp
        else:
            etag = md5(str(items).encode('utf-8')).hexdigest()
            resp = jsonify(items)
            resp.headers['ETag'] = etag
            resp.headers['Cache-Control'] = 'public, max-age=60'
            return resp
    except Exception as e:
        logger.error(f'Error obteniendo ventas: {str(e)}')
        return jsonify({'error': 'Error interno del servidor'}), 500

@api_bp.route('/sales/export', methods=['GET'])
def export_sales_csv():
    """Exporta ventas a CSV con filtros opcionales"""
    try:
        days = request.args.get('days', type=int)
        start_date_param = request.args.get('start_date')
        end_date_param = request.args.get('end_date')
        customer_id = request.args.get('customer_id', type=int)

        if start_date_param:
            start_date = datetime.fromisoformat(start_date_param)
        elif days is not None:
            start_date = datetime.utcnow() - timedelta(days=days)
        else:
            start_date = datetime.utcnow() - timedelta(days=30)

        end_date = None
        if end_date_param:
            end_date = datetime.fromisoformat(end_date_param) + timedelta(days=1)

        query = SalesRecord.query.filter(SalesRecord.sale_date >= start_date)
        if end_date:
            query = query.filter(SalesRecord.sale_date < end_date)
        if customer_id:
            query = query.filter(SalesRecord.customer_id == customer_id)
        query = query.order_by(SalesRecord.sale_date.desc())

        sales = query.all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'product_id', 'product_name', 'quantity_sold', 'unit_price', 'total_amount', 'sale_date', 'customer_id'])
        for s in sales:
            writer.writerow([
                s.id, s.product_id, s.product.name if getattr(s, 'product', None) else None,
                s.quantity_sold, s.unit_price, s.total_amount,
                s.sale_date.isoformat(), s.customer_id
            ])
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=sales.csv'
            }
        )
    except Exception as e:
        logger.error(f'Error exportando ventas: {str(e)}')
        return jsonify({'error': 'Error exportando ventas'}), 500

@api_bp.route('/sales/export/excel', methods=['GET'])
def export_sales_excel():
    """Exporta ventas a Excel con filtros opcionales"""
    try:
        days = request.args.get('days', type=int)
        start_date_param = request.args.get('start_date')
        end_date_param = request.args.get('end_date')
        customer_id = request.args.get('customer_id', type=int)

        if start_date_param:
            start_date = datetime.fromisoformat(start_date_param)
        elif days is not None:
            start_date = datetime.utcnow() - timedelta(days=days)
        else:
            start_date = datetime.utcnow() - timedelta(days=30)

        end_date = None
        if end_date_param:
            end_date = datetime.fromisoformat(end_date_param) + timedelta(days=1)

        query = SalesRecord.query.filter(SalesRecord.sale_date >= start_date)
        if end_date:
            query = query.filter(SalesRecord.sale_date < end_date)
        if customer_id:
            query = query.filter(SalesRecord.customer_id == customer_id)
        query = query.order_by(SalesRecord.sale_date.desc())

        sales = query.all()

        # Crear Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ['ID', 'Producto ID', 'Nombre Producto', 'Cantidad', 'Precio Unitario', 'Total', 'Fecha Venta', 'Cliente ID']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Datos
        for row, s in enumerate(sales, 2):
            ws.cell(row=row, column=1, value=s.id)
            ws.cell(row=row, column=2, value=s.product_id)
            ws.cell(row=row, column=3, value=s.product.name if getattr(s, 'product', None) else '')
            ws.cell(row=row, column=4, value=s.quantity_sold)
            ws.cell(row=row, column=5, value=s.unit_price)
            ws.cell(row=row, column=6, value=s.total_amount)
            ws.cell(row=row, column=7, value=s.sale_date.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=8, value=s.customer_id)

        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=sales.xlsx'
            }
        )
    except Exception as e:
        logger.error(f'Error exportando ventas a Excel: {str(e)}')
        return jsonify({'error': 'Error exportando ventas a Excel'}), 500

# ==================== REPORTES AGREGADOS ====================

@api_bp.route('/reports/sales/daily', methods=['GET'])
def get_daily_sales_report():
    """Reporte de ventas por día"""
    try:
        days = request.args.get('days', 30, type=int)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Agrupar ventas por día
        from sqlalchemy import func, cast, Date
        daily_sales = db.session.query(
            cast(SalesRecord.sale_date, Date).label('date'),
            func.sum(SalesRecord.total_amount).label('total_revenue'),
            func.sum(SalesRecord.quantity_sold).label('total_quantity'),
            func.count(SalesRecord.id).label('total_transactions')
        ).filter(
            SalesRecord.sale_date >= start_date
        ).group_by(
            cast(SalesRecord.sale_date, Date)
        ).order_by('date').all()
        
        result = []
        for day in daily_sales:
            result.append({
                'date': day.date.isoformat(),
                'total_revenue': float(day.total_revenue or 0),
                'total_quantity': int(day.total_quantity or 0),
                'total_transactions': int(day.total_transactions or 0)
            })
        
        return jsonify({
            'period_days': days,
            'daily_sales': result
        })
    except Exception as e:
        logger.error(f'Error generando reporte de ventas diarias: {str(e)}')
        return jsonify({'error': 'Error generando reporte'}), 500

@api_bp.route('/reports/products/top', methods=['GET'])
def get_top_products_report():
    """Reporte de productos más vendidos"""
    try:
        days = request.args.get('days', 30, type=int)
        limit = request.args.get('limit', 10, type=int)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        # Top productos por cantidad vendida
        top_products = db.session.query(
            Product.id,
            Product.name,
            Product.sku,
            Product.category,
            func.sum(SalesRecord.quantity_sold).label('total_quantity'),
            func.sum(SalesRecord.total_amount).label('total_revenue'),
            func.count(SalesRecord.id).label('total_sales')
        ).join(
            SalesRecord, Product.id == SalesRecord.product_id
        ).filter(
            SalesRecord.sale_date >= start_date
        ).group_by(
            Product.id, Product.name, Product.sku, Product.category
        ).order_by(
            func.sum(SalesRecord.quantity_sold).desc()
        ).limit(limit).all()
        
        result = []
        for product in top_products:
            result.append({
                'product_id': product.id,
                'name': product.name,
                'sku': product.sku,
                'category': product.category,
                'total_quantity_sold': int(product.total_quantity or 0),
                'total_revenue': float(product.total_revenue or 0),
                'total_sales': int(product.total_sales or 0)
            })
        
        return jsonify({
            'period_days': days,
            'top_products': result
        })
    except Exception as e:
        logger.error(f'Error generando reporte de top productos: {str(e)}')
        return jsonify({'error': 'Error generando reporte'}), 500

@api_bp.route('/reports/categories/summary', methods=['GET'])
def get_categories_summary():
    """Resumen de ventas por categoría"""
    try:
        days = request.args.get('days', 30, type=int)
        start_date = datetime.utcnow() - timedelta(days=days)
        
        category_sales = db.session.query(
            Product.category,
            func.sum(SalesRecord.total_amount).label('total_revenue'),
            func.sum(SalesRecord.quantity_sold).label('total_quantity'),
            func.count(SalesRecord.id).label('total_sales'),
            func.count(func.distinct(Product.id)).label('unique_products')
        ).join(
            SalesRecord, Product.id == SalesRecord.product_id
        ).filter(
            SalesRecord.sale_date >= start_date
        ).group_by(
            Product.category
        ).order_by(
            func.sum(SalesRecord.total_amount).desc()
        ).all()
        
        result = []
        for cat in category_sales:
            result.append({
                'category': cat.category or 'Sin categoría',
                'total_revenue': float(cat.total_revenue or 0),
                'total_quantity': int(cat.total_quantity or 0),
                'total_sales': int(cat.total_sales or 0),
                'unique_products': int(cat.unique_products or 0)
            })
        
        return jsonify({
            'period_days': days,
            'categories': result
        })
    except Exception as e:
        logger.error(f'Error generando reporte de categorías: {str(e)}')
        return jsonify({'error': 'Error generando reporte'}), 500

@api_bp.route('/reports/inventory/status', methods=['GET'])
def get_inventory_status_report():
    """Reporte del estado del inventario"""
    try:
        # Productos con stock bajo
        low_stock_products = Product.query.filter(
            Product.min_stock_level > 0
        ).all()
        
        low_stock = []
        for product in low_stock_products:
            current_stock = alert_system.get_current_stock(product.id)
            if current_stock <= product.min_stock_level:
                low_stock.append({
                    'product_id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'current_stock': current_stock,
                    'min_stock_level': product.min_stock_level,
                    'status': 'low_stock'
                })
        
        # Productos sin stock
        out_of_stock = []
        for product in low_stock_products:
            current_stock = alert_system.get_current_stock(product.id)
            if current_stock == 0:
                out_of_stock.append({
                    'product_id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'current_stock': current_stock,
                    'status': 'out_of_stock'
                })
        
        # Resumen general
        total_products = Product.query.count()
        total_value = 0
        for product in Product.query.all():
            current_stock = alert_system.get_current_stock(product.id)
            total_value += current_stock * product.cost_price
        
        return jsonify({
            'summary': {
                'total_products': total_products,
                'low_stock_count': len(low_stock),
                'out_of_stock_count': len(out_of_stock),
                'total_inventory_value': round(total_value, 2)
            },
            'low_stock_products': low_stock,
            'out_of_stock_products': out_of_stock
        })
    except Exception as e:
        logger.error(f'Error generando reporte de inventario: {str(e)}')
        return jsonify({'error': 'Error generando reporte'}), 500

# ==================== DASHBOARD ====================

@api_bp.route('/dashboard/summary', methods=['GET'])
def get_dashboard_summary():
    """Obtiene resumen completo para el dashboard con KPIs y gráficos"""
    try:
        # Obtener estadísticas rápidas
        total_products = Product.query.count()
        active_alerts = Alert.query.filter(Alert.is_resolved == False).count()
        
        # Calcular valor total del inventario
        total_inventory_value = 0
        low_stock_count = 0
        for product in Product.query.all():
            current_stock = alert_system.get_current_stock(product.id)
            total_inventory_value += current_stock * product.cost_price
            if current_stock <= product.min_stock_level:
                low_stock_count += 1
        
        # Ventas de los últimos 30 días
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        recent_sales = SalesRecord.query.filter(
            SalesRecord.sale_date >= thirty_days_ago
        ).all()
        
        total_revenue_30d = sum(sale.total_amount for sale in recent_sales)
        total_quantity_30d = sum(sale.quantity_sold for sale in recent_sales)
        
        # Ventas de los últimos 7 días para comparación
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_sales_7d = SalesRecord.query.filter(
            SalesRecord.sale_date >= seven_days_ago
        ).all()
        
        total_revenue_7d = sum(sale.total_amount for sale in recent_sales_7d)
        
        # Calcular crecimiento de ventas
        if total_revenue_7d > 0:
            growth_rate = ((total_revenue_30d - total_revenue_7d) / total_revenue_7d) * 100
        else:
            growth_rate = 0
        
        # Top 5 productos más vendidos (últimos 30 días)
        from sqlalchemy import func
        top_products = db.session.query(
            Product.id,
            Product.name,
            Product.sku,
            func.sum(SalesRecord.quantity_sold).label('total_quantity')
        ).join(
            SalesRecord, Product.id == SalesRecord.product_id
        ).filter(
            SalesRecord.sale_date >= thirty_days_ago
        ).group_by(
            Product.id, Product.name, Product.sku
        ).order_by(
            func.sum(SalesRecord.quantity_sold).desc()
        ).limit(5).all()
        
        top_products_data = []
        for product in top_products:
            top_products_data.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'quantity_sold': int(product.total_quantity or 0)
            })
        
        # Ventas por día (últimos 7 días) para gráfico
        from sqlalchemy import cast, Date
        daily_sales = db.session.query(
            cast(SalesRecord.sale_date, Date).label('date'),
            func.sum(SalesRecord.total_amount).label('revenue')
        ).filter(
            SalesRecord.sale_date >= seven_days_ago
        ).group_by(
            cast(SalesRecord.sale_date, Date)
        ).order_by('date').all()
        
        daily_sales_data = []
        for day in daily_sales:
            daily_sales_data.append({
                'date': day.date.isoformat(),
                'revenue': float(day.revenue or 0)
            })
        
        # Ventas por categoría (últimos 30 días)
        category_sales = db.session.query(
            Product.category,
            func.sum(SalesRecord.total_amount).label('revenue')
        ).join(
            SalesRecord, Product.id == SalesRecord.product_id
        ).filter(
            SalesRecord.sale_date >= thirty_days_ago
        ).group_by(
            Product.category
        ).order_by(
            func.sum(SalesRecord.total_amount).desc()
        ).all()
        
        category_data = []
        for cat in category_sales:
            category_data.append({
                'category': cat.category or 'Sin categoría',
                'revenue': float(cat.revenue or 0)
            })
        
        return jsonify({
            'summary': {
                'total_products': total_products,
                'active_alerts': active_alerts,
                'low_stock_count': low_stock_count,
                'total_inventory_value': round(total_inventory_value, 2),
                'total_revenue_30d': round(total_revenue_30d, 2),
                'total_quantity_30d': total_quantity_30d,
                'growth_rate': round(growth_rate, 2)
            },
            'charts': {
                'daily_sales': daily_sales_data,
                'category_sales': category_data
            },
            'top_products': top_products_data
        })
    except Exception as e:
        logger.error(f'Error obteniendo resumen del dashboard: {str(e)}')
        return jsonify({'error': 'Error obteniendo resumen del dashboard'}), 500

@api_bp.route('/dashboard/kpis', methods=['GET'])
def get_dashboard_kpis():
    """Obtiene KPIs detallados para el dashboard"""
    try:
        # KPIs de inventario
        inventory_kpis = kpi_service.get_inventory_kpis()
        
        # KPIs de ventas
        sales_kpis = kpi_service.get_sales_kpis()
        
        # KPIs financieros
        financial_kpis = kpi_service.get_financial_kpis()
        
        return jsonify({
            'inventory': inventory_kpis,
            'sales': sales_kpis,
            'financial': financial_kpis,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        logger.error(f'Error obteniendo KPIs del dashboard: {str(e)}')
        return jsonify({'error': 'Error obteniendo KPIs'}), 500

# ==================== NOTIFICACIONES ====================

@api_bp.route('/notifications/test', methods=['POST'])
@require_auth
@require_permission('admin')
def test_notifications():
    """Prueba el sistema de notificaciones"""
    try:
        notification_service = get_notification_service()
        if not notification_service:
            return jsonify({'error': 'Servicio de notificaciones no disponible'}), 500
        
        success, message = notification_service.test_email_connection()
        
        return jsonify({
            'success': success,
            'message': message
        })
    except Exception as e:
        logger.error(f'Error probando notificaciones: {str(e)}')
        return jsonify({'error': 'Error probando notificaciones'}), 500

@api_bp.route('/notifications/send-daily-summary', methods=['POST'])
@require_auth
@require_permission('admin')
def send_daily_summary():
    """Envía resumen diario manualmente"""
    try:
        notification_service = get_notification_service()
        if not notification_service:
            return jsonify({'error': 'Servicio de notificaciones no disponible'}), 500
        
        # Obtener datos del resumen
        from datetime import datetime, timedelta
        today = datetime.utcnow().date()
        
        # Ventas de hoy
        today_sales = SalesRecord.query.filter(
            db.func.date(SalesRecord.sale_date) == today
        ).all()
        
        revenue = sum(sale.total_amount for sale in today_sales)
        transactions = len(today_sales)
        
        # Productos con stock bajo
        low_stock_count = 0
        for product in Product.query.all():
            current_stock = alert_system.get_current_stock(product.id)
            if current_stock <= product.min_stock_level:
                low_stock_count += 1
        
        # Alertas
        active_alerts = Alert.query.filter(Alert.is_resolved == False).count()
        resolved_alerts = Alert.query.filter(
            Alert.is_resolved == True,
            Alert.resolved_at >= today
        ).count()
        
        summary_data = {
            'revenue': revenue,
            'transactions': transactions,
            'total_products': Product.query.count(),
            'low_stock': low_stock_count,
            'active_alerts': active_alerts,
            'resolved_alerts': resolved_alerts
        }
        
        success = notification_service.send_daily_summary(summary_data)
        
        return jsonify({
            'success': success,
            'message': 'Resumen diario enviado' if success else 'Error enviando resumen'
        })
    except Exception as e:
        logger.error(f'Error enviando resumen diario: {str(e)}')
        return jsonify({'error': 'Error enviando resumen diario'}), 500

# ==================== INTEGRACIONES (SLACK / WEBHOOKS) ====================

@api_bp.route('/notifications/slack/test', methods=['POST'])
@require_auth
@require_permission('admin')
def slack_test():
    try:
        ok = slack_notifier.send_message(':white_check_mark: Prueba Slack OK - Inventario')
        return jsonify({'success': ok, 'configured': slack_notifier.is_configured()})
    except Exception as e:
        logger.error(f'Error Slack test: {str(e)}')
        return jsonify({'error': 'Error enviando a Slack'}), 500

@api_bp.route('/webhooks/test', methods=['POST'])
@require_auth
@require_permission('admin')
def webhooks_test():
    try:
        sample = {'message': 'Webhook de prueba', 'ts': datetime.utcnow().isoformat()}
        results = webhook_publisher.publish('test.event', sample)
        return jsonify({'results': results})
    except Exception as e:
        logger.error(f'Error webhooks test: {str(e)}')
        return jsonify({'error': 'Error publicando webhook'}), 500

@api_bp.route('/notifications/telegram/test', methods=['POST'])
@require_auth
@require_permission('admin')
def telegram_test():
    try:
        ok = telegram_notifier.send_message('✅ Prueba Telegram OK - Inventario')
        return jsonify({'success': ok, 'configured': telegram_notifier.is_configured()})
    except Exception as e:
        logger.error(f'Error Telegram test: {str(e)}')
        return jsonify({'error': 'Error enviando a Telegram'}), 500

# ==================== HEALTH CHECK ====================

@api_bp.route('/health', methods=['GET'])
def health_check():
    """Verificación de salud del sistema"""
    try:
        # Estado base
        status = {
            'status': 'ok',
            'timestamp': datetime.utcnow().isoformat(),
            'services': {}
        }

        # Base de datos
        try:
            db.session.execute('SELECT 1')
            status['services']['database'] = 'ok'
        except Exception as e:
            status['services']['database'] = f'error: {str(e)}'
            status['status'] = 'degraded'

        # Notificaciones
        try:
            from utils.notifications import get_notification_service
            status['services']['notifications'] = 'ok' if get_notification_service() is not None else 'disabled'
        except Exception as e:
            status['services']['notifications'] = f'error: {str(e)}'
            status['status'] = 'degraded'

        # Version (si existe en config)
        try:
            from flask import current_app
            status['version'] = current_app.config.get('APP_VERSION', 'unknown')
        except Exception:
            status['version'] = 'unknown'

        code = 200 if status['status'] == 'ok' else 503
        return jsonify(status), code
    except Exception as e:
        logger.error(f'Error en health check: {str(e)}')
        return jsonify({'status': 'error', 'message': 'Health check failed'}), 500

# ==================== OPENAPI / SWAGGER ====================

def _get_openapi_spec():
    base_url = 'http://localhost:5000'
    spec = {
        'openapi': '3.0.3',
        'info': {
            'title': 'Inventario API',
            'version': '1.0.0'
        },
        'servers': [{'url': base_url + '/api'}],
        'paths': {
            '/auth/login': {
                'post': {
                    'summary': 'Login',
                    'requestBody': {
                        'required': True,
                        'content': {
                            'application/json': {
                                'schema': {
                                    'type': 'object',
                                    'properties': {
                                        'username': {'type': 'string'},
                                        'password': {'type': 'string'}
                                    },
                                    'required': ['username', 'password']
                                }
                            }
                        }
                    },
                    'responses': {
                        '200': {'description': 'OK'},
                        '401': {'description': 'Unauthorized'}
                    }
                }
            },
            '/products': {
                'get': {
                    'summary': 'Listar productos',
                    'parameters': [
                        {'in': 'query', 'name': 'page', 'schema': {'type': 'integer'}},
                        {'in': 'query', 'name': 'per_page', 'schema': {'type': 'integer'}},
                        {'in': 'query', 'name': 'q', 'schema': {'type': 'string'}},
                        {'in': 'query', 'name': 'sku', 'schema': {'type': 'string'}},
                        {'in': 'query', 'name': 'category', 'schema': {'type': 'string'}},
                        {'in': 'query', 'name': 'min_price', 'schema': {'type': 'number'}},
                        {'in': 'query', 'name': 'max_price', 'schema': {'type': 'number'}}
                    ],
                    'responses': {'200': {'description': 'OK'}}
                }
            },
            '/sales': {
                'get': {
                    'summary': 'Listar ventas',
                    'parameters': [
                        {'in': 'query', 'name': 'days', 'schema': {'type': 'integer'}},
                        {'in': 'query', 'name': 'start_date', 'schema': {'type': 'string', 'format': 'date'}},
                        {'in': 'query', 'name': 'end_date', 'schema': {'type': 'string', 'format': 'date'}},
                        {'in': 'query', 'name': 'customer_id', 'schema': {'type': 'integer'}},
                        {'in': 'query', 'name': 'page', 'schema': {'type': 'integer'}},
                        {'in': 'query', 'name': 'per_page', 'schema': {'type': 'integer'}}
                    ],
                    'responses': {'200': {'description': 'OK'}}
                }
            },
            '/health': {
                'get': {
                    'summary': 'Health check',
                    'responses': {'200': {'description': 'OK'}}
                }
            }
        },
        'components': {
            'securitySchemes': {
                'bearerAuth': {
                    'type': 'http',
                    'scheme': 'bearer',
                    'bearerFormat': 'JWT'
                }
            }
        },
        'security': [{'bearerAuth': []}]
    }
    return spec

@api_bp.route('/openapi.json', methods=['GET'])
def openapi_spec():
    return jsonify(_get_openapi_spec())

@api_bp.route('/docs', methods=['GET'])
def swagger_ui():
    html = """
<!doctype html>
<html>
  <head>
    <meta charset="utf-8" />
    <title>Inventario API Docs</title>
    <link rel="stylesheet" href="https://unpkg.com/swagger-ui-dist@5/swagger-ui.css" />
  </head>
  <body>
    <div id="swagger-ui"></div>
    <script src="https://unpkg.com/swagger-ui-dist@5/swagger-ui-bundle.js"></script>
    <script>
      window.ui = SwaggerUIBundle({
        url: '/api/openapi.json',
        dom_id: '#swagger-ui',
        presets: [SwaggerUIBundle.presets.apis],
        layout: 'BaseLayout'
      });
    </script>
  </body>
</html>
"""
    from flask import Response
    return Response(html, mimetype='text/html')
