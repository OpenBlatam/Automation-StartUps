#!/usr/bin/env python3
"""
Social Media Sentiment Analysis System - Enhanced Version
Analyzes social media discussions, customer reviews, and online forums
Tracks sentiment trends over 6 months and identifies emerging preferences
Features: API integrations, advanced NLP, visualizations, competitor analysis
"""

import json
import sqlite3
import re
import csv
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
import statistics
import math
import os
from pathlib import Path

# Optional imports for advanced features
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('sentiment_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class SentimentDataPoint:
    """Data point for sentiment analysis"""
    date: str
    sentiment_score: float
    sentiment_label: str
    source: str
    platform: str
    content: str
    emotions: Dict[str, float]
    topics: List[str]
    engagement: int = 0
    author_type: str = "user"  # user, influencer, expert, etc.

class SocialMediaSentimentAnalyzer:
    """Comprehensive social media sentiment analyzer with 6-month trend tracking"""
    
    def __init__(self, db_path: str = "social_sentiment_analysis.db", 
                 enable_api_integrations: bool = False,
                 api_credentials: Optional[Dict] = None):
        self.db_path = db_path
        self.enable_api_integrations = enable_api_integrations
        self.api_credentials = api_credentials or {}
        self.sentiment_lexicons = self._load_sentiment_lexicons()
        self.emotion_models = self._load_emotion_models()
        self.topic_keywords = self._load_topic_keywords()
        self.competitor_keywords = {}
        self.init_database()
        self._setup_output_directories()
        
        if enable_api_integrations:
            self._initialize_api_clients()
    
    def _setup_output_directories(self):
        """Create output directories for reports and visualizations"""
        self.output_dir = Path("sentiment_analysis_output")
        self.output_dir.mkdir(exist_ok=True)
        (self.output_dir / "reports").mkdir(exist_ok=True)
        (self.output_dir / "visualizations").mkdir(exist_ok=True)
        (self.output_dir / "exports").mkdir(exist_ok=True)
    
    def _initialize_api_clients(self):
        """Initialize API clients for social media platforms"""
        self.api_clients = {}
        
        # Twitter API (if credentials provided)
        if 'twitter' in self.api_credentials:
            try:
                import tweepy
                auth = tweepy.OAuthHandler(
                    self.api_credentials['twitter'].get('consumer_key'),
                    self.api_credentials['twitter'].get('consumer_secret')
                )
                auth.set_access_token(
                    self.api_credentials['twitter'].get('access_token'),
                    self.api_credentials['twitter'].get('access_token_secret')
                )
                self.api_clients['twitter'] = tweepy.API(auth, wait_on_rate_limit=True)
                logger.info("Twitter API client initialized")
            except Exception as e:
                logger.warning(f"Could not initialize Twitter API: {e}")
        
        # Reddit API (if credentials provided)
        if 'reddit' in self.api_credentials:
            try:
                import praw
                self.api_clients['reddit'] = praw.Reddit(
                    client_id=self.api_credentials['reddit'].get('client_id'),
                    client_secret=self.api_credentials['reddit'].get('client_secret'),
                    user_agent=self.api_credentials['reddit'].get('user_agent', 'SentimentAnalyzer/1.0')
                )
                logger.info("Reddit API client initialized")
            except Exception as e:
                logger.warning(f"Could not initialize Reddit API: {e}")
    
    def init_database(self):
        """Initialize database for storing sentiment data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Main sentiment data table with additional fields
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sentiment_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                sentiment_score REAL,
                sentiment_label TEXT,
                source TEXT,
                platform TEXT,
                content TEXT,
                emotions TEXT,
                topics TEXT,
                engagement INTEGER,
                author_type TEXT,
                product_service TEXT,
                location TEXT,
                language TEXT,
                author_id TEXT,
                author_followers INTEGER,
                competitor_mention TEXT,
                created_at TEXT,
                INDEX idx_date (date),
                INDEX idx_platform (platform),
                INDEX idx_sentiment (sentiment_label),
                INDEX idx_product (product_service)
            )
        ''')
        
        # Monthly aggregations
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS monthly_sentiment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER,
                month INTEGER,
                avg_sentiment REAL,
                positive_count INTEGER,
                negative_count INTEGER,
                neutral_count INTEGER,
                total_count INTEGER,
                top_emotions TEXT,
                top_topics TEXT,
                trend_direction TEXT,
                created_at TEXT
            )
        ''')
        
        # Trend analysis
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sentiment_trends (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trend_type TEXT,
                trend_data TEXT,
                time_period TEXT,
                significance REAL,
                created_at TEXT
            )
        ''')
        
        # Platform-specific metrics
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS platform_metrics (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                platform TEXT,
                date TEXT,
                avg_sentiment REAL,
                volume INTEGER,
                engagement_rate REAL,
                top_influencers TEXT,
                created_at TEXT
            )
        ''')
        
        # Competitor analysis
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS competitor_analysis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                competitor_name TEXT,
                date TEXT,
                mention_count INTEGER,
                avg_sentiment REAL,
                comparison_score REAL,
                created_at TEXT
            )
        ''')
        
        # Influencer tracking
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS influencer_mentions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                author_id TEXT,
                author_name TEXT,
                platform TEXT,
                followers_count INTEGER,
                mention_count INTEGER,
                avg_sentiment REAL,
                influence_score REAL,
                created_at TEXT
            )
        ''')
        
        # Alerts and anomalies
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sentiment_alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                alert_type TEXT,
                severity TEXT,
                message TEXT,
                data TEXT,
                created_at TEXT,
                resolved INTEGER DEFAULT 0
            )
        ''')
        
        conn.commit()
        conn.close()
        logger.info("Database initialized successfully")
    
    def _load_sentiment_lexicons(self) -> Dict:
        """Load sentiment lexicons for analysis"""
        return {
            'positive_words': {
                'excellent': 0.9, 'amazing': 0.9, 'fantastic': 0.9, 'wonderful': 0.8,
                'great': 0.8, 'good': 0.7, 'nice': 0.6, 'awesome': 0.9, 'brilliant': 0.8,
                'outstanding': 0.9, 'perfect': 1.0, 'love': 0.9, 'like': 0.7, 'enjoy': 0.8,
                'happy': 0.8, 'pleased': 0.7, 'satisfied': 0.7, 'impressed': 0.8,
                'recommend': 0.8, 'superb': 0.9, 'marvelous': 0.9, 'exceptional': 0.9,
                'delighted': 0.8, 'thrilled': 0.9, 'ecstatic': 0.95, 'phenomenal': 0.9,
                'incredible': 0.9, 'remarkable': 0.85, 'splendid': 0.85, 'magnificent': 0.9
            },
            'negative_words': {
                'terrible': -0.9, 'awful': -0.9, 'horrible': -0.9, 'bad': -0.7,
                'worst': -0.9, 'hate': -0.9, 'dislike': -0.6, 'disappointed': -0.7,
                'angry': -0.8, 'frustrated': -0.7, 'annoyed': -0.6, 'upset': -0.7,
                'sad': -0.6, 'unhappy': -0.7, 'disgusted': -0.8, 'furious': -0.9,
                'disgusting': -0.8, 'pathetic': -0.8, 'useless': -0.7, 'waste': -0.7,
                'broken': -0.7, 'defective': -0.8, 'poor': -0.6, 'inferior': -0.7,
                'unacceptable': -0.8, 'ridiculous': -0.7, 'absurd': -0.7, 'outrageous': -0.8
            },
            'intensifiers': {
                'very': 1.5, 'extremely': 2.0, 'incredibly': 2.0, 'absolutely': 1.8,
                'totally': 1.5, 'completely': 1.5, 'really': 1.3, 'quite': 1.2,
                'somewhat': 0.8, 'slightly': 0.7, 'barely': 0.5, 'hardly': 0.3,
                'super': 1.4, 'ultra': 1.6, 'mega': 1.5, 'insanely': 2.1
            },
            'negators': {
                'not': -1.0, 'no': -1.0, 'never': -1.0, 'none': -1.0,
                'nothing': -1.0, 'nobody': -1.0, 'nowhere': -1.0, 'neither': -1.0,
                'without': -0.8, 'lack': -0.7, 'missing': -0.6
            }
        }
    
    def _load_emotion_models(self) -> Dict:
        """Load emotion detection models"""
        return {
            'joy': {
                'keywords': ['happy', 'joy', 'excited', 'thrilled', 'delighted', 'cheerful', 
                           'elated', 'ecstatic', 'jubilant', 'blissful', 'euphoric'],
                'weight': 1.0
            },
            'sadness': {
                'keywords': ['sad', 'depressed', 'melancholy', 'gloomy', 'sorrowful', 
                            'dejected', 'miserable', 'heartbroken', 'disheartened'],
                'weight': 1.0
            },
            'anger': {
                'keywords': ['angry', 'mad', 'furious', 'rage', 'irritated', 'annoyed', 
                           'outraged', 'livid', 'enraged', 'incensed'],
                'weight': 1.0
            },
            'fear': {
                'keywords': ['afraid', 'scared', 'terrified', 'worried', 'anxious', 
                           'nervous', 'frightened', 'apprehensive', 'panicked'],
                'weight': 1.0
            },
            'surprise': {
                'keywords': ['surprised', 'amazed', 'astonished', 'shocked', 'stunned', 
                           'bewildered', 'astounded', 'flabbergasted'],
                'weight': 1.0
            },
            'disgust': {
                'keywords': ['disgusted', 'revolted', 'repulsed', 'sickened', 'appalled', 
                           'horrified', 'nauseated'],
                'weight': 1.0
            },
            'trust': {
                'keywords': ['trust', 'reliable', 'credible', 'honest', 'authentic', 
                           'genuine', 'transparent', 'dependable', 'trustworthy'],
                'weight': 1.2
            },
            'anticipation': {
                'keywords': ['excited', 'eager', 'hopeful', 'optimistic', 'anticipating', 
                           'expecting', 'looking forward'],
                'weight': 1.1
            }
        }
    
    def _load_topic_keywords(self) -> Dict:
        """Load topic keywords for categorization"""
        return {
            'product_quality': ['quality', 'durable', 'reliable', 'sturdy', 'well-made', 
                              'premium', 'cheap', 'flimsy', 'break', 'defect'],
            'customer_service': ['service', 'support', 'help', 'response', 'staff', 
                               'representative', 'assistant', 'care', 'attention'],
            'pricing': ['price', 'cost', 'expensive', 'affordable', 'value', 'worth', 
                       'cheap', 'overpriced', 'budget', 'deal'],
            'features': ['feature', 'function', 'capability', 'option', 'setting', 
                        'tool', 'functionality', 'ability'],
            'usability': ['easy', 'simple', 'intuitive', 'complicated', 'difficult', 
                         'confusing', 'user-friendly', 'straightforward'],
            'shipping': ['delivery', 'shipping', 'arrived', 'package', 'fast', 'slow', 
                        'tracking', 'delayed'],
            'design': ['design', 'look', 'appearance', 'style', 'aesthetic', 'beautiful', 
                      'ugly', 'modern', 'attractive'],
            'performance': ['fast', 'slow', 'speed', 'performance', 'efficient', 
                          'laggy', 'responsive', 'smooth']
        }
    
    def analyze_sentiment(self, text: str, source: str = "unknown", 
                         platform: str = "unknown", product_service: str = "",
                         location: Optional[str] = None,
                         author_id: Optional[str] = None,
                         author_followers: int = 0) -> Dict:
        """Analyze sentiment of a single text with enhanced features"""
        if not text or len(text.strip()) < 3:
            return {
                'sentiment_score': 0.0,
                'sentiment_label': 'neutral',
                'confidence': 0.0,
                'emotions': {},
                'topics': [],
                'error': 'Text too short'
            }
        
        processed_text = self._preprocess_text(text)
        sentiment_score = self._calculate_sentiment_score(processed_text)
        sentiment_label = self._get_sentiment_label(sentiment_score)
        confidence = self._calculate_confidence(processed_text, sentiment_score)
        emotions = self._analyze_emotions(processed_text)
        topics = self._extract_topics(processed_text)
        language = self._detect_language(text)
        competitor_mention = self._detect_competitor_mention(text)
        
        # Calculate influence score if author info provided
        influence_score = self._calculate_influence_score(author_followers, platform) if author_followers else 0
        
        return {
            'sentiment_score': round(sentiment_score, 4),
            'sentiment_label': sentiment_label,
            'confidence': round(confidence, 4),
            'emotions': emotions,
            'topics': topics,
            'source': source,
            'platform': platform,
            'product_service': product_service,
            'language': language,
            'competitor_mention': competitor_mention,
            'influence_score': round(influence_score, 4),
            'text_length': len(text),
            'word_count': len(processed_text.split())
        }
    
    def _preprocess_text(self, text: str) -> str:
        """Preprocess text for analysis"""
        text = text.lower()
        # Remove URLs
        text = re.sub(r'http\S+|www\.\S+', '', text)
        # Remove mentions and hashtags (keep content)
        text = re.sub(r'@\w+|#\w+', '', text)
        # Remove special characters but keep basic punctuation
        text = re.sub(r'[^\w\s\.\!\?\,\;\:]', '', text)
        # Normalize whitespace
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    
    def _calculate_sentiment_score(self, text: str) -> float:
        """Calculate sentiment score from text"""
        words = text.split()
        total_score = 0.0
        word_count = 0
        
        i = 0
        while i < len(words):
            word = words[i]
            word_score = 0.0
            
            # Check sentiment words
            if word in self.sentiment_lexicons['positive_words']:
                word_score = self.sentiment_lexicons['positive_words'][word]
            elif word in self.sentiment_lexicons['negative_words']:
                word_score = self.sentiment_lexicons['negative_words'][word]
            
            # Check intensifiers
            if i > 0 and words[i-1] in self.sentiment_lexicons['intensifiers']:
                intensifier = self.sentiment_lexicons['intensifiers'][words[i-1]]
                word_score *= intensifier
            
            # Check negators
            if i > 0 and words[i-1] in self.sentiment_lexicons['negators']:
                word_score *= -1.0
            
            # Check double negation
            if i > 1 and (words[i-1] in self.sentiment_lexicons['negators'] and 
                         words[i-2] in self.sentiment_lexicons['negators']):
                word_score *= -1.0
            
            total_score += word_score
            if word_score != 0:
                word_count += 1
            
            i += 1
        
        # Normalize score
        if word_count > 0:
            normalized_score = total_score / word_count
            return max(-1.0, min(1.0, normalized_score))
        
        return 0.0
    
    def _get_sentiment_label(self, score: float) -> str:
        """Get sentiment label from score"""
        if score >= 0.3:
            return 'positive'
        elif score <= -0.3:
            return 'negative'
        else:
            return 'neutral'
    
    def _calculate_confidence(self, text: str, sentiment_score: float) -> float:
        """Calculate confidence in sentiment analysis"""
        words = text.split()
        sentiment_words = 0
        
        for word in words:
            if (word in self.sentiment_lexicons['positive_words'] or 
                word in self.sentiment_lexicons['negative_words']):
                sentiment_words += 1
        
        if len(words) > 0:
            density = sentiment_words / len(words)
            confidence = min(1.0, density * 2)
        else:
            confidence = 0.0
        
        magnitude = abs(sentiment_score)
        confidence = (confidence + magnitude) / 2
        
        return round(confidence, 3)
    
    def _analyze_emotions(self, text: str) -> Dict[str, float]:
        """Analyze emotions in text"""
        emotions = {}
        words = text.split()
        
        for emotion, data in self.emotion_models.items():
            emotion_score = 0.0
            for keyword in data['keywords']:
                count = words.count(keyword)
                emotion_score += count * data['weight']
            
            if emotion_score > 0:
                emotions[emotion] = round(emotion_score, 3)
        
        return emotions
    
    def _extract_topics(self, text: str) -> List[str]:
        """Extract topics from text"""
        topics = []
        words = text.split()
        text_lower = text.lower()
        
        for topic, keywords in self.topic_keywords.items():
            for keyword in keywords:
                if keyword in text_lower:
                    topics.append(topic)
                    break
        
        return topics
    
    def _detect_language(self, text: str) -> str:
        """Detect language of text"""
        # Simple language detection based on common words
        text_lower = text.lower()
        
        # Spanish indicators
        spanish_words = ['el', 'la', 'de', 'que', 'y', 'en', 'un', 'es', 'se', 'no', 'te', 'lo']
        spanish_count = sum(1 for word in spanish_words if word in text_lower)
        
        # English indicators
        english_words = ['the', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with']
        english_count = sum(1 for word in english_words if word in text_lower)
        
        if spanish_count > english_count:
            return 'es'
        elif english_count > spanish_count:
            return 'en'
        else:
            return 'unknown'
    
    def _detect_competitor_mention(self, text: str) -> Optional[str]:
        """Detect if text mentions a competitor"""
        text_lower = text.lower()
        for competitor, keywords in self.competitor_keywords.items():
            for keyword in keywords:
                if keyword in text_lower:
                    return competitor
        return None
    
    def _calculate_influence_score(self, followers: int, platform: str) -> float:
        """Calculate influence score based on follower count and platform"""
        # Platform multipliers
        platform_multipliers = {
            'twitter': 1.0,
            'instagram': 1.2,
            'youtube': 1.5,
            'tiktok': 1.3,
            'facebook': 0.8,
            'linkedin': 1.1,
            'reddit': 0.5
        }
        
        multiplier = platform_multipliers.get(platform.lower(), 1.0)
        
        # Logarithmic scale for followers
        if followers == 0:
            return 0.0
        
        # Normalize to 0-1 scale using log
        log_followers = math.log10(followers + 1)
        normalized = min(1.0, log_followers / 7)  # 10M followers = ~7 on log scale
        
        return normalized * multiplier
    
    def set_competitor_keywords(self, competitors: Dict[str, List[str]]):
        """Set competitor keywords for detection"""
        self.competitor_keywords = competitors
        logger.info(f"Competitor keywords updated: {list(competitors.keys())}")
    
    def save_sentiment_data(self, data_point: SentimentDataPoint, product_service: str = "",
                           location: Optional[str] = None, language: Optional[str] = None,
                           author_id: Optional[str] = None, author_followers: int = 0,
                           competitor_mention: Optional[str] = None):
        """Save sentiment data point to database with enhanced fields"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO sentiment_data 
            (date, sentiment_score, sentiment_label, source, platform, content, 
             emotions, topics, engagement, author_type, product_service, location,
             language, author_id, author_followers, competitor_mention, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data_point.date,
            data_point.sentiment_score,
            data_point.sentiment_label,
            data_point.source,
            data_point.platform,
            data_point.content,
            json.dumps(data_point.emotions),
            json.dumps(data_point.topics),
            data_point.engagement,
            data_point.author_type,
            product_service,
            location,
            language,
            author_id,
            author_followers,
            competitor_mention,
            datetime.now().isoformat()
        ))
        
        conn.commit()
        conn.close()
        
        # Check for alerts
        self._check_alerts(data_point)
    
    def analyze_6_month_trends(self, product_service: str = "") -> Dict:
        """Analyze sentiment trends over the past 6 months"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, sentiment_score, sentiment_label, emotions, topics, platform, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' ORDER BY date'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data found for the 6-month period'}
        
        # Organize by month
        monthly_data = defaultdict(lambda: {
            'scores': [],
            'labels': [],
            'emotions': defaultdict(list),
            'topics': defaultdict(int),
            'platforms': defaultdict(int),
            'total_engagement': 0
        })
        
        for row in data:
            date_str = row[0]
            date_obj = datetime.fromisoformat(date_str)
            month_key = f"{date_obj.year}-{date_obj.month:02d}"
            
            monthly_data[month_key]['scores'].append(row[1])
            monthly_data[month_key]['labels'].append(row[2])
            
            emotions = json.loads(row[3]) if row[3] else {}
            for emotion, score in emotions.items():
                monthly_data[month_key]['emotions'][emotion].append(score)
            
            topics = json.loads(row[4]) if row[4] else []
            for topic in topics:
                monthly_data[month_key]['topics'][topic] += 1
            
            monthly_data[month_key]['platforms'][row[5]] += 1
            monthly_data[month_key]['total_engagement'] += row[6] or 0
        
        # Calculate monthly metrics
        monthly_metrics = {}
        for month, data in monthly_data.items():
            avg_sentiment = statistics.mean(data['scores']) if data['scores'] else 0
            label_counts = Counter(data['labels'])
            
            # Top emotions
            top_emotions = {}
            for emotion, scores in data['emotions'].items():
                top_emotions[emotion] = statistics.mean(scores)
            top_emotions = dict(sorted(top_emotions.items(), key=lambda x: x[1], reverse=True)[:5])
            
            # Top topics
            top_topics = dict(sorted(data['topics'].items(), key=lambda x: x[1], reverse=True)[:5])
            
            monthly_metrics[month] = {
                'avg_sentiment': round(avg_sentiment, 3),
                'positive_count': label_counts.get('positive', 0),
                'negative_count': label_counts.get('negative', 0),
                'neutral_count': label_counts.get('neutral', 0),
                'total_count': len(data['scores']),
                'top_emotions': top_emotions,
                'top_topics': top_topics,
                'platform_distribution': dict(data['platforms']),
                'total_engagement': data['total_engagement']
            }
        
        # Calculate overall trends
        months = sorted(monthly_metrics.keys())
        if len(months) >= 2:
            first_month_avg = monthly_metrics[months[0]]['avg_sentiment']
            last_month_avg = monthly_metrics[months[-1]]['avg_sentiment']
            trend_direction = 'improving' if last_month_avg > first_month_avg else 'declining'
            trend_magnitude = abs(last_month_avg - first_month_avg)
        else:
            trend_direction = 'stable'
            trend_magnitude = 0.0
        
        # Identify emerging preferences
        emerging_preferences = self._identify_emerging_preferences(monthly_data)
        
        # Identify dissatisfaction trends
        dissatisfaction_trends = self._identify_dissatisfaction_trends(monthly_data)
        
        # Calculate sentiment shifts
        sentiment_shifts = self._calculate_sentiment_shifts(monthly_metrics)
        
        return {
            'success': True,
            'period': '6 months',
            'start_date': six_months_ago.isoformat(),
            'end_date': datetime.now().isoformat(),
            'monthly_metrics': monthly_metrics,
            'overall_trend': {
                'direction': trend_direction,
                'magnitude': round(trend_magnitude, 3),
                'first_month_avg': round(first_month_avg, 3) if len(months) >= 2 else 0,
                'last_month_avg': round(last_month_avg, 3) if len(months) >= 2 else 0
            },
            'emerging_preferences': emerging_preferences,
            'dissatisfaction_trends': dissatisfaction_trends,
            'sentiment_shifts': sentiment_shifts,
            'top_insights': self._generate_top_insights(monthly_metrics, emerging_preferences, 
                                                       dissatisfaction_trends, sentiment_shifts)
        }
    
    def _identify_emerging_preferences(self, monthly_data: Dict) -> List[Dict]:
        """Identify emerging preferences over time"""
        preferences = []
        
        # Track topic mentions over time
        topic_trends = defaultdict(list)
        for month, data in monthly_data.items():
            for topic, count in data['topics'].items():
                topic_trends[topic].append((month, count))
        
        # Find topics with increasing mentions
        for topic, trend_data in topic_trends.items():
            if len(trend_data) >= 3:
                recent_avg = statistics.mean([count for _, count in trend_data[-2:]])
                earlier_avg = statistics.mean([count for _, count in trend_data[:2]])
                
                if recent_avg > earlier_avg * 1.5:  # 50% increase
                    preferences.append({
                        'topic': topic,
                        'trend': 'increasing',
                        'growth_rate': round((recent_avg / earlier_avg - 1) * 100, 1),
                        'description': f"Growing interest in {topic.replace('_', ' ')}"
                    })
        
        return sorted(preferences, key=lambda x: x['growth_rate'], reverse=True)
    
    def _identify_dissatisfaction_trends(self, monthly_data: Dict) -> List[Dict]:
        """Identify dissatisfaction trends"""
        trends = []
        
        # Track negative sentiment over time
        negative_trends = []
        for month in sorted(monthly_data.keys()):
            data = monthly_data[month]
            negative_pct = (data['labels'].count('negative') / len(data['labels']) * 100) if data['labels'] else 0
            negative_trends.append((month, negative_pct))
        
        if len(negative_trends) >= 3:
            recent_avg = statistics.mean([pct for _, pct in negative_trends[-2:]])
            earlier_avg = statistics.mean([pct for _, pct in negative_trends[:2]])
            
            if recent_avg > earlier_avg + 5:  # 5% increase in negative sentiment
                trends.append({
                    'type': 'increasing_dissatisfaction',
                    'current_level': round(recent_avg, 1),
                    'previous_level': round(earlier_avg, 1),
                    'increase': round(recent_avg - earlier_avg, 1),
                    'severity': 'high' if recent_avg > 30 else 'moderate' if recent_avg > 20 else 'low'
                })
        
        # Identify recurring negative topics
        negative_topics = defaultdict(int)
        for month, data in monthly_data.items():
            negative_count = data['labels'].count('negative')
            if negative_count > 0:
                for topic, count in data['topics'].items():
                    negative_topics[topic] += count
        
        top_negative_topics = sorted(negative_topics.items(), key=lambda x: x[1], reverse=True)[:3]
        for topic, count in top_negative_topics:
            trends.append({
                'type': 'recurring_issue',
                'topic': topic.replace('_', ' '),
                'frequency': count,
                'description': f"Frequent complaints about {topic.replace('_', ' ')}"
            })
        
        return trends
    
    def _calculate_sentiment_shifts(self, monthly_metrics: Dict) -> List[Dict]:
        """Calculate significant sentiment shifts"""
        shifts = []
        months = sorted(monthly_metrics.keys())
        
        if len(months) < 2:
            return shifts
        
        for i in range(1, len(months)):
            prev_month = months[i-1]
            curr_month = months[i]
            
            prev_avg = monthly_metrics[prev_month]['avg_sentiment']
            curr_avg = monthly_metrics[curr_month]['avg_sentiment']
            
            change = curr_avg - prev_avg
            
            if abs(change) > 0.15:  # Significant shift
                shifts.append({
                    'period': f"{prev_month} to {curr_month}",
                    'change': round(change, 3),
                    'direction': 'positive' if change > 0 else 'negative',
                    'magnitude': 'large' if abs(change) > 0.3 else 'moderate',
                    'previous_sentiment': round(prev_avg, 3),
                    'current_sentiment': round(curr_avg, 3)
                })
        
        return shifts
    
    def _generate_top_insights(self, monthly_metrics: Dict, emerging_preferences: List[Dict],
                               dissatisfaction_trends: List[Dict], sentiment_shifts: List[Dict]) -> List[str]:
        """Generate top insights summary"""
        insights = []
        
        # Overall sentiment trend
        months = sorted(monthly_metrics.keys())
        if len(months) >= 2:
            first_avg = monthly_metrics[months[0]]['avg_sentiment']
            last_avg = monthly_metrics[months[-1]]['avg_sentiment']
            
            if last_avg > first_avg + 0.2:
                insights.append(f"Sentiment has significantly improved over 6 months (+{round(last_avg - first_avg, 2)})")
            elif last_avg < first_avg - 0.2:
                insights.append(f"Sentiment has declined over 6 months ({round(last_avg - first_avg, 2)})")
            else:
                insights.append("Sentiment has remained relatively stable over the 6-month period")
        
        # Emerging preferences
        if emerging_preferences:
            top_preference = emerging_preferences[0]
            insights.append(f"Emerging preference: {top_preference['description']} (growing {top_preference['growth_rate']}%)")
        
        # Dissatisfaction trends
        if dissatisfaction_trends:
            for trend in dissatisfaction_trends[:2]:
                if trend['type'] == 'increasing_dissatisfaction':
                    insights.append(f"Dissatisfaction is increasing: {trend['current_level']}% negative sentiment (up from {trend['previous_level']}%)")
                elif trend['type'] == 'recurring_issue':
                    insights.append(f"Recurring issue: {trend['description']}")
        
        # Sentiment shifts
        if sentiment_shifts:
            largest_shift = max(sentiment_shifts, key=lambda x: abs(x['change']))
            insights.append(f"Largest sentiment shift: {largest_shift['direction']} change of {abs(largest_shift['change'])} in {largest_shift['period']}")
        
        # Platform insights
        if months:
            last_month = months[-1]
            platform_dist = monthly_metrics[last_month].get('platform_distribution', {})
            if platform_dist:
                top_platform = max(platform_dist.items(), key=lambda x: x[1])
                insights.append(f"Most active platform: {top_platform[0]} with {top_platform[1]} mentions")
        
        return insights
    
    def generate_summary_report(self, product_service: str = "") -> str:
        """Generate a comprehensive summary report"""
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            return f"Error: {trends.get('error', 'Unknown error')}"
        
        report = []
        report.append("=" * 80)
        report.append("SOCIAL MEDIA SENTIMENT ANALYSIS REPORT")
        report.append("6-Month Trend Analysis")
        report.append("=" * 80)
        report.append("")
        
        # Executive Summary
        report.append("EXECUTIVE SUMMARY")
        report.append("-" * 80)
        report.append(f"Analysis Period: {trends['start_date']} to {trends['end_date']}")
        report.append(f"Overall Trend: {trends['overall_trend']['direction'].upper()}")
        report.append(f"Sentiment Change: {trends['overall_trend']['first_month_avg']} → {trends['overall_trend']['last_month_avg']}")
        report.append("")
        
        # Top Insights
        report.append("TOP INSIGHTS")
        report.append("-" * 80)
        for i, insight in enumerate(trends['top_insights'], 1):
            report.append(f"{i}. {insight}")
        report.append("")
        
        # Monthly Breakdown
        report.append("MONTHLY BREAKDOWN")
        report.append("-" * 80)
        for month in sorted(trends['monthly_metrics'].keys()):
            metrics = trends['monthly_metrics'][month]
            report.append(f"\n{month}:")
            report.append(f"  Average Sentiment: {metrics['avg_sentiment']}")
            report.append(f"  Total Mentions: {metrics['total_count']}")
            report.append(f"  Positive: {metrics['positive_count']} ({metrics['positive_count']/metrics['total_count']*100:.1f}%)")
            report.append(f"  Negative: {metrics['negative_count']} ({metrics['negative_count']/metrics['total_count']*100:.1f}%)")
            report.append(f"  Top Topics: {', '.join(list(metrics['top_topics'].keys())[:3])}")
        
        report.append("")
        
        # Emerging Preferences
        if trends['emerging_preferences']:
            report.append("EMERGING PREFERENCES")
            report.append("-" * 80)
            for pref in trends['emerging_preferences']:
                report.append(f"• {pref['description']} (Growth: {pref['growth_rate']}%)")
            report.append("")
        
        # Dissatisfaction Trends
        if trends['dissatisfaction_trends']:
            report.append("DISSATISFACTION TRENDS")
            report.append("-" * 80)
            for trend in trends['dissatisfaction_trends']:
                if trend['type'] == 'increasing_dissatisfaction':
                    report.append(f"• Increasing dissatisfaction: {trend['current_level']}% (up {trend['increase']}%)")
                elif trend['type'] == 'recurring_issue':
                    report.append(f"• {trend['description']}")
            report.append("")
        
        # Sentiment Shifts
        if trends['sentiment_shifts']:
            report.append("SIGNIFICANT SENTIMENT SHIFTS")
            report.append("-" * 80)
            for shift in trends['sentiment_shifts']:
                report.append(f"• {shift['period']}: {shift['direction']} shift of {abs(shift['change'])}")
            report.append("")
        
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def _check_alerts(self, data_point: SentimentDataPoint):
        """Check for alert conditions and create alerts"""
        alerts = []
        
        # Negative sentiment spike alert
        if data_point.sentiment_score < -0.7:
            alerts.append({
                'type': 'high_negative_sentiment',
                'severity': 'high',
                'message': f'Very negative sentiment detected: {data_point.sentiment_score:.2f}',
                'data': json.dumps({
                    'platform': data_point.platform,
                    'content_preview': data_point.content[:100]
                })
            })
        
        # High engagement negative content
        if data_point.sentiment_label == 'negative' and data_point.engagement > 1000:
            alerts.append({
                'type': 'viral_negative_content',
                'severity': 'critical',
                'message': f'High engagement negative content detected: {data_point.engagement} engagements',
                'data': json.dumps({
                    'platform': data_point.platform,
                    'engagement': data_point.engagement
                })
            })
        
        # Save alerts
        if alerts:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            for alert in alerts:
                cursor.execute('''
                    INSERT INTO sentiment_alerts 
                    (alert_type, severity, message, data, created_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (
                    alert['type'],
                    alert['severity'],
                    alert['message'],
                    alert['data'],
                    datetime.now().isoformat()
                ))
            conn.commit()
            conn.close()
            logger.warning(f"Created {len(alerts)} alerts")
    
    def fetch_twitter_data(self, query: str, count: int = 100, 
                          product_service: str = "") -> List[Dict]:
        """Fetch data from Twitter API"""
        if 'twitter' not in self.api_clients:
            logger.warning("Twitter API client not initialized")
            return []
        
        try:
            import tweepy
            tweets = []
            api = self.api_clients['twitter']
            
            for tweet in tweepy.Cursor(api.search_tweets,
                                     q=query,
                                     lang="en",
                                     tweet_mode="extended",
                                     result_type="recent").items(count):
                analysis = self.analyze_sentiment(
                    tweet.full_text,
                    source=f"twitter_{tweet.id}",
                    platform="twitter",
                    product_service=product_service,
                    author_id=str(tweet.user.id),
                    author_followers=tweet.user.followers_count
                )
                
                data_point = SentimentDataPoint(
                    date=tweet.created_at.isoformat(),
                    sentiment_score=analysis['sentiment_score'],
                    sentiment_label=analysis['sentiment_label'],
                    source=f"twitter_{tweet.id}",
                    platform="twitter",
                    content=tweet.full_text,
                    emotions=analysis['emotions'],
                    topics=analysis['topics'],
                    engagement=tweet.retweet_count + tweet.favorite_count,
                    author_type="influencer" if tweet.user.followers_count > 10000 else "user"
                )
                
                self.save_sentiment_data(
                    data_point,
                    product_service=product_service,
                    author_id=str(tweet.user.id),
                    author_followers=tweet.user.followers_count
                )
                
                tweets.append({
                    'id': tweet.id,
                    'text': tweet.full_text,
                    'sentiment': analysis,
                    'engagement': tweet.retweet_count + tweet.favorite_count
                })
            
            logger.info(f"Fetched {len(tweets)} tweets")
            return tweets
            
        except Exception as e:
            logger.error(f"Error fetching Twitter data: {e}")
            return []
    
    def fetch_reddit_data(self, subreddit: str, limit: int = 100,
                          product_service: str = "") -> List[Dict]:
        """Fetch data from Reddit API"""
        if 'reddit' not in self.api_clients:
            logger.warning("Reddit API client not initialized")
            return []
        
        try:
            posts = []
            reddit = self.api_clients['reddit']
            subreddit_obj = reddit.subreddit(subreddit)
            
            for submission in subreddit_obj.hot(limit=limit):
                # Analyze post
                analysis = self.analyze_sentiment(
                    submission.title + " " + (submission.selftext or ""),
                    source=f"reddit_{submission.id}",
                    platform="reddit",
                    product_service=product_service
                )
                
                data_point = SentimentDataPoint(
                    date=datetime.fromtimestamp(submission.created_utc).isoformat(),
                    sentiment_score=analysis['sentiment_score'],
                    sentiment_label=analysis['sentiment_label'],
                    source=f"reddit_{submission.id}",
                    platform="reddit",
                    content=submission.title + " " + (submission.selftext or ""),
                    emotions=analysis['emotions'],
                    topics=analysis['topics'],
                    engagement=submission.score + submission.num_comments,
                    author_type="user"
                )
                
                self.save_sentiment_data(data_point, product_service=product_service)
                
                posts.append({
                    'id': submission.id,
                    'title': submission.title,
                    'sentiment': analysis,
                    'engagement': submission.score + submission.num_comments
                })
            
            logger.info(f"Fetched {len(posts)} Reddit posts")
            return posts
            
        except Exception as e:
            logger.error(f"Error fetching Reddit data: {e}")
            return []
    
    def export_to_csv(self, product_service: str = "", filename: Optional[str] = None):
        """Export sentiment data to CSV"""
        if not PANDAS_AVAILABLE:
            logger.warning("Pandas not available, using basic CSV export")
            return self._export_to_csv_basic(product_service, filename)
        
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        query = '''
            SELECT date, sentiment_score, sentiment_label, platform, content,
                   emotions, topics, engagement, product_service
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        if filename is None:
            filename = self.output_dir / "exports" / f"sentiment_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            filename = self.output_dir / "exports" / filename
        
        df.to_csv(filename, index=False)
        logger.info(f"Data exported to {filename}")
        return str(filename)
    
    def _export_to_csv_basic(self, product_service: str = "", filename: Optional[str] = None):
        """Basic CSV export without pandas"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, sentiment_score, sentiment_label, platform, content,
                   emotions, topics, engagement, product_service
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        conn.close()
        
        if filename is None:
            filename = self.output_dir / "exports" / f"sentiment_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            filename = self.output_dir / "exports" / filename
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(rows)
        
        logger.info(f"Data exported to {filename}")
        return str(filename)
    
    def export_to_json(self, product_service: str = "", filename: Optional[str] = None):
        """Export sentiment data to JSON"""
        trends = self.analyze_6_month_trends(product_service)
        
        if filename is None:
            filename = self.output_dir / "exports" / f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        else:
            filename = self.output_dir / "exports" / filename
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(trends, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Report exported to {filename}")
        return str(filename)
    
    def create_visualizations(self, product_service: str = ""):
        """Create visualization charts"""
        if not MATPLOTLIB_AVAILABLE:
            logger.warning("Matplotlib not available, skipping visualizations")
            return
        
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            logger.error("Cannot create visualizations: no data available")
            return
        
        # Set style
        plt.style.use('seaborn-v0_8-darkgrid')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle(f'Sentiment Analysis - {product_service or "All Products"}', 
                     fontsize=16, fontweight='bold')
        
        # 1. Monthly sentiment trend
        months = sorted(trends['monthly_metrics'].keys())
        avg_sentiments = [trends['monthly_metrics'][m]['avg_sentiment'] for m in months]
        
        axes[0, 0].plot(months, avg_sentiments, marker='o', linewidth=2, markersize=8)
        axes[0, 0].axhline(y=0, color='r', linestyle='--', alpha=0.5)
        axes[0, 0].set_title('Monthly Average Sentiment Trend', fontweight='bold')
        axes[0, 0].set_xlabel('Month')
        axes[0, 0].set_ylabel('Sentiment Score')
        axes[0, 0].grid(True, alpha=0.3)
        axes[0, 0].tick_params(axis='x', rotation=45)
        
        # 2. Sentiment distribution
        last_month = months[-1] if months else None
        if last_month:
            metrics = trends['monthly_metrics'][last_month]
            labels = ['Positive', 'Negative', 'Neutral']
            sizes = [
                metrics['positive_count'],
                metrics['negative_count'],
                metrics['neutral_count']
            ]
            colors = ['#2ecc71', '#e74c3c', '#95a5a6']
            
            axes[0, 1].pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
                          startangle=90, textprops={'fontsize': 10, 'fontweight': 'bold'})
            axes[0, 1].set_title(f'Sentiment Distribution - {last_month}', fontweight='bold')
        
        # 3. Platform comparison
        if last_month:
            platform_dist = trends['monthly_metrics'][last_month].get('platform_distribution', {})
            if platform_dist:
                platforms = list(platform_dist.keys())
                counts = list(platform_dist.values())
                
                axes[1, 0].bar(platforms, counts, color='#3498db', alpha=0.7)
                axes[1, 0].set_title('Mentions by Platform', fontweight='bold')
                axes[1, 0].set_xlabel('Platform')
                axes[1, 0].set_ylabel('Number of Mentions')
                axes[1, 0].tick_params(axis='x', rotation=45)
                axes[1, 0].grid(True, alpha=0.3, axis='y')
        
        # 4. Topic frequency
        if last_month:
            top_topics = trends['monthly_metrics'][last_month].get('top_topics', {})
            if top_topics:
                topics = list(top_topics.keys())[:5]
                frequencies = list(top_topics.values())[:5]
                
                axes[1, 1].barh(topics, frequencies, color='#9b59b6', alpha=0.7)
                axes[1, 1].set_title('Top Topics', fontweight='bold')
                axes[1, 1].set_xlabel('Frequency')
                axes[1, 1].grid(True, alpha=0.3, axis='x')
        
        plt.tight_layout()
        
        filename = self.output_dir / "visualizations" / f"sentiment_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"Visualization saved to {filename}")
        return str(filename)
    
    def get_active_alerts(self, severity: Optional[str] = None) -> List[Dict]:
        """Get active (unresolved) alerts"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = 'SELECT * FROM sentiment_alerts WHERE resolved = 0'
        params = []
        
        if severity:
            query += ' AND severity = ?'
            params.append(severity)
        
        query += ' ORDER BY created_at DESC LIMIT 50'
        
        cursor.execute(query, params)
        columns = [desc[0] for desc in cursor.description]
        alerts = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        conn.close()
        return alerts
    
    def analyze_competitor_comparison(self, competitors: List[str], 
                                     product_service: str = "") -> Dict:
        """Compare sentiment against competitors"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        results = {}
        
        for competitor in competitors:
            query = '''
                SELECT AVG(sentiment_score) as avg_sentiment, COUNT(*) as count
                FROM sentiment_data
                WHERE date >= ? AND competitor_mention = ?
            '''
            params = [six_months_ago.isoformat(), competitor]
            
            if product_service:
                query += ' AND product_service = ?'
                params.append(product_service)
            
            cursor.execute(query, params)
            row = cursor.fetchone()
            
            if row and row[0] is not None:
                results[competitor] = {
                    'avg_sentiment': round(row[0], 3),
                    'mention_count': row[1]
                }
        
        conn.close()
        
        # Calculate comparison scores
        if results:
            main_product_sentiment = self._get_main_product_sentiment(product_service)
            
            for competitor, data in results.items():
                comparison = data['avg_sentiment'] - main_product_sentiment
                results[competitor]['comparison_score'] = round(comparison, 3)
                results[competitor]['advantage'] = 'competitor' if comparison > 0 else 'main_product'
        
        return results
    
    def _get_main_product_sentiment(self, product_service: str) -> float:
        """Get average sentiment for main product"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT AVG(sentiment_score) as avg_sentiment
            FROM sentiment_data
            WHERE date >= ? AND (competitor_mention IS NULL OR competitor_mention = '')
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        row = cursor.fetchone()
        conn.close()
        
        return row[0] if row and row[0] is not None else 0.0
    
    def predict_future_trends(self, product_service: str = "", days_ahead: int = 30) -> Dict:
        """Predict future sentiment trends using time series analysis"""
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for prediction'}
        
        # Extract monthly sentiment scores
        months = sorted(trends['monthly_metrics'].keys())
        sentiment_scores = [trends['monthly_metrics'][m]['avg_sentiment'] for m in months]
        
        if len(sentiment_scores) < 3:
            return {'success': False, 'error': 'Need at least 3 months of data'}
        
        # Simple linear regression for trend prediction
        x = list(range(len(sentiment_scores)))
        y = sentiment_scores
        
        # Calculate slope and intercept
        n = len(x)
        sum_x = sum(x)
        sum_y = sum(y)
        sum_xy = sum(x[i] * y[i] for i in range(n))
        sum_x2 = sum(x[i] ** 2 for i in range(n))
        
        slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2) if (n * sum_x2 - sum_x ** 2) != 0 else 0
        intercept = (sum_y - slope * sum_x) / n
        
        # Predict future values
        future_months = len(months) + (days_ahead // 30)
        predicted_scores = []
        for i in range(len(months), future_months):
            predicted = slope * i + intercept
            predicted_scores.append(max(-1.0, min(1.0, predicted)))  # Clamp to [-1, 1]
        
        # Calculate confidence based on historical variance
        variance = statistics.variance(sentiment_scores) if len(sentiment_scores) > 1 else 0
        confidence = max(0.0, min(1.0, 1.0 - variance))
        
        return {
            'success': True,
            'current_trend': trends['overall_trend']['direction'],
            'predicted_sentiment': round(predicted_scores[-1] if predicted_scores else 0, 3),
            'predicted_change': round(predicted_scores[-1] - sentiment_scores[-1] if predicted_scores else 0, 3),
            'confidence': round(confidence, 3),
            'trend_slope': round(slope, 4),
            'forecast_period_days': days_ahead,
            'prediction': 'improving' if slope > 0 else 'declining' if slope < 0 else 'stable'
        }
    
    def analyze_by_time_of_day(self, product_service: str = "") -> Dict:
        """Analyze sentiment patterns by time of day"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, sentiment_score, sentiment_label
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by hour of day
        hourly_sentiment = defaultdict(list)
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                hour = date_obj.hour
                hourly_sentiment[hour].append(row[1])
            except:
                continue
        
        # Calculate averages
        hourly_avg = {}
        for hour in range(24):
            if hour in hourly_sentiment:
                hourly_avg[hour] = {
                    'avg_sentiment': round(statistics.mean(hourly_sentiment[hour]), 3),
                    'count': len(hourly_sentiment[hour])
                }
        
        # Find best and worst hours
        if hourly_avg:
            best_hour = max(hourly_avg.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_hour = min(hourly_avg.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'hourly_breakdown': hourly_avg,
                'best_hour': {
                    'hour': best_hour[0],
                    'sentiment': best_hour[1]['avg_sentiment'],
                    'count': best_hour[1]['count']
                },
                'worst_hour': {
                    'hour': worst_hour[0],
                    'sentiment': worst_hour[1]['avg_sentiment'],
                    'count': worst_hour[1]['count']
                },
                'insights': self._generate_time_insights(hourly_avg)
            }
        
        return {'success': False, 'error': 'Insufficient time data'}
    
    def _generate_time_insights(self, hourly_avg: Dict) -> List[str]:
        """Generate insights from hourly analysis"""
        insights = []
        
        if not hourly_avg:
            return insights
        
        # Find peak hours (most activity)
        peak_hours = sorted(hourly_avg.items(), key=lambda x: x[1]['count'], reverse=True)[:3]
        insights.append(f"Peak activity hours: {', '.join([f'{h[0]}:00' for h in peak_hours])}")
        
        # Find sentiment patterns
        morning_hours = [h for h in range(6, 12) if h in hourly_avg]
        afternoon_hours = [h for h in range(12, 18) if h in hourly_avg]
        evening_hours = [h for h in range(18, 24) if h in hourly_avg]
        
        if morning_hours and afternoon_hours and evening_hours:
            morning_avg = statistics.mean([hourly_avg[h]['avg_sentiment'] for h in morning_hours])
            afternoon_avg = statistics.mean([hourly_avg[h]['avg_sentiment'] for h in afternoon_hours])
            evening_avg = statistics.mean([hourly_avg[h]['avg_sentiment'] for h in evening_hours])
            
            best_period = max([
                ('morning', morning_avg),
                ('afternoon', afternoon_avg),
                ('evening', evening_avg)
            ], key=lambda x: x[1])
            
            insights.append(f"Best sentiment period: {best_period[0]} ({best_period[1]:.2f})")
        
        return insights
    
    def analyze_hashtags_and_mentions(self, product_service: str = "") -> Dict:
        """Extract and analyze hashtags and mentions"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, sentiment_score, engagement, platform
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        hashtags = defaultdict(lambda: {'count': 0, 'total_sentiment': 0, 'total_engagement': 0})
        mentions = defaultdict(lambda: {'count': 0, 'total_sentiment': 0, 'total_engagement': 0})
        
        for row in data:
            content = row[0] or ""
            sentiment = row[1]
            engagement = row[2] or 0
            platform = row[3] or ""
            
            # Extract hashtags
            hashtag_pattern = r'#(\w+)'
            found_hashtags = re.findall(hashtag_pattern, content, re.IGNORECASE)
            for tag in found_hashtags:
                hashtags[tag.lower()]['count'] += 1
                hashtags[tag.lower()]['total_sentiment'] += sentiment
                hashtags[tag.lower()]['total_engagement'] += engagement
            
            # Extract mentions
            mention_pattern = r'@(\w+)'
            found_mentions = re.findall(mention_pattern, content, re.IGNORECASE)
            for mention in found_mentions:
                mentions[mention.lower()]['count'] += 1
                mentions[mention.lower()]['total_sentiment'] += sentiment
                mentions[mention.lower()]['total_engagement'] += engagement
        
        # Calculate averages and rankings
        hashtag_analysis = {}
        for tag, data in hashtags.items():
            hashtag_analysis[tag] = {
                'count': data['count'],
                'avg_sentiment': round(data['total_sentiment'] / data['count'], 3),
                'total_engagement': data['total_engagement'],
                'avg_engagement': round(data['total_engagement'] / data['count'], 1)
            }
        
        mention_analysis = {}
        for mention, data in mentions.items():
            mention_analysis[mention] = {
                'count': data['count'],
                'avg_sentiment': round(data['total_sentiment'] / data['count'], 3),
                'total_engagement': data['total_engagement'],
                'avg_engagement': round(data['total_engagement'] / data['count'], 1)
            }
        
        # Top hashtags and mentions
        top_hashtags = sorted(hashtag_analysis.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
        top_mentions = sorted(mention_analysis.items(), key=lambda x: x[1]['count'], reverse=True)[:10]
        
        return {
            'success': True,
            'total_hashtags': len(hashtag_analysis),
            'total_mentions': len(mention_analysis),
            'top_hashtags': dict(top_hashtags),
            'top_mentions': dict(top_mentions),
            'hashtag_analysis': hashtag_analysis,
            'mention_analysis': mention_analysis
        }
    
    def calculate_virality_score(self, content: str, engagement: int, 
                                author_followers: int, platform: str) -> float:
        """Calculate virality score for content"""
        # Base score from engagement
        engagement_score = min(1.0, math.log10(engagement + 1) / 5)  # Normalize to 0-1
        
        # Follower ratio (engagement relative to followers)
        if author_followers > 0:
            follower_ratio = engagement / author_followers
            follower_score = min(1.0, follower_ratio * 10)  # 10% engagement = max score
        else:
            follower_score = 0.0
        
        # Platform multiplier
        platform_multipliers = {
            'twitter': 1.0,
            'instagram': 1.2,
            'tiktok': 1.5,
            'youtube': 1.3,
            'facebook': 0.9,
            'reddit': 0.8
        }
        platform_mult = platform_multipliers.get(platform.lower(), 1.0)
        
        # Content length factor (optimal around 100-200 chars)
        content_length = len(content)
        if 100 <= content_length <= 200:
            length_score = 1.0
        elif content_length < 50:
            length_score = 0.7
        elif content_length > 500:
            length_score = 0.8
        else:
            length_score = 0.9
        
        # Calculate final virality score
        virality = (engagement_score * 0.4 + follower_score * 0.4 + length_score * 0.2) * platform_mult
        
        return round(min(1.0, virality), 4)
    
    def analyze_influencer_impact(self, min_followers: int = 10000, 
                                  product_service: str = "") -> Dict:
        """Analyze impact of influencers vs regular users"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT sentiment_score, engagement, author_followers, platform, content
            FROM sentiment_data
            WHERE date >= ? AND author_followers >= ?
        '''
        params = [six_months_ago.isoformat(), min_followers]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        influencer_data = cursor.fetchall()
        
        # Regular users
        query2 = '''
            SELECT sentiment_score, engagement, author_followers, platform, content
            FROM sentiment_data
            WHERE date >= ? AND (author_followers < ? OR author_followers IS NULL)
        '''
        params2 = [six_months_ago.isoformat(), min_followers]
        
        if product_service:
            query2 += ' AND product_service = ?'
            params2.append(product_service)
        
        cursor.execute(query2, params2)
        regular_data = cursor.fetchall()
        conn.close()
        
        # Analyze influencer data
        if influencer_data:
            influencer_sentiments = [row[0] for row in influencer_data]
            influencer_engagements = [row[1] or 0 for row in influencer_data]
            avg_influencer_sentiment = statistics.mean(influencer_sentiments)
            avg_influencer_engagement = statistics.mean(influencer_engagements)
        else:
            avg_influencer_sentiment = 0
            avg_influencer_engagement = 0
        
        # Analyze regular user data
        if regular_data:
            regular_sentiments = [row[0] for row in regular_data]
            regular_engagements = [row[1] or 0 for row in regular_data]
            avg_regular_sentiment = statistics.mean(regular_sentiments)
            avg_regular_engagement = statistics.mean(regular_engagements)
        else:
            avg_regular_sentiment = 0
            avg_regular_engagement = 0
        
        # Calculate impact
        sentiment_difference = avg_influencer_sentiment - avg_regular_sentiment
        engagement_multiplier = (avg_influencer_engagement / avg_regular_engagement) if avg_regular_engagement > 0 else 0
        
        return {
            'success': True,
            'influencer_count': len(influencer_data),
            'regular_user_count': len(regular_data),
            'influencer_metrics': {
                'avg_sentiment': round(avg_influencer_sentiment, 3),
                'avg_engagement': round(avg_influencer_engagement, 1)
            },
            'regular_user_metrics': {
                'avg_sentiment': round(avg_regular_sentiment, 3),
                'avg_engagement': round(avg_regular_engagement, 1)
            },
            'impact_analysis': {
                'sentiment_difference': round(sentiment_difference, 3),
                'engagement_multiplier': round(engagement_multiplier, 2),
                'influencer_advantage': 'sentiment' if sentiment_difference > 0 else 'engagement' if engagement_multiplier > 1 else 'none'
            }
        }
    
    def calculate_risk_score(self, product_service: str = "") -> Dict:
        """Calculate overall risk score based on multiple factors"""
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data'}
        
        risk_factors = []
        total_risk = 0.0
        
        # Factor 1: Negative sentiment trend
        if trends['overall_trend']['direction'] == 'declining':
            risk_score = min(1.0, abs(trends['overall_trend']['magnitude']) * 2)
            risk_factors.append({
                'factor': 'declining_sentiment',
                'score': round(risk_score, 3),
                'weight': 0.3
            })
            total_risk += risk_score * 0.3
        
        # Factor 2: High negative sentiment percentage
        last_month = sorted(trends['monthly_metrics'].keys())[-1] if trends['monthly_metrics'] else None
        if last_month:
            metrics = trends['monthly_metrics'][last_month]
            negative_pct = (metrics['negative_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            if negative_pct > 30:
                risk_score = min(1.0, (negative_pct - 30) / 50)  # 30% = 0, 80% = 1.0
                risk_factors.append({
                    'factor': 'high_negative_percentage',
                    'score': round(risk_score, 3),
                    'weight': 0.25
                })
                total_risk += risk_score * 0.25
        
        # Factor 3: Increasing dissatisfaction
        if trends.get('dissatisfaction_trends'):
            for trend in trends['dissatisfaction_trends']:
                if trend.get('type') == 'increasing_dissatisfaction':
                    risk_score = min(1.0, trend.get('increase', 0) / 20)  # 20% increase = max risk
                    risk_factors.append({
                        'factor': 'increasing_dissatisfaction',
                        'score': round(risk_score, 3),
                        'weight': 0.25
                    })
                    total_risk += risk_score * 0.25
                    break
        
        # Factor 4: Critical alerts
        alerts = self.get_active_alerts(severity='critical')
        if alerts:
            risk_score = min(1.0, len(alerts) / 5)  # 5+ critical alerts = max risk
            risk_factors.append({
                'factor': 'critical_alerts',
                'score': round(risk_score, 3),
                'weight': 0.2
            })
            total_risk += risk_score * 0.2
        
        # Normalize to 0-100 scale
        risk_score_100 = round(total_risk * 100, 1)
        
        # Determine risk level
        if risk_score_100 >= 70:
            risk_level = 'critical'
        elif risk_score_100 >= 50:
            risk_level = 'high'
        elif risk_score_100 >= 30:
            risk_level = 'moderate'
        else:
            risk_level = 'low'
        
        return {
            'success': True,
            'risk_score': risk_score_100,
            'risk_level': risk_level,
            'risk_factors': risk_factors,
            'recommendations': self._generate_risk_recommendations(risk_level, risk_factors)
        }
    
    def _generate_risk_recommendations(self, risk_level: str, risk_factors: List[Dict]) -> List[str]:
        """Generate recommendations based on risk level"""
        recommendations = []
        
        if risk_level == 'critical':
            recommendations.append("URGENT: Immediate action required to address negative sentiment")
            recommendations.append("Consider crisis communication strategy")
            recommendations.append("Engage with negative feedback directly and publicly")
        
        if risk_level in ['critical', 'high']:
            recommendations.append("Review and address recurring issues mentioned in feedback")
            recommendations.append("Increase monitoring frequency")
            recommendations.append("Consider product/service improvements based on feedback")
        
        # Specific recommendations based on risk factors
        for factor in risk_factors:
            if factor['factor'] == 'declining_sentiment':
                recommendations.append("Investigate causes of sentiment decline")
            elif factor['factor'] == 'high_negative_percentage':
                recommendations.append("Focus on improving customer satisfaction")
            elif factor['factor'] == 'increasing_dissatisfaction':
                recommendations.append("Address dissatisfaction trends immediately")
            elif factor['factor'] == 'critical_alerts':
                recommendations.append("Review and resolve critical alerts")
        
        return recommendations
    
    def export_to_excel(self, product_service: str = "", filename: Optional[str] = None):
        """Export comprehensive report to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            EXCEL_AVAILABLE = True
        except ImportError:
            logger.warning("openpyxl not available, cannot export to Excel")
            return None
        
        if filename is None:
            filename = self.output_dir / "exports" / f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = self.output_dir / "exports" / filename
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        trends = self.analyze_6_month_trends(product_service)
        
        ws1['A1'] = "Social Media Sentiment Analysis Report"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:D1')
        
        row = 3
        ws1[f'A{row}'] = "Overall Trend:"
        ws1[f'B{row}'] = trends.get('overall_trend', {}).get('direction', 'N/A')
        row += 1
        
        ws1[f'A{row}'] = "Risk Score:"
        risk = self.calculate_risk_score(product_service)
        ws1[f'B{row}'] = f"{risk.get('risk_score', 0)} ({risk.get('risk_level', 'N/A')})"
        row += 2
        
        # Sheet 2: Monthly Trends
        ws2 = wb.create_sheet("Monthly Trends")
        ws2.append(["Month", "Avg Sentiment", "Positive", "Negative", "Neutral", "Total"])
        
        for month in sorted(trends.get('monthly_metrics', {}).keys()):
            metrics = trends['monthly_metrics'][month]
            ws2.append([
                month,
                metrics['avg_sentiment'],
                metrics['positive_count'],
                metrics['negative_count'],
                metrics['neutral_count'],
                metrics['total_count']
            ])
        
        # Sheet 3: Top Insights
        ws3 = wb.create_sheet("Top Insights")
        ws3.append(["Insight"])
        for insight in trends.get('top_insights', []):
            ws3.append([insight])
        
        wb.save(filename)
        logger.info(f"Excel report exported to {filename}")
        return str(filename)
    
    def analyze_by_day_of_week(self, product_service: str = "") -> Dict:
        """Analyze sentiment patterns by day of the week"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by day of week (0=Monday, 6=Sunday)
        daily_sentiment = defaultdict(lambda: {'scores': [], 'labels': [], 'engagement': []})
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                day_of_week = date_obj.weekday()
                daily_sentiment[day_of_week]['scores'].append(row[1])
                daily_sentiment[day_of_week]['labels'].append(row[2])
                daily_sentiment[day_of_week]['engagement'].append(row[3] or 0)
            except:
                continue
        
        # Calculate metrics per day
        daily_metrics = {}
        for day_num in range(7):
            if day_num in daily_sentiment:
                day_data = daily_sentiment[day_num]
                daily_metrics[day_names[day_num]] = {
                    'avg_sentiment': round(statistics.mean(day_data['scores']), 3),
                    'total_count': len(day_data['scores']),
                    'positive_pct': round((day_data['labels'].count('positive') / len(day_data['labels']) * 100), 1),
                    'negative_pct': round((day_data['labels'].count('negative') / len(day_data['labels']) * 100), 1),
                    'avg_engagement': round(statistics.mean(day_data['engagement']), 1)
                }
        
        # Find best and worst days
        if daily_metrics:
            best_day = max(daily_metrics.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_day = min(daily_metrics.items(), key=lambda x: x[1]['avg_sentiment'])
            most_active_day = max(daily_metrics.items(), key=lambda x: x[1]['total_count'])
            
            return {
                'success': True,
                'daily_breakdown': daily_metrics,
                'best_day': {
                    'day': best_day[0],
                    'sentiment': best_day[1]['avg_sentiment'],
                    'count': best_day[1]['total_count']
                },
                'worst_day': {
                    'day': worst_day[0],
                    'sentiment': worst_day[1]['avg_sentiment'],
                    'count': worst_day[1]['total_count']
                },
                'most_active_day': {
                    'day': most_active_day[0],
                    'count': most_active_day[1]['total_count']
                },
                'insights': self._generate_day_insights(daily_metrics)
            }
        
        return {'success': False, 'error': 'Insufficient data'}
    
    def _generate_day_insights(self, daily_metrics: Dict) -> List[str]:
        """Generate insights from daily analysis"""
        insights = []
        
        if not daily_metrics:
            return insights
        
        # Weekend vs weekday analysis
        weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        weekends = ['Saturday', 'Sunday']
        
        weekday_sentiments = [daily_metrics[d]['avg_sentiment'] for d in weekdays if d in daily_metrics]
        weekend_sentiments = [daily_metrics[d]['avg_sentiment'] for d in weekends if d in daily_metrics]
        
        if weekday_sentiments and weekend_sentiments:
            weekday_avg = statistics.mean(weekday_sentiments)
            weekend_avg = statistics.mean(weekend_sentiments)
            
            if weekend_avg > weekday_avg + 0.1:
                insights.append(f"Weekend sentiment is significantly better (+{weekend_avg - weekday_avg:.2f})")
            elif weekday_avg > weekend_avg + 0.1:
                insights.append(f"Weekday sentiment is significantly better (+{weekday_avg - weekend_avg:.2f})")
        
        return insights
    
    def analyze_by_language(self, product_service: str = "") -> Dict:
        """Analyze sentiment by language"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT language, sentiment_score, sentiment_label, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY language'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        language_analysis = {}
        total_count = 0
        
        for row in data:
            lang = row[0] or 'unknown'
            avg_sentiment = row[1]
            label = row[2]
            count = row[3]
            total_count += count
            
            language_analysis[lang] = {
                'avg_sentiment': round(avg_sentiment, 3),
                'count': count,
                'percentage': 0  # Will calculate after
            }
        
        # Calculate percentages
        for lang in language_analysis:
            language_analysis[lang]['percentage'] = round(
                (language_analysis[lang]['count'] / total_count * 100), 1
            )
        
        # Find dominant language
        dominant_lang = max(language_analysis.items(), key=lambda x: x[1]['count'])
        
        return {
            'success': True,
            'language_breakdown': language_analysis,
            'dominant_language': {
                'language': dominant_lang[0],
                'count': dominant_lang[1]['count'],
                'percentage': dominant_lang[1]['percentage']
            },
            'total_languages': len(language_analysis)
        }
    
    def detect_bot_patterns(self, product_service: str = "") -> Dict:
        """Detect potential bot accounts based on patterns"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT author_id, COUNT(*) as post_count, 
                   AVG(sentiment_score) as avg_sentiment,
                   AVG(engagement) as avg_engagement,
                   MIN(date) as first_post,
                   MAX(date) as last_post
            FROM sentiment_data
            WHERE date >= ? AND author_id IS NOT NULL
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY author_id HAVING post_count > 1'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        suspicious_accounts = []
        
        for row in data:
            author_id = row[0]
            post_count = row[1]
            avg_sentiment = row[2]
            avg_engagement = row[3]
            first_post = row[4]
            last_post = row[5]
            
            # Calculate time span
            try:
                first_date = datetime.fromisoformat(first_post)
                last_date = datetime.fromisoformat(last_post)
                time_span = (last_date - first_date).total_seconds() / 3600  # hours
                
                # Bot indicators
                bot_score = 0
                indicators = []
                
                # High posting frequency (more than 10 posts per hour)
                if time_span > 0:
                    posts_per_hour = post_count / time_span
                    if posts_per_hour > 10:
                        bot_score += 30
                        indicators.append('high_posting_frequency')
                
                # Very consistent sentiment (all same)
                if abs(avg_sentiment) < 0.1:
                    bot_score += 20
                    indicators.append('neutral_sentiment_only')
                
                # Low engagement despite many posts
                if post_count > 20 and avg_engagement < 5:
                    bot_score += 25
                    indicators.append('low_engagement')
                
                # Very short time span with many posts
                if time_span < 24 and post_count > 50:
                    bot_score += 25
                    indicators.append('burst_posting')
                
                if bot_score >= 30:  # Threshold for suspicious
                    suspicious_accounts.append({
                        'author_id': author_id,
                        'bot_score': bot_score,
                        'post_count': post_count,
                        'indicators': indicators,
                        'avg_sentiment': round(avg_sentiment, 3),
                        'avg_engagement': round(avg_engagement, 1)
                    })
            except:
                continue
        
        # Sort by bot score
        suspicious_accounts.sort(key=lambda x: x['bot_score'], reverse=True)
        
        return {
            'success': True,
            'suspicious_accounts': suspicious_accounts[:20],  # Top 20
            'total_suspicious': len(suspicious_accounts),
            'total_accounts_analyzed': len(data)
        }
    
    def generate_automated_recommendations(self, product_service: str = "") -> Dict:
        """Generate automated actionable recommendations"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        time_analysis = self.analyze_by_time_of_day(product_service)
        day_analysis = self.analyze_by_day_of_week(product_service)
        
        recommendations = []
        priority = []
        
        # High priority recommendations based on risk
        if risk.get('success') and risk.get('risk_level') in ['critical', 'high']:
            priority.append({
                'priority': 'HIGH',
                'category': 'Risk Management',
                'recommendation': 'Immediate action required: Address negative sentiment trends',
                'action_items': [
                    'Review all critical alerts',
                    'Engage with negative feedback',
                    'Implement crisis communication plan'
                ]
            })
        
        # Time-based recommendations
        if time_analysis.get('success') and 'best_hour' in time_analysis:
            best_hour = time_analysis['best_hour']['hour']
            recommendations.append({
                'priority': 'MEDIUM',
                'category': 'Content Timing',
                'recommendation': f'Optimize posting times: Best sentiment at {best_hour}:00',
                'action_items': [
                    f'Schedule important content for {best_hour}:00',
                    'Test posting at optimal hours',
                    'Monitor engagement at different times'
                ]
            })
        
        # Day-based recommendations
        if day_analysis.get('success') and 'best_day' in day_analysis:
            best_day = day_analysis['best_day']['day']
            recommendations.append({
                'priority': 'MEDIUM',
                'category': 'Content Scheduling',
                'recommendation': f'Best day for positive sentiment: {best_day}',
                'action_items': [
                    f'Plan major announcements for {best_day}',
                    'Increase engagement activities on best days'
                ]
            })
        
        # Trend-based recommendations
        if trends.get('success'):
            if trends['overall_trend']['direction'] == 'declining':
                recommendations.append({
                    'priority': 'HIGH',
                    'category': 'Sentiment Improvement',
                    'recommendation': 'Sentiment is declining - investigate causes',
                    'action_items': [
                        'Review recent product/service changes',
                        'Analyze competitor activities',
                        'Survey customers for feedback'
                    ]
                })
            
            # Emerging preferences
            if trends.get('emerging_preferences'):
                top_pref = trends['emerging_preferences'][0]
                recommendations.append({
                    'priority': 'MEDIUM',
                    'category': 'Product Development',
                    'recommendation': f"Emerging preference detected: {top_pref['description']}",
                    'action_items': [
                        f'Consider enhancing {top_pref["topic"].replace("_", " ")} features',
                        'Monitor this trend closely',
                        'Gather more customer feedback on this topic'
                    ]
                })
            
            # Dissatisfaction trends
            if trends.get('dissatisfaction_trends'):
                for trend in trends['dissatisfaction_trends'][:2]:
                    if trend.get('type') == 'recurring_issue':
                        recommendations.append({
                            'priority': 'HIGH',
                            'category': 'Issue Resolution',
                            'recommendation': f"Recurring issue: {trend.get('description', '')}",
                            'action_items': [
                                'Investigate root cause',
                                'Develop solution plan',
                                'Communicate fix to customers'
                            ]
                        })
        
        # Competitor recommendations
        if trends.get('success') and trends.get('sentiment_shifts'):
            largest_shift = max(trends['sentiment_shifts'], key=lambda x: abs(x['change']))
            if largest_shift['direction'] == 'negative':
                recommendations.append({
                    'priority': 'MEDIUM',
                    'category': 'Competitive Analysis',
                    'recommendation': 'Significant negative sentiment shift detected',
                    'action_items': [
                        'Review competitor activities during this period',
                        'Check for external factors (news, events)',
                        'Consider competitive response'
                    ]
                })
        
        return {
            'success': True,
            'total_recommendations': len(recommendations) + len(priority),
            'high_priority': priority,
            'recommendations': recommendations,
            'summary': f"Generated {len(recommendations) + len(priority)} actionable recommendations"
        }
    
    def analyze_emotion_trends(self, product_service: str = "") -> Dict:
        """Analyze trends in specific emotions over time"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, emotions
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by month
        monthly_emotions = defaultdict(lambda: defaultdict(list))
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                month_key = f"{date_obj.year}-{date_obj.month:02d}"
                emotions = json.loads(row[1]) if row[1] else {}
                
                for emotion, score in emotions.items():
                    monthly_emotions[month_key][emotion].append(score)
            except:
                continue
        
        # Calculate monthly averages
        emotion_trends = {}
        all_emotions = set()
        
        for month, emotions in monthly_emotions.items():
            for emotion in emotions:
                all_emotions.add(emotion)
        
        for emotion in all_emotions:
            emotion_trends[emotion] = []
            for month in sorted(monthly_emotions.keys()):
                if emotion in monthly_emotions[month]:
                    avg_score = statistics.mean(monthly_emotions[month][emotion])
                    emotion_trends[emotion].append({
                        'month': month,
                        'avg_score': round(avg_score, 3)
                    })
        
        # Find trending emotions (increasing)
        trending_emotions = []
        for emotion, trend_data in emotion_trends.items():
            if len(trend_data) >= 2:
                first_score = trend_data[0]['avg_score']
                last_score = trend_data[-1]['avg_score']
                change = last_score - first_score
                
                if change > 0.1:  # Significant increase
                    trending_emotions.append({
                        'emotion': emotion,
                        'change': round(change, 3),
                        'trend': 'increasing'
                    })
        
        trending_emotions.sort(key=lambda x: x['change'], reverse=True)
        
        return {
            'success': True,
            'emotion_trends': emotion_trends,
            'trending_emotions': trending_emotions[:5],
            'total_emotions_tracked': len(all_emotions)
        }
    
    def create_comprehensive_dashboard_data(self, product_service: str = "") -> Dict:
        """Create comprehensive data for dashboard visualization"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        time_analysis = self.analyze_by_time_of_day(product_service)
        day_analysis = self.analyze_by_day_of_week(product_service)
        language_analysis = self.analyze_by_language(product_service)
        hashtag_analysis = self.analyze_hashtags_and_mentions(product_service)
        recommendations = self.generate_automated_recommendations(product_service)
        emotion_trends = self.analyze_emotion_trends(product_service)
        
        return {
            'success': True,
            'timestamp': datetime.now().isoformat(),
            'product_service': product_service or 'All Products',
            'overview': {
                'overall_trend': trends.get('overall_trend', {}),
                'risk_score': risk.get('risk_score', 0),
                'risk_level': risk.get('risk_level', 'unknown')
            },
            'trends': {
                'monthly_metrics': trends.get('monthly_metrics', {}),
                'emerging_preferences': trends.get('emerging_preferences', []),
                'dissatisfaction_trends': trends.get('dissatisfaction_trends', [])
            },
            'temporal_analysis': {
                'time_of_day': time_analysis.get('hourly_breakdown', {}),
                'day_of_week': day_analysis.get('daily_breakdown', {})
            },
            'content_analysis': {
                'hashtags': hashtag_analysis.get('top_hashtags', {}),
                'mentions': hashtag_analysis.get('top_mentions', {}),
                'languages': language_analysis.get('language_breakdown', {})
            },
            'emotions': {
                'trends': emotion_trends.get('emotion_trends', {}),
                'trending': emotion_trends.get('trending_emotions', [])
            },
            'recommendations': {
                'high_priority': recommendations.get('high_priority', []),
                'all': recommendations.get('recommendations', [])
            },
            'alerts': {
                'critical': len(self.get_active_alerts(severity='critical')),
                'high': len(self.get_active_alerts(severity='high')),
                'all': len(self.get_active_alerts())
            }
        }
    
    def analyze_by_geographic_region(self, product_service: str = "") -> Dict:
        """Analyze sentiment by geographic region"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT location, sentiment_score, sentiment_label, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ? AND location IS NOT NULL AND location != ''
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY location'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No geographic data available'}
        
        region_analysis = {}
        total_count = 0
        
        for row in data:
            location = row[0]
            avg_sentiment = row[1]
            label = row[2]
            count = row[3]
            total_count += count
            
            region_analysis[location] = {
                'avg_sentiment': round(avg_sentiment, 3),
                'count': count,
                'percentage': 0,
                'positive_pct': 0,
                'negative_pct': 0
            }
        
        # Calculate percentages
        for region in region_analysis:
            region_analysis[region]['percentage'] = round(
                (region_analysis[region]['count'] / total_count * 100), 1
            )
        
        # Find best and worst regions
        if region_analysis:
            best_region = max(region_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_region = min(region_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            most_active_region = max(region_analysis.items(), key=lambda x: x[1]['count'])
            
            return {
                'success': True,
                'region_breakdown': region_analysis,
                'best_region': {
                    'region': best_region[0],
                    'sentiment': best_region[1]['avg_sentiment'],
                    'count': best_region[1]['count']
                },
                'worst_region': {
                    'region': worst_region[0],
                    'sentiment': worst_region[1]['avg_sentiment'],
                    'count': worst_region[1]['count']
                },
                'most_active_region': {
                    'region': most_active_region[0],
                    'count': most_active_region[1]['count']
                },
                'total_regions': len(region_analysis)
            }
        
        return {'success': False, 'error': 'Insufficient geographic data'}
    
    def analyze_sentiment_by_topic(self, topic: str, product_service: str = "") -> Dict:
        """Analyze sentiment for a specific topic"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT sentiment_score, sentiment_label, engagement, date, platform
            FROM sentiment_data
            WHERE date >= ? AND topics LIKE ?
        '''
        params = [six_months_ago.isoformat(), f'%{topic}%']
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': f'No data found for topic: {topic}'}
        
        # Calculate metrics
        sentiments = [row[0] for row in data]
        labels = [row[2] for row in data]
        engagements = [row[2] or 0 for row in data]
        
        avg_sentiment = statistics.mean(sentiments)
        label_counts = Counter(labels)
        total_engagement = sum(engagements)
        avg_engagement = statistics.mean(engagements) if engagements else 0
        
        # Platform distribution
        platforms = Counter([row[4] for row in data])
        
        # Time trend
        monthly_sentiment = defaultdict(list)
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[3])
                month_key = f"{date_obj.year}-{date_obj.month:02d}"
                monthly_sentiment[month_key].append(row[0])
            except:
                continue
        
        monthly_avg = {}
        for month, scores in monthly_sentiment.items():
            monthly_avg[month] = round(statistics.mean(scores), 3)
        
        return {
            'success': True,
            'topic': topic,
            'total_mentions': len(data),
            'avg_sentiment': round(avg_sentiment, 3),
            'sentiment_distribution': dict(label_counts),
            'positive_pct': round((label_counts.get('positive', 0) / len(data) * 100), 1),
            'negative_pct': round((label_counts.get('negative', 0) / len(data) * 100), 1),
            'total_engagement': total_engagement,
            'avg_engagement': round(avg_engagement, 1),
            'platform_distribution': dict(platforms),
            'monthly_trend': monthly_avg
        }
    
    def analyze_sentiment_by_platform_comparison(self, product_service: str = "") -> Dict:
        """Compare sentiment across different platforms"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT platform, 
                   AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count,
                   AVG(engagement) as avg_engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY platform'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No platform data available'}
        
        platform_comparison = {}
        for row in data:
            platform = row[0] or 'unknown'
            platform_comparison[platform] = {
                'avg_sentiment': round(row[1], 3),
                'total_count': row[2],
                'positive_count': row[3],
                'negative_count': row[4],
                'positive_pct': round((row[3] / row[2] * 100), 1) if row[2] > 0 else 0,
                'negative_pct': round((row[4] / row[2] * 100), 1) if row[2] > 0 else 0,
                'avg_engagement': round(row[5] or 0, 1)
            }
        
        # Find best and worst platforms
        if platform_comparison:
            best_platform = max(platform_comparison.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_platform = min(platform_comparison.items(), key=lambda x: x[1]['avg_sentiment'])
            most_active_platform = max(platform_comparison.items(), key=lambda x: x[1]['total_count'])
            
            return {
                'success': True,
                'platform_comparison': platform_comparison,
                'best_platform': {
                    'platform': best_platform[0],
                    'sentiment': best_platform[1]['avg_sentiment'],
                    'count': best_platform[1]['total_count']
                },
                'worst_platform': {
                    'platform': worst_platform[0],
                    'sentiment': worst_platform[1]['avg_sentiment'],
                    'count': worst_platform[1]['total_count']
                },
                'most_active_platform': {
                    'platform': most_active_platform[0],
                    'count': most_active_platform[1]['total_count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient platform data'}
    
    def analyze_sentiment_by_engagement_level(self, product_service: str = "") -> Dict:
        """Analyze sentiment correlation with engagement levels"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Categorize by engagement levels
        engagement_levels = {
            'viral': {'min': 10000, 'scores': [], 'count': 0},
            'high': {'min': 1000, 'max': 9999, 'scores': [], 'count': 0},
            'medium': {'min': 100, 'max': 999, 'scores': [], 'count': 0},
            'low': {'min': 10, 'max': 99, 'scores': [], 'count': 0},
            'minimal': {'max': 9, 'scores': [], 'count': 0}
        }
        
        for row in data:
            engagement = row[2] or 0
            sentiment = row[1]
            
            if engagement >= 10000:
                level = 'viral'
            elif engagement >= 1000:
                level = 'high'
            elif engagement >= 100:
                level = 'medium'
            elif engagement >= 10:
                level = 'low'
            else:
                level = 'minimal'
            
            engagement_levels[level]['scores'].append(row[0])
            engagement_levels[level]['count'] += 1
        
        # Calculate averages
        level_analysis = {}
        for level, data_dict in engagement_levels.items():
            if data_dict['count'] > 0:
                level_analysis[level] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'count': data_dict['count'],
                    'percentage': round((data_dict['count'] / len(data) * 100), 1)
                }
        
        # Calculate correlation
        sentiments = [row[0] for row in data]
        engagements = [row[2] or 0 for row in data]
        
        # Simple correlation calculation
        if len(sentiments) > 1 and max(engagements) > 0:
            # Normalize for correlation
            norm_sentiments = [(s + 1) / 2 for s in sentiments]  # -1 to 1 -> 0 to 1
            norm_engagements = [e / max(engagements) for e in engagements] if max(engagements) > 0 else engagements
            
            # Calculate correlation coefficient
            if len(norm_sentiments) == len(norm_engagements):
                correlation = self._calculate_correlation(norm_sentiments, norm_engagements)
            else:
                correlation = 0
        else:
            correlation = 0
        
        return {
            'success': True,
            'engagement_level_analysis': level_analysis,
            'correlation': round(correlation, 3),
            'correlation_interpretation': self._interpret_correlation(correlation),
            'insights': self._generate_engagement_insights(level_analysis, correlation)
        }
    
    def _calculate_correlation(self, x: List[float], y: List[float]) -> float:
        """Calculate Pearson correlation coefficient"""
        if len(x) != len(y) or len(x) < 2:
            return 0.0
        
        n = len(x)
        sum_x = sum(x)
        sum_y = sum(y)
        sum_xy = sum(x[i] * y[i] for i in range(n))
        sum_x2 = sum(x[i] ** 2 for i in range(n))
        sum_y2 = sum(y[i] ** 2 for i in range(n))
        
        numerator = n * sum_xy - sum_x * sum_y
        denominator = math.sqrt((n * sum_x2 - sum_x ** 2) * (n * sum_y2 - sum_y ** 2))
        
        if denominator == 0:
            return 0.0
        
        return numerator / denominator
    
    def _interpret_correlation(self, correlation: float) -> str:
        """Interpret correlation coefficient"""
        abs_corr = abs(correlation)
        if abs_corr >= 0.7:
            strength = 'strong'
        elif abs_corr >= 0.4:
            strength = 'moderate'
        elif abs_corr >= 0.2:
            strength = 'weak'
        else:
            strength = 'very weak'
        
        direction = 'positive' if correlation > 0 else 'negative' if correlation < 0 else 'no'
        
        return f"{strength} {direction} correlation"
    
    def _generate_engagement_insights(self, level_analysis: Dict, correlation: float) -> List[str]:
        """Generate insights from engagement analysis"""
        insights = []
        
        if not level_analysis:
            return insights
        
        # Find level with best sentiment
        if level_analysis:
            best_level = max(level_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_level = min(level_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            insights.append(f"Best sentiment at {best_level[0]} engagement level: {best_level[1]['avg_sentiment']:.3f}")
            insights.append(f"Worst sentiment at {worst_level[0]} engagement level: {worst_level[1]['avg_sentiment']:.3f}")
        
        # Correlation insights
        if abs(correlation) > 0.3:
            if correlation > 0:
                insights.append("Positive correlation: Higher engagement correlates with better sentiment")
            else:
                insights.append("Negative correlation: Higher engagement correlates with worse sentiment")
        else:
            insights.append("Weak correlation: Engagement and sentiment are largely independent")
        
        return insights
    
    def analyze_sentiment_by_author_type(self, product_service: str = "") -> Dict:
        """Analyze sentiment by author type (user, influencer, expert, etc.)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT author_type,
                   AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   AVG(engagement) as avg_engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY author_type'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No author type data available'}
        
        author_analysis = {}
        for row in data:
            author_type = row[0] or 'user'
            author_analysis[author_type] = {
                'avg_sentiment': round(row[1], 3),
                'count': row[2],
                'avg_engagement': round(row[3] or 0, 1)
            }
        
        # Find best author type
        if author_analysis:
            best_type = max(author_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'author_type_analysis': author_analysis,
                'best_author_type': {
                    'type': best_type[0],
                    'sentiment': best_type[1]['avg_sentiment'],
                    'count': best_type[1]['count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient author type data'}
    
    def create_sentiment_heatmap_data(self, product_service: str = "") -> Dict:
        """Create data for sentiment heatmap visualization (time vs platform)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, platform, sentiment_score
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by month and platform
        heatmap_data = defaultdict(lambda: defaultdict(list))
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                month_key = f"{date_obj.year}-{date_obj.month:02d}"
                platform = row[1] or 'unknown'
                sentiment = row[2]
                
                heatmap_data[month_key][platform].append(sentiment)
            except:
                continue
        
        # Calculate averages
        heatmap_matrix = {}
        all_platforms = set()
        
        for month, platforms in heatmap_data.items():
            for platform in platforms:
                all_platforms.add(platform)
        
        for month in sorted(heatmap_data.keys()):
            heatmap_matrix[month] = {}
            for platform in all_platforms:
                if platform in heatmap_data[month]:
                    avg_sentiment = statistics.mean(heatmap_data[month][platform])
                    heatmap_matrix[month][platform] = round(avg_sentiment, 3)
                else:
                    heatmap_matrix[month][platform] = None
        
        return {
            'success': True,
            'heatmap_data': heatmap_matrix,
            'platforms': sorted(list(all_platforms)),
            'months': sorted(heatmap_data.keys())
        }
    
    def generate_executive_summary(self, product_service: str = "") -> str:
        """Generate comprehensive executive summary report"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        platform_comparison = self.analyze_sentiment_by_platform_comparison(product_service)
        recommendations = self.generate_automated_recommendations(product_service)
        alerts = self.get_active_alerts()
        
        summary = []
        summary.append("=" * 80)
        summary.append("EXECUTIVE SUMMARY - SOCIAL MEDIA SENTIMENT ANALYSIS")
        summary.append("=" * 80)
        summary.append(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary.append(f"Product/Service: {product_service or 'All Products'}")
        summary.append("")
        
        # Overview
        summary.append("OVERVIEW")
        summary.append("-" * 80)
        if trends.get('success'):
            overall_trend = trends.get('overall_trend', {})
            summary.append(f"Overall Trend: {overall_trend.get('direction', 'N/A').upper()}")
            summary.append(f"Sentiment Change: {overall_trend.get('first_month_avg', 0):.3f} → {overall_trend.get('last_month_avg', 0):.3f}")
        
        if risk.get('success'):
            summary.append(f"Risk Score: {risk.get('risk_score', 0)}/100 ({risk.get('risk_level', 'N/A').upper()})")
        
        summary.append(f"Active Alerts: {len(alerts)} ({len([a for a in alerts if a.get('severity') == 'critical'])} critical)")
        summary.append("")
        
        # Key Metrics
        summary.append("KEY METRICS")
        summary.append("-" * 80)
        if trends.get('success') and trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            metrics = trends['monthly_metrics'][last_month]
            summary.append(f"Last Month ({last_month}):")
            summary.append(f"  Total Mentions: {metrics['total_count']}")
            summary.append(f"  Positive: {metrics['positive_count']} ({metrics['positive_count']/metrics['total_count']*100:.1f}%)")
            summary.append(f"  Negative: {metrics['negative_count']} ({metrics['negative_count']/metrics['total_count']*100:.1f}%)")
            summary.append(f"  Average Sentiment: {metrics['avg_sentiment']:.3f}")
        summary.append("")
        
        # Platform Performance
        if platform_comparison.get('success'):
            summary.append("PLATFORM PERFORMANCE")
            summary.append("-" * 80)
            summary.append(f"Best Platform: {platform_comparison.get('best_platform', {}).get('platform', 'N/A')} (sentiment: {platform_comparison.get('best_platform', {}).get('sentiment', 0):.3f})")
            summary.append(f"Most Active: {platform_comparison.get('most_active_platform', {}).get('platform', 'N/A')} ({platform_comparison.get('most_active_platform', {}).get('count', 0)} mentions)")
            summary.append("")
        
        # Top Recommendations
        if recommendations.get('success'):
            summary.append("TOP RECOMMENDATIONS")
            summary.append("-" * 80)
            high_priority = recommendations.get('high_priority', [])
            if high_priority:
                for rec in high_priority[:3]:
                    summary.append(f"🔴 {rec.get('recommendation', '')}")
            else:
                all_recs = recommendations.get('recommendations', [])
                for rec in all_recs[:3]:
                    summary.append(f"• {rec.get('recommendation', '')}")
            summary.append("")
        
        # Risk Factors
        if risk.get('success') and risk.get('risk_factors'):
            summary.append("RISK FACTORS")
            summary.append("-" * 80)
            for factor in risk['risk_factors'][:3]:
                summary.append(f"• {factor.get('factor', '')}: Score {factor.get('score', 0):.2f}")
            summary.append("")
        
        summary.append("=" * 80)
        
        return "\n".join(summary)
    
    def analyze_sentiment_by_content_length(self, product_service: str = "") -> Dict:
        """Analyze sentiment correlation with content length"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, sentiment_score, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Categorize by content length
        length_categories = {
            'very_short': {'max': 50, 'scores': [], 'count': 0, 'engagement': []},
            'short': {'min': 51, 'max': 140, 'scores': [], 'count': 0, 'engagement': []},
            'medium': {'min': 141, 'max': 280, 'scores': [], 'count': 0, 'engagement': []},
            'long': {'min': 281, 'max': 500, 'scores': [], 'count': 0, 'engagement': []},
            'very_long': {'min': 501, 'scores': [], 'count': 0, 'engagement': []}
        }
        
        for row in data:
            content = row[0] or ""
            length = len(content)
            sentiment = row[1]
            engagement = row[2] or 0
            
            if length <= 50:
                cat = 'very_short'
            elif length <= 140:
                cat = 'short'
            elif length <= 280:
                cat = 'medium'
            elif length <= 500:
                cat = 'long'
            else:
                cat = 'very_long'
            
            length_categories[cat]['scores'].append(sentiment)
            length_categories[cat]['engagement'].append(engagement)
            length_categories[cat]['count'] += 1
        
        # Calculate metrics
        length_analysis = {}
        for cat, data_dict in length_categories.items():
            if data_dict['count'] > 0:
                length_analysis[cat] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1),
                    'count': data_dict['count'],
                    'percentage': round((data_dict['count'] / len(data) * 100), 1)
                }
        
        # Find optimal length
        if length_analysis:
            best_length = max(length_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'length_analysis': length_analysis,
                'optimal_length_category': best_length[0],
                'optimal_sentiment': best_length[1]['avg_sentiment'],
                'insights': self._generate_length_insights(length_analysis)
            }
        
        return {'success': False, 'error': 'Insufficient data'}
    
    def _generate_length_insights(self, length_analysis: Dict) -> List[str]:
        """Generate insights from content length analysis"""
        insights = []
        
        if not length_analysis:
            return insights
        
        # Find category with best engagement
        if length_analysis:
            best_engagement = max(length_analysis.items(), key=lambda x: x[1]['avg_engagement'])
            insights.append(f"Best engagement at {best_engagement[0]} length: {best_engagement[1]['avg_engagement']:.1f}")
        
        # Check if there's a clear optimal range
        medium_sentiment = length_analysis.get('medium', {}).get('avg_sentiment', 0)
        short_sentiment = length_analysis.get('short', {}).get('avg_sentiment', 0)
        
        if medium_sentiment > short_sentiment + 0.1:
            insights.append("Medium-length content (141-280 chars) shows better sentiment than short content")
        
        return insights
    
    def analyze_seasonal_patterns(self, product_service: str = "") -> Dict:
        """Analyze seasonal patterns in sentiment"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT date, sentiment_score, sentiment_label
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by month
        monthly_sentiment = defaultdict(list)
        monthly_labels = defaultdict(list)
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                month = date_obj.month
                monthly_sentiment[month].append(row[1])
                monthly_labels[month].append(row[2])
            except:
                continue
        
        # Calculate monthly averages
        seasonal_patterns = {}
        month_names = {
            1: 'January', 2: 'February', 3: 'March', 4: 'April',
            5: 'May', 6: 'June', 7: 'July', 8: 'August',
            9: 'September', 10: 'October', 11: 'November', 12: 'December'
        }
        
        for month in range(1, 13):
            if month in monthly_sentiment:
                seasonal_patterns[month_names[month]] = {
                    'avg_sentiment': round(statistics.mean(monthly_sentiment[month]), 3),
                    'count': len(monthly_sentiment[month]),
                    'positive_pct': round((monthly_labels[month].count('positive') / len(monthly_labels[month]) * 100), 1)
                }
        
        # Find best and worst months
        if seasonal_patterns:
            best_month = max(seasonal_patterns.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_month = min(seasonal_patterns.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'seasonal_patterns': seasonal_patterns,
                'best_month': {
                    'month': best_month[0],
                    'sentiment': best_month[1]['avg_sentiment']
                },
                'worst_month': {
                    'month': worst_month[0],
                    'sentiment': worst_month[1]['avg_sentiment']
                },
                'insights': self._generate_seasonal_insights(seasonal_patterns)
            }
        
        return {'success': False, 'error': 'Insufficient seasonal data'}
    
    def _generate_seasonal_insights(self, seasonal_patterns: Dict) -> List[str]:
        """Generate insights from seasonal analysis"""
        insights = []
        
        if len(seasonal_patterns) < 3:
            return insights
        
        # Calculate overall average
        all_sentiments = [data['avg_sentiment'] for data in seasonal_patterns.values()]
        overall_avg = statistics.mean(all_sentiments)
        
        # Find months significantly above/below average
        above_avg = [month for month, data in seasonal_patterns.items() 
                    if data['avg_sentiment'] > overall_avg + 0.1]
        below_avg = [month for month, data in seasonal_patterns.items() 
                    if data['avg_sentiment'] < overall_avg - 0.1]
        
        if above_avg:
            insights.append(f"Months with above-average sentiment: {', '.join(above_avg)}")
        if below_avg:
            insights.append(f"Months with below-average sentiment: {', '.join(below_avg)}")
        
        return insights
    
    def analyze_keyword_frequency(self, product_service: str = "", top_n: int = 20) -> Dict:
        """Analyze most frequent keywords and their sentiment"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, sentiment_score, sentiment_label
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Extract keywords (excluding common stop words)
        stop_words = {'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 
                     'of', 'with', 'by', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
                     'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
                     'should', 'may', 'might', 'must', 'can', 'this', 'that', 'these', 'those',
                     'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them'}
        
        keyword_sentiment = defaultdict(lambda: {'count': 0, 'total_sentiment': 0, 'positive': 0, 'negative': 0})
        
        for row in data:
            content = (row[0] or "").lower()
            sentiment = row[1]
            label = row[2]
            
            # Extract words
            words = re.findall(r'\b\w+\b', content)
            
            for word in words:
                if len(word) > 3 and word not in stop_words:  # Only meaningful words
                    keyword_sentiment[word]['count'] += 1
                    keyword_sentiment[word]['total_sentiment'] += sentiment
                    if label == 'positive':
                        keyword_sentiment[word]['positive'] += 1
                    elif label == 'negative':
                        keyword_sentiment[word]['negative'] += 1
        
        # Calculate metrics and filter by frequency
        keyword_analysis = {}
        for word, data_dict in keyword_sentiment.items():
            if data_dict['count'] >= 3:  # Minimum frequency
                keyword_analysis[word] = {
                    'frequency': data_dict['count'],
                    'avg_sentiment': round(data_dict['total_sentiment'] / data_dict['count'], 3),
                    'positive_pct': round((data_dict['positive'] / data_dict['count'] * 100), 1),
                    'negative_pct': round((data_dict['negative'] / data_dict['count'] * 100), 1)
                }
        
        # Sort and get top N
        top_keywords = sorted(keyword_analysis.items(), key=lambda x: x[1]['frequency'], reverse=True)[:top_n]
        
        # Categorize keywords
        positive_keywords = [kw for kw, data in top_keywords if data['avg_sentiment'] > 0.2]
        negative_keywords = [kw for kw, data in top_keywords if data['avg_sentiment'] < -0.2]
        
        return {
            'success': True,
            'top_keywords': dict(top_keywords),
            'positive_keywords': positive_keywords[:10],
            'negative_keywords': negative_keywords[:10],
            'total_unique_keywords': len(keyword_analysis)
        }
    
    def calculate_sentiment_velocity(self, product_service: str = "", days: int = 7) -> Dict:
        """Calculate how quickly sentiment is changing (velocity)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get recent data
        recent_date = datetime.now() - timedelta(days=days)
        
        query = '''
            SELECT date, sentiment_score
            FROM sentiment_data
            WHERE date >= ? AND date <= ?
        '''
        params = [recent_date.isoformat(), datetime.now().isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' ORDER BY date'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if len(data) < 2:
            return {'success': False, 'error': 'Insufficient data for velocity calculation'}
        
        # Calculate daily averages
        daily_avg = defaultdict(list)
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                date_key = date_obj.date().isoformat()
                daily_avg[date_key].append(row[1])
            except:
                continue
        
        daily_sentiments = {}
        for date_key, scores in daily_avg.items():
            daily_sentiments[date_key] = statistics.mean(scores)
        
        # Calculate velocity (rate of change)
        dates = sorted(daily_sentiments.keys())
        if len(dates) < 2:
            return {'success': False, 'error': 'Need at least 2 days of data'}
        
        velocities = []
        for i in range(1, len(dates)):
            days_diff = (datetime.fromisoformat(dates[i]) - datetime.fromisoformat(dates[i-1])).days
            if days_diff > 0:
                sentiment_diff = daily_sentiments[dates[i]] - daily_sentiments[dates[i-1]]
                velocity = sentiment_diff / days_diff
                velocities.append(velocity)
        
        if not velocities:
            return {'success': False, 'error': 'Could not calculate velocity'}
        
        avg_velocity = statistics.mean(velocities)
        
        # Interpret velocity
        if abs(avg_velocity) > 0.05:
            direction = 'rapidly improving' if avg_velocity > 0 else 'rapidly declining'
        elif abs(avg_velocity) > 0.02:
            direction = 'moderately improving' if avg_velocity > 0 else 'moderately declining'
        else:
            direction = 'stable'
        
        return {
            'success': True,
            'velocity': round(avg_velocity, 4),
            'direction': direction,
            'period_days': days,
            'data_points': len(data),
            'interpretation': f"Sentiment is {direction} ({avg_velocity:+.4f} per day)"
        }
    
    def analyze_sentiment_clusters(self, product_service: str = "", n_clusters: int = 5) -> Dict:
        """Cluster similar sentiment patterns together"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT sentiment_score, engagement, 
                   CASE WHEN sentiment_label = 'positive' THEN 1 
                        WHEN sentiment_label = 'negative' THEN -1 
                        ELSE 0 END as label_numeric
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data or len(data) < n_clusters:
            return {'success': False, 'error': 'Insufficient data for clustering'}
        
        # Simple clustering based on sentiment and engagement
        # Group into clusters
        clusters = defaultdict(list)
        
        for row in data:
            sentiment = row[0]
            engagement = row[2] or 0
            
            # Normalize engagement (log scale)
            norm_engagement = math.log10(engagement + 1) if engagement > 0 else 0
            
            # Simple clustering: combine sentiment and engagement
            cluster_score = (sentiment * 0.7) + (norm_engagement * 0.3)
            
            # Assign to cluster
            cluster_id = min(n_clusters - 1, max(0, int((cluster_score + 1) * n_clusters / 2)))
            clusters[cluster_id].append({
                'sentiment': sentiment,
                'engagement': engagement
            })
        
        # Analyze clusters
        cluster_analysis = {}
        for cluster_id, items in clusters.items():
            if items:
                avg_sentiment = statistics.mean([item['sentiment'] for item in items])
                avg_engagement = statistics.mean([item['engagement'] for item in items])
                
                cluster_analysis[cluster_id] = {
                    'size': len(items),
                    'avg_sentiment': round(avg_sentiment, 3),
                    'avg_engagement': round(avg_engagement, 1),
                    'percentage': round((len(items) / len(data) * 100), 1)
                }
        
        # Identify cluster characteristics
        cluster_labels = {}
        for cluster_id, data_dict in cluster_analysis.items():
            sentiment = data_dict['avg_sentiment']
            engagement = data_dict['avg_engagement']
            
            if sentiment > 0.3 and engagement > 1000:
                label = 'High Positive Engagement'
            elif sentiment > 0.3:
                label = 'Positive Low Engagement'
            elif sentiment < -0.3 and engagement > 1000:
                label = 'High Negative Engagement'
            elif sentiment < -0.3:
                label = 'Negative Low Engagement'
            else:
                label = 'Neutral'
            
            cluster_labels[cluster_id] = label
        
        return {
            'success': True,
            'clusters': cluster_analysis,
            'cluster_labels': cluster_labels,
            'total_items': len(data),
            'n_clusters': n_clusters
        }
    
    def generate_comparative_report(self, period1_days: int = 90, period2_days: int = 90, 
                                   product_service: str = "") -> Dict:
        """Compare sentiment between two time periods"""
        now = datetime.now()
        period1_start = now - timedelta(days=period1_days + period2_days)
        period1_end = now - timedelta(days=period2_days)
        period2_start = period1_end
        period2_end = now
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Period 1
        query1 = '''
            SELECT AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count,
                   AVG(engagement) as avg_engagement
            FROM sentiment_data
            WHERE date >= ? AND date < ?
        '''
        params1 = [period1_start.isoformat(), period1_end.isoformat()]
        
        if product_service:
            query1 += ' AND product_service = ?'
            params1.append(product_service)
        
        cursor.execute(query1, params1)
        period1_data = cursor.fetchone()
        
        # Period 2
        query2 = query1.replace('date >= ? AND date < ?', 'date >= ? AND date <= ?')
        params2 = [period2_start.isoformat(), period2_end.isoformat()]
        
        if product_service:
            query2 += ' AND product_service = ?'
            params2.append(product_service)
        
        cursor.execute(query2, params2)
        period2_data = cursor.fetchone()
        conn.close()
        
        if not period1_data or not period2_data or period1_data[1] == 0 or period2_data[1] == 0:
            return {'success': False, 'error': 'Insufficient data for comparison'}
        
        # Calculate metrics
        period1 = {
            'avg_sentiment': round(period1_data[0] or 0, 3),
            'count': period1_data[1],
            'positive_count': period1_data[2],
            'negative_count': period1_data[3],
            'positive_pct': round((period1_data[2] / period1_data[1] * 100), 1),
            'negative_pct': round((period1_data[3] / period1_data[1] * 100), 1),
            'avg_engagement': round(period1_data[4] or 0, 1)
        }
        
        period2 = {
            'avg_sentiment': round(period2_data[0] or 0, 3),
            'count': period2_data[1],
            'positive_count': period2_data[2],
            'negative_count': period2_data[3],
            'positive_pct': round((period2_data[2] / period2_data[1] * 100), 1),
            'negative_pct': round((period2_data[3] / period2_data[1] * 100), 1),
            'avg_engagement': round(period2_data[4] or 0, 1)
        }
        
        # Calculate changes
        changes = {
            'sentiment_change': round(period2['avg_sentiment'] - period1['avg_sentiment'], 3),
            'count_change': period2['count'] - period1['count'],
            'count_change_pct': round(((period2['count'] - period1['count']) / period1['count'] * 100), 1) if period1['count'] > 0 else 0,
            'positive_pct_change': round(period2['positive_pct'] - period1['positive_pct'], 1),
            'negative_pct_change': round(period2['negative_pct'] - period1['negative_pct'], 1),
            'engagement_change': round(period2['avg_engagement'] - period1['avg_engagement'], 1)
        }
        
        # Determine trend
        if changes['sentiment_change'] > 0.1:
            trend = 'significantly improved'
        elif changes['sentiment_change'] > 0.05:
            trend = 'improved'
        elif changes['sentiment_change'] < -0.1:
            trend = 'significantly declined'
        elif changes['sentiment_change'] < -0.05:
            trend = 'declined'
        else:
            trend = 'remained stable'
        
        return {
            'success': True,
            'period1': {
                'label': f'Previous {period1_days} days',
                'data': period1
            },
            'period2': {
                'label': f'Recent {period2_days} days',
                'data': period2
            },
            'changes': changes,
            'trend': trend,
            'summary': f"Sentiment has {trend} from {period1['avg_sentiment']:.3f} to {period2['avg_sentiment']:.3f}"
        }
    
    def analyze_sentiment_by_event_campaign(self, event_name: str, days_before: int = 7, 
                                           days_after: int = 7, product_service: str = "") -> Dict:
        """Analyze sentiment around a specific event or campaign"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Find event date (search in content for event mentions)
        query = '''
            SELECT MIN(date) as event_date
            FROM sentiment_data
            WHERE content LIKE ? AND date >= ?
        '''
        params = [f'%{event_name}%', (datetime.now() - timedelta(days=180)).isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        
        if not result or not result[0]:
            conn.close()
            return {'success': False, 'error': f'Event "{event_name}" not found in data'}
        
        event_date = datetime.fromisoformat(result[0])
        before_start = event_date - timedelta(days=days_before)
        after_end = event_date + timedelta(days=days_after)
        
        # Before event
        query_before = '''
            SELECT AVG(sentiment_score) as avg_sentiment, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ? AND date < ?
        '''
        params_before = [before_start.isoformat(), event_date.isoformat()]
        
        if product_service:
            query_before += ' AND product_service = ?'
            params_before.append(product_service)
        
        cursor.execute(query_before, params_before)
        before_data = cursor.fetchone()
        
        # During event (day of)
        query_during = '''
            SELECT AVG(sentiment_score) as avg_sentiment, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ? AND date < ?
        '''
        during_start = event_date.replace(hour=0, minute=0, second=0)
        during_end = during_start + timedelta(days=1)
        params_during = [during_start.isoformat(), during_end.isoformat()]
        
        if product_service:
            query_during += ' AND product_service = ?'
            params_during.append(product_service)
        
        cursor.execute(query_during, params_during)
        during_data = cursor.fetchone()
        
        # After event
        query_after = '''
            SELECT AVG(sentiment_score) as avg_sentiment, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ? AND date <= ?
        '''
        params_after = [event_date.isoformat(), after_end.isoformat()]
        
        if product_service:
            query_after += ' AND product_service = ?'
            params_after.append(product_service)
        
        cursor.execute(query_after, params_after)
        after_data = cursor.fetchone()
        conn.close()
        
        # Calculate impact
        before_sentiment = before_data[0] if before_data and before_data[0] else 0
        during_sentiment = during_data[0] if during_data and during_data[0] else 0
        after_sentiment = after_data[0] if after_data and after_data[0] else 0
        
        impact = after_sentiment - before_sentiment
        
        return {
            'success': True,
            'event_name': event_name,
            'event_date': event_date.isoformat(),
            'before_event': {
                'avg_sentiment': round(before_sentiment, 3),
                'count': before_data[1] if before_data else 0,
                'period': f'{days_before} days before'
            },
            'during_event': {
                'avg_sentiment': round(during_sentiment, 3),
                'count': during_data[1] if during_data else 0,
                'period': 'Day of event'
            },
            'after_event': {
                'avg_sentiment': round(after_sentiment, 3),
                'count': after_data[1] if after_data else 0,
                'period': f'{days_after} days after'
            },
            'impact': {
                'sentiment_change': round(impact, 3),
                'interpretation': 'positive' if impact > 0.1 else 'negative' if impact < -0.1 else 'neutral'
            }
        }
    
    def calculate_sentiment_health_score(self, product_service: str = "") -> Dict:
        """Calculate overall sentiment health score (0-100)"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for health score'}
        
        health_score = 100.0
        factors = []
        
        # Factor 1: Current sentiment (40% weight)
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            current_sentiment = trends['monthly_metrics'][last_month]['avg_sentiment']
            # Convert -1 to 1 scale to 0-100
            sentiment_score = ((current_sentiment + 1) / 2) * 100
            health_score = (health_score * 0.6) + (sentiment_score * 0.4)
            factors.append({
                'factor': 'current_sentiment',
                'score': round(sentiment_score, 1),
                'weight': 0.4
            })
        
        # Factor 2: Trend direction (20% weight)
        trend_direction = trends.get('overall_trend', {}).get('direction', 'stable')
        if trend_direction == 'improving':
            trend_score = 100
        elif trend_direction == 'declining':
            trend_score = 50
        else:
            trend_score = 75
        
        health_score = (health_score * 0.8) + (trend_score * 0.2)
        factors.append({
            'factor': 'trend_direction',
            'score': trend_score,
            'weight': 0.2
        })
        
        # Factor 3: Risk level (20% weight)
        if risk.get('success'):
            risk_score_value = risk.get('risk_score', 0)
            risk_penalty = risk_score_value  # Subtract risk from health
            health_score -= (risk_penalty * 0.2)
            factors.append({
                'factor': 'risk_level',
                'score': round(100 - risk_penalty, 1),
                'weight': 0.2
            })
        
        # Factor 4: Positive percentage (20% weight)
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            metrics = trends['monthly_metrics'][last_month]
            positive_pct = (metrics['positive_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            health_score = (health_score * 0.8) + (positive_pct * 0.2)
            factors.append({
                'factor': 'positive_percentage',
                'score': round(positive_pct, 1),
                'weight': 0.2
            })
        
        # Clamp to 0-100
        health_score = max(0, min(100, health_score))
        
        # Determine health level
        if health_score >= 80:
            health_level = 'excellent'
        elif health_score >= 65:
            health_level = 'good'
        elif health_score >= 50:
            health_level = 'fair'
        elif health_score >= 35:
            health_level = 'poor'
        else:
            health_level = 'critical'
        
        return {
            'success': True,
            'health_score': round(health_score, 1),
            'health_level': health_level,
            'factors': factors,
            'recommendations': self._generate_health_recommendations(health_score, health_level)
        }
    
    def _generate_health_recommendations(self, health_score: float, health_level: str) -> List[str]:
        """Generate recommendations based on health score"""
        recommendations = []
        
        if health_level == 'critical':
            recommendations.append("URGENT: Immediate action required - sentiment health is critical")
            recommendations.append("Implement crisis management protocol")
            recommendations.append("Engage directly with negative feedback")
        elif health_level == 'poor':
            recommendations.append("Focus on improving customer satisfaction")
            recommendations.append("Address recurring negative themes")
            recommendations.append("Increase positive engagement activities")
        elif health_level == 'fair':
            recommendations.append("Monitor trends closely")
            recommendations.append("Proactively address emerging issues")
            recommendations.append("Strengthen positive messaging")
        elif health_level == 'good':
            recommendations.append("Maintain current positive trends")
            recommendations.append("Continue monitoring for early warning signs")
        else:  # excellent
            recommendations.append("Excellent sentiment health - maintain current strategies")
            recommendations.append("Leverage positive sentiment for marketing")
        
        return recommendations
    
    def analyze_sentiment_by_competitor_mention(self, competitor_name: str, 
                                               product_service: str = "") -> Dict:
        """Analyze sentiment when competitor is mentioned"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Posts mentioning competitor
        query_competitor = '''
            SELECT AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count
            FROM sentiment_data
            WHERE date >= ? AND competitor_mention = ?
        '''
        params_comp = [six_months_ago.isoformat(), competitor_name]
        
        if product_service:
            query_competitor += ' AND product_service = ?'
            params_comp.append(product_service)
        
        cursor.execute(query_competitor, params_comp)
        competitor_data = cursor.fetchone()
        
        # Posts without competitor mention
        query_no_competitor = '''
            SELECT AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count
            FROM sentiment_data
            WHERE date >= ? AND (competitor_mention IS NULL OR competitor_mention != ?)
        '''
        params_no_comp = [six_months_ago.isoformat(), competitor_name]
        
        if product_service:
            query_no_competitor += ' AND product_service = ?'
            params_no_comp.append(product_service)
        
        cursor.execute(query_no_competitor, params_no_comp)
        no_competitor_data = cursor.fetchone()
        conn.close()
        
        if not competitor_data or competitor_data[1] == 0:
            return {'success': False, 'error': f'No mentions of competitor "{competitor_name}" found'}
        
        competitor_metrics = {
            'avg_sentiment': round(competitor_data[0] or 0, 3),
            'count': competitor_data[1],
            'positive_pct': round((competitor_data[2] / competitor_data[1] * 100), 1) if competitor_data[1] > 0 else 0,
            'negative_pct': round((competitor_data[3] / competitor_data[1] * 100), 1) if competitor_data[1] > 0 else 0
        }
        
        no_competitor_metrics = {
            'avg_sentiment': round(no_competitor_data[0] or 0, 3) if no_competitor_data else 0,
            'count': no_competitor_data[1] if no_competitor_data else 0,
            'positive_pct': round((no_competitor_data[2] / no_competitor_data[1] * 100), 1) if no_competitor_data and no_competitor_data[1] > 0 else 0,
            'negative_pct': round((no_competitor_data[3] / no_competitor_data[1] * 100), 1) if no_competitor_data and no_competitor_data[1] > 0 else 0
        }
        
        # Calculate difference
        sentiment_diff = competitor_metrics['avg_sentiment'] - no_competitor_metrics['avg_sentiment']
        
        return {
            'success': True,
            'competitor_name': competitor_name,
            'with_competitor_mention': competitor_metrics,
            'without_competitor_mention': no_competitor_metrics,
            'difference': {
                'sentiment_diff': round(sentiment_diff, 3),
                'interpretation': 'better' if sentiment_diff > 0.1 else 'worse' if sentiment_diff < -0.1 else 'similar'
            }
        }
    
    def generate_benchmark_report(self, product_service: str = "") -> Dict:
        """Generate benchmark report comparing against industry standards"""
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for benchmarking'}
        
        # Industry benchmarks (configurable)
        industry_benchmarks = {
            'positive_sentiment_pct': 60.0,  # Industry average: 60% positive
            'negative_sentiment_pct': 20.0,  # Industry average: 20% negative
            'avg_sentiment_score': 0.3,  # Industry average sentiment score
            'engagement_rate': 5.0,  # Industry average engagement rate
            'response_time_hours': 24.0  # Industry average response time
        }
        
        # Get current metrics
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            current_metrics = trends['monthly_metrics'][last_month]
            
            current_positive_pct = (current_metrics['positive_count'] / current_metrics['total_count'] * 100) if current_metrics['total_count'] > 0 else 0
            current_negative_pct = (current_metrics['negative_count'] / current_metrics['total_count'] * 100) if current_metrics['total_count'] > 0 else 0
            
            # Compare against benchmarks
            benchmarks = {
                'positive_sentiment': {
                    'current': round(current_positive_pct, 1),
                    'industry_avg': industry_benchmarks['positive_sentiment_pct'],
                    'difference': round(current_positive_pct - industry_benchmarks['positive_sentiment_pct'], 1),
                    'status': 'above' if current_positive_pct > industry_benchmarks['positive_sentiment_pct'] else 'below'
                },
                'negative_sentiment': {
                    'current': round(current_negative_pct, 1),
                    'industry_avg': industry_benchmarks['negative_sentiment_pct'],
                    'difference': round(current_negative_pct - industry_benchmarks['negative_sentiment_pct'], 1),
                    'status': 'above' if current_negative_pct > industry_benchmarks['negative_sentiment_pct'] else 'below'
                },
                'avg_sentiment': {
                    'current': round(current_metrics['avg_sentiment'], 3),
                    'industry_avg': industry_benchmarks['avg_sentiment_score'],
                    'difference': round(current_metrics['avg_sentiment'] - industry_benchmarks['avg_sentiment_score'], 3),
                    'status': 'above' if current_metrics['avg_sentiment'] > industry_benchmarks['avg_sentiment_score'] else 'below'
                }
            }
            
            # Calculate overall benchmark score
            score_components = []
            if benchmarks['positive_sentiment']['status'] == 'above':
                score_components.append(25)
            else:
                score_components.append(15)
            
            if benchmarks['negative_sentiment']['status'] == 'below':
                score_components.append(25)
            else:
                score_components.append(15)
            
            if benchmarks['avg_sentiment']['status'] == 'above':
                score_components.append(50)
            else:
                score_components.append(30)
            
            benchmark_score = sum(score_components)
            
            return {
                'success': True,
                'benchmark_score': benchmark_score,
                'benchmarks': benchmarks,
                'overall_status': 'exceeding' if benchmark_score >= 80 else 'meeting' if benchmark_score >= 60 else 'below',
                'recommendations': self._generate_benchmark_recommendations(benchmarks)
            }
        
        return {'success': False, 'error': 'Could not calculate benchmarks'}
    
    def _generate_benchmark_recommendations(self, benchmarks: Dict) -> List[str]:
        """Generate recommendations based on benchmark comparison"""
        recommendations = []
        
        if benchmarks['positive_sentiment']['status'] == 'below':
            recommendations.append(f"Improve positive sentiment: Currently {benchmarks['positive_sentiment']['difference']:.1f}% below industry average")
        
        if benchmarks['negative_sentiment']['status'] == 'above':
            recommendations.append(f"Reduce negative sentiment: Currently {benchmarks['negative_sentiment']['difference']:.1f}% above industry average")
        
        if benchmarks['avg_sentiment']['status'] == 'below':
            recommendations.append(f"Improve average sentiment: Currently {benchmarks['avg_sentiment']['difference']:.3f} below industry average")
        
        if all(b['status'] == 'above' for b in benchmarks.values() if 'status' in b):
            recommendations.append("Excellent performance - exceeding industry benchmarks across all metrics")
        
        return recommendations
    
    def analyze_sentiment_momentum(self, product_service: str = "", window_days: int = 30) -> Dict:
        """Analyze sentiment momentum (acceleration of change)"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=window_days)
        
        query = '''
            SELECT date, AVG(sentiment_score) as daily_avg
            FROM sentiment_data
            WHERE date >= ? AND date <= ?
        '''
        params = [start_date.isoformat(), end_date.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY DATE(date) ORDER BY date'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if len(data) < 3:
            return {'success': False, 'error': 'Insufficient data for momentum analysis'}
        
        # Calculate daily averages
        daily_sentiments = {}
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                date_key = date_obj.date().isoformat()
                daily_sentiments[date_key] = row[1]
            except:
                continue
        
        dates = sorted(daily_sentiments.keys())
        
        # Calculate velocity (first derivative)
        velocities = []
        for i in range(1, len(dates)):
            days_diff = (datetime.fromisoformat(dates[i]) - datetime.fromisoformat(dates[i-1])).days
            if days_diff > 0:
                sentiment_diff = daily_sentiments[dates[i]] - daily_sentiments[dates[i-1]]
                velocity = sentiment_diff / days_diff
                velocities.append(velocity)
        
        if len(velocities) < 2:
            return {'success': False, 'error': 'Insufficient data for momentum calculation'}
        
        # Calculate momentum (second derivative - acceleration)
        momentums = []
        for i in range(1, len(velocities)):
            momentum = velocities[i] - velocities[i-1]
            momentums.append(momentum)
        
        avg_velocity = statistics.mean(velocities)
        avg_momentum = statistics.mean(momentums) if momentums else 0
        
        # Interpret momentum
        if avg_momentum > 0.01:
            momentum_direction = 'accelerating positive'
        elif avg_momentum < -0.01:
            momentum_direction = 'accelerating negative'
        else:
            momentum_direction = 'stable momentum'
        
        return {
            'success': True,
            'velocity': round(avg_velocity, 4),
            'momentum': round(avg_momentum, 4),
            'momentum_direction': momentum_direction,
            'window_days': window_days,
            'interpretation': f"Sentiment is {momentum_direction} (velocity: {avg_velocity:+.4f}, momentum: {avg_momentum:+.4f})"
        }
    
    def generate_insights_summary(self, product_service: str = "") -> str:
        """Generate comprehensive insights summary"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        health = self.calculate_sentiment_health_score(product_service)
        recommendations = self.generate_automated_recommendations(product_service)
        
        summary = []
        summary.append("=" * 80)
        summary.append("COMPREHENSIVE SENTIMENT INSIGHTS SUMMARY")
        summary.append("=" * 80)
        summary.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary.append(f"Product/Service: {product_service or 'All Products'}")
        summary.append("")
        
        # Health Score
        if health.get('success'):
            summary.append("SENTIMENT HEALTH")
            summary.append("-" * 80)
            summary.append(f"Health Score: {health['health_score']}/100 ({health['health_level'].upper()})")
            summary.append("")
        
        # Risk Assessment
        if risk.get('success'):
            summary.append("RISK ASSESSMENT")
            summary.append("-" * 80)
            summary.append(f"Risk Score: {risk['risk_score']}/100 ({risk['risk_level'].upper()})")
            if risk.get('risk_factors'):
                summary.append("Key Risk Factors:")
                for factor in risk['risk_factors'][:3]:
                    summary.append(f"  • {factor.get('factor', '')}: {factor.get('score', 0):.2f}")
            summary.append("")
        
        # Trend Summary
        if trends.get('success'):
            summary.append("TREND SUMMARY")
            summary.append("-" * 80)
            overall_trend = trends.get('overall_trend', {})
            summary.append(f"Overall Direction: {overall_trend.get('direction', 'N/A').upper()}")
            summary.append(f"Sentiment Change: {overall_trend.get('first_month_avg', 0):.3f} → {overall_trend.get('last_month_avg', 0):.3f}")
            
            if trends.get('emerging_preferences'):
                summary.append("Emerging Preferences:")
                for pref in trends['emerging_preferences'][:3]:
                    summary.append(f"  • {pref.get('description', '')} (Growth: {pref.get('growth_rate', 0)}%)")
            
            if trends.get('dissatisfaction_trends'):
                summary.append("Dissatisfaction Trends:")
                for trend in trends['dissatisfaction_trends'][:2]:
                    if trend.get('type') == 'increasing_dissatisfaction':
                        summary.append(f"  • Increasing: {trend.get('current_level', 0)}% (up {trend.get('increase', 0)}%)")
            summary.append("")
        
        # Top Recommendations
        if recommendations.get('success'):
            summary.append("TOP RECOMMENDATIONS")
            summary.append("-" * 80)
            high_priority = recommendations.get('high_priority', [])
            if high_priority:
                for rec in high_priority[:3]:
                    summary.append(f"🔴 {rec.get('recommendation', '')}")
            
            all_recs = recommendations.get('recommendations', [])
            for rec in all_recs[:3]:
                summary.append(f"• {rec.get('recommendation', '')}")
            summary.append("")
        
        summary.append("=" * 80)
        
        return "\n".join(summary)
    
    def detect_crisis_situations(self, product_service: str = "", threshold: float = 0.3) -> Dict:
        """Detect potential crisis situations based on multiple indicators"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        alerts = self.get_active_alerts(severity='critical')
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for crisis detection'}
        
        crisis_indicators = []
        crisis_score = 0.0
        
        # Indicator 1: Rapid negative sentiment decline
        if trends.get('overall_trend', {}).get('direction') == 'declining':
            decline_magnitude = trends.get('overall_trend', {}).get('magnitude', 0)
            if decline_magnitude > 0.2:
                crisis_score += 30
                crisis_indicators.append({
                    'indicator': 'rapid_sentiment_decline',
                    'severity': 'high',
                    'description': f'Rapid sentiment decline of {decline_magnitude:.3f}',
                    'score': 30
                })
        
        # Indicator 2: High negative sentiment percentage
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            metrics = trends['monthly_metrics'][last_month]
            negative_pct = (metrics['negative_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            
            if negative_pct > threshold * 100:
                crisis_score += 25
                crisis_indicators.append({
                    'indicator': 'high_negative_percentage',
                    'severity': 'high',
                    'description': f'Negative sentiment at {negative_pct:.1f}% (threshold: {threshold*100}%)',
                    'score': 25
                })
        
        # Indicator 3: Critical alerts
        if alerts:
            crisis_score += min(25, len(alerts) * 5)
            crisis_indicators.append({
                'indicator': 'critical_alerts',
                'severity': 'critical',
                'description': f'{len(alerts)} critical alerts active',
                'score': min(25, len(alerts) * 5)
            })
        
        # Indicator 4: High risk score
        if risk.get('success') and risk.get('risk_score', 0) > 70:
            crisis_score += 20
            crisis_indicators.append({
                'indicator': 'high_risk_score',
                'severity': 'high',
                'description': f'Risk score at {risk.get("risk_score", 0)}/100',
                'score': 20
            })
        
        # Indicator 5: Viral negative content
        if trends.get('dissatisfaction_trends'):
            for trend in trends['dissatisfaction_trends']:
                if trend.get('type') == 'viral_negative_content':
                    crisis_score += 15
                    crisis_indicators.append({
                        'indicator': 'viral_negative',
                        'severity': 'high',
                        'description': 'Viral negative content detected',
                        'score': 15
                    })
                    break
        
        # Determine crisis level
        if crisis_score >= 70:
            crisis_level = 'critical'
            action_required = 'immediate'
        elif crisis_score >= 50:
            crisis_level = 'high'
            action_required = 'urgent'
        elif crisis_score >= 30:
            crisis_level = 'moderate'
            action_required = 'soon'
        else:
            crisis_level = 'low'
            action_required = 'monitor'
        
        return {
            'success': True,
            'crisis_score': round(crisis_score, 1),
            'crisis_level': crisis_level,
            'action_required': action_required,
            'indicators': crisis_indicators,
            'recommendations': self._generate_crisis_recommendations(crisis_level, crisis_score)
        }
    
    def _generate_crisis_recommendations(self, crisis_level: str, crisis_score: float) -> List[str]:
        """Generate crisis management recommendations"""
        recommendations = []
        
        if crisis_level == 'critical':
            recommendations.append("🚨 CRITICAL: Activate crisis management team immediately")
            recommendations.append("Issue public statement addressing concerns")
            recommendations.append("Engage directly with all negative feedback")
            recommendations.append("Monitor all channels 24/7")
            recommendations.append("Prepare escalation plan")
        elif crisis_level == 'high':
            recommendations.append("⚠️ HIGH: Increase monitoring frequency")
            recommendations.append("Proactively address negative sentiment")
            recommendations.append("Review and respond to all critical alerts")
            recommendations.append("Consider public communication")
        elif crisis_level == 'moderate':
            recommendations.append("📊 MODERATE: Monitor trends closely")
            recommendations.append("Address emerging issues proactively")
            recommendations.append("Increase positive engagement")
        else:
            recommendations.append("✅ LOW: Continue normal monitoring")
            recommendations.append("Maintain current engagement levels")
        
        return recommendations
    
    def analyze_sentiment_by_frequency(self, product_service: str = "") -> Dict:
        """Analyze how posting frequency affects sentiment"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get posting frequency by author
        query = '''
            SELECT author_id, COUNT(*) as post_count,
                   AVG(sentiment_score) as avg_sentiment,
                   AVG(engagement) as avg_engagement
            FROM sentiment_data
            WHERE date >= ? AND author_id IS NOT NULL
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY author_id'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No author data available'}
        
        # Categorize by posting frequency
        frequency_categories = {
            'very_active': {'min': 50, 'scores': [], 'count': 0, 'engagement': []},
            'active': {'min': 20, 'max': 49, 'scores': [], 'count': 0, 'engagement': []},
            'moderate': {'min': 10, 'max': 19, 'scores': [], 'count': 0, 'engagement': []},
            'occasional': {'min': 5, 'max': 9, 'scores': [], 'count': 0, 'engagement': []},
            'rare': {'max': 4, 'scores': [], 'count': 0, 'engagement': []}
        }
        
        for row in data:
            post_count = row[1]
            sentiment = row[2]
            engagement = row[3] or 0
            
            if post_count >= 50:
                cat = 'very_active'
            elif post_count >= 20:
                cat = 'active'
            elif post_count >= 10:
                cat = 'moderate'
            elif post_count >= 5:
                cat = 'occasional'
            else:
                cat = 'rare'
            
            frequency_categories[cat]['scores'].append(sentiment)
            frequency_categories[cat]['engagement'].append(engagement)
            frequency_categories[cat]['count'] += 1
        
        # Calculate metrics
        frequency_analysis = {}
        for cat, data_dict in frequency_categories.items():
            if data_dict['count'] > 0:
                frequency_analysis[cat] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1),
                    'author_count': data_dict['count']
                }
        
        return {
            'success': True,
            'frequency_analysis': frequency_analysis,
            'insights': self._generate_frequency_insights(frequency_analysis)
        }
    
    def _generate_frequency_insights(self, frequency_analysis: Dict) -> List[str]:
        """Generate insights from frequency analysis"""
        insights = []
        
        if not frequency_analysis:
            return insights
        
        # Find category with best sentiment
        if frequency_analysis:
            best_cat = max(frequency_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            insights.append(f"Best sentiment from {best_cat[0].replace('_', ' ')} posters: {best_cat[1]['avg_sentiment']:.3f}")
        
        return insights
    
    def calculate_content_quality_score(self, content: str, sentiment_score: float, 
                                       engagement: int, author_followers: int) -> Dict:
        """Calculate overall content quality score"""
        quality_score = 0.0
        factors = []
        
        # Factor 1: Sentiment (30%)
        sentiment_factor = ((sentiment_score + 1) / 2) * 100  # Convert -1 to 1 -> 0 to 100
        quality_score += sentiment_factor * 0.3
        factors.append({
            'factor': 'sentiment',
            'score': round(sentiment_factor, 1),
            'weight': 0.3
        })
        
        # Factor 2: Engagement (25%)
        if author_followers > 0:
            engagement_rate = (engagement / author_followers) * 100
            engagement_score = min(100, engagement_rate * 10)  # 10% engagement = 100
        else:
            engagement_score = min(100, math.log10(engagement + 1) * 20)
        
        quality_score += engagement_score * 0.25
        factors.append({
            'factor': 'engagement',
            'score': round(engagement_score, 1),
            'weight': 0.25
        })
        
        # Factor 3: Content length (15%)
        content_length = len(content)
        if 100 <= content_length <= 280:
            length_score = 100
        elif 50 <= content_length < 100 or 280 < content_length <= 500:
            length_score = 80
        else:
            length_score = 60
        
        quality_score += length_score * 0.15
        factors.append({
            'factor': 'content_length',
            'score': length_score,
            'weight': 0.15
        })
        
        # Factor 4: Readability (15%)
        # Simple readability: check for proper punctuation, capitalization
        sentences = content.split('.')
        if len(sentences) > 1:
            readability_score = min(100, len(sentences) * 10)
        else:
            readability_score = 70
        
        quality_score += readability_score * 0.15
        factors.append({
            'factor': 'readability',
            'score': readability_score,
            'weight': 0.15
        })
        
        # Factor 5: Uniqueness (15%)
        # Check for repeated words (simplified)
        words = content.lower().split()
        unique_words = len(set(words))
        total_words = len(words)
        uniqueness = (unique_words / total_words * 100) if total_words > 0 else 0
        uniqueness_score = min(100, uniqueness * 1.5)
        
        quality_score += uniqueness_score * 0.15
        factors.append({
            'factor': 'uniqueness',
            'score': round(uniqueness_score, 1),
            'weight': 0.15
        })
        
        # Determine quality level
        if quality_score >= 80:
            quality_level = 'excellent'
        elif quality_score >= 65:
            quality_level = 'good'
        elif quality_score >= 50:
            quality_level = 'fair'
        else:
            quality_level = 'poor'
        
        return {
            'quality_score': round(quality_score, 1),
            'quality_level': quality_level,
            'factors': factors
        }
    
    def analyze_sentiment_by_interaction_type(self, product_service: str = "") -> Dict:
        """Analyze sentiment by type of interaction (post, comment, review, etc.)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Infer interaction type from source/platform patterns
        query = '''
            SELECT source, platform, sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Categorize by interaction type (inferred from source/platform)
        interaction_types = defaultdict(lambda: {'scores': [], 'labels': [], 'engagement': [], 'count': 0})
        
        for row in data:
            source = (row[0] or "").lower()
            platform = (row[1] or "").lower()
            sentiment = row[2]
            label = row[3]
            engagement = row[4] or 0
            
            # Infer interaction type
            if 'review' in source or 'review' in platform:
                interaction_type = 'review'
            elif 'comment' in source:
                interaction_type = 'comment'
            elif 'tweet' in source or platform == 'twitter':
                interaction_type = 'post'
            elif 'reddit' in platform:
                interaction_type = 'discussion'
            else:
                interaction_type = 'general'
            
            interaction_types[interaction_type]['scores'].append(sentiment)
            interaction_types[interaction_type]['labels'].append(label)
            interaction_types[interaction_type]['engagement'].append(engagement)
            interaction_types[interaction_type]['count'] += 1
        
        # Calculate metrics
        type_analysis = {}
        for interaction_type, data_dict in interaction_types.items():
            if data_dict['count'] > 0:
                type_analysis[interaction_type] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'positive_pct': round((data_dict['labels'].count('positive') / len(data_dict['labels']) * 100), 1),
                    'negative_pct': round((data_dict['labels'].count('negative') / len(data_dict['labels']) * 100), 1),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1),
                    'count': data_dict['count']
                }
        
        # Find best interaction type
        if type_analysis:
            best_type = max(type_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'interaction_type_analysis': type_analysis,
                'best_interaction_type': {
                    'type': best_type[0],
                    'sentiment': best_type[1]['avg_sentiment'],
                    'count': best_type[1]['count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient data'}
    
    def predict_churn_risk(self, product_service: str = "") -> Dict:
        """Predict customer churn risk based on sentiment patterns"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for churn prediction'}
        
        churn_risk_score = 0.0
        risk_factors = []
        
        # Factor 1: Declining sentiment trend (40%)
        if trends.get('overall_trend', {}).get('direction') == 'declining':
            decline_magnitude = trends.get('overall_trend', {}).get('magnitude', 0)
            churn_risk = min(100, decline_magnitude * 200)  # Scale to 0-100
            churn_risk_score += churn_risk * 0.4
            risk_factors.append({
                'factor': 'declining_sentiment',
                'contribution': round(churn_risk * 0.4, 1),
                'weight': 0.4
            })
        
        # Factor 2: High negative sentiment (30%)
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            metrics = trends['monthly_metrics'][last_month]
            negative_pct = (metrics['negative_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            
            if negative_pct > 30:
                churn_risk = min(100, (negative_pct - 30) * 2)  # 30% = 0, 80% = 100
                churn_risk_score += churn_risk * 0.3
                risk_factors.append({
                    'factor': 'high_negative_sentiment',
                    'contribution': round(churn_risk * 0.3, 1),
                    'weight': 0.3
                })
        
        # Factor 3: Increasing dissatisfaction (20%)
        if trends.get('dissatisfaction_trends'):
            for trend in trends['dissatisfaction_trends']:
                if trend.get('type') == 'increasing_dissatisfaction':
                    increase = trend.get('increase', 0)
                    churn_risk = min(100, increase * 5)  # 20% increase = 100
                    churn_risk_score += churn_risk * 0.2
                    risk_factors.append({
                        'factor': 'increasing_dissatisfaction',
                        'contribution': round(churn_risk * 0.2, 1),
                        'weight': 0.2
                    })
                    break
        
        # Factor 4: High risk score (10%)
        if risk.get('success'):
            risk_score = risk.get('risk_score', 0)
            churn_risk_score += risk_score * 0.1
            risk_factors.append({
                'factor': 'overall_risk',
                'contribution': round(risk_score * 0.1, 1),
                'weight': 0.1
            })
        
        # Clamp to 0-100
        churn_risk_score = max(0, min(100, churn_risk_score))
        
        # Determine churn risk level
        if churn_risk_score >= 70:
            churn_level = 'critical'
            action = 'immediate_intervention'
        elif churn_risk_score >= 50:
            churn_level = 'high'
            action = 'proactive_engagement'
        elif churn_risk_score >= 30:
            churn_level = 'moderate'
            action = 'monitor_closely'
        else:
            churn_level = 'low'
            action = 'maintain_engagement'
        
        return {
            'success': True,
            'churn_risk_score': round(churn_risk_score, 1),
            'churn_risk_level': churn_level,
            'recommended_action': action,
            'risk_factors': risk_factors,
            'recommendations': self._generate_churn_recommendations(churn_level, churn_risk_score)
        }
    
    def _generate_churn_recommendations(self, churn_level: str, churn_score: float) -> List[str]:
        """Generate recommendations to prevent churn"""
        recommendations = []
        
        if churn_level == 'critical':
            recommendations.append("🚨 CRITICAL: High churn risk - immediate customer retention actions needed")
            recommendations.append("Reach out to at-risk customers directly")
            recommendations.append("Offer personalized solutions to address concerns")
            recommendations.append("Implement win-back campaigns")
        elif churn_level == 'high':
            recommendations.append("⚠️ HIGH: Proactive engagement required")
            recommendations.append("Identify and address root causes of dissatisfaction")
            recommendations.append("Increase positive touchpoints")
            recommendations.append("Consider loyalty programs or incentives")
        elif churn_level == 'moderate':
            recommendations.append("📊 MODERATE: Monitor customer satisfaction closely")
            recommendations.append("Address emerging issues before they escalate")
            recommendations.append("Strengthen customer relationships")
        else:
            recommendations.append("✅ LOW: Maintain current engagement levels")
            recommendations.append("Continue monitoring for early warning signs")
        
        return recommendations
    
    def analyze_sentiment_by_value_perception(self, product_service: str = "") -> Dict:
        """Analyze sentiment related to value perception (price, quality, worth)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Search for value-related keywords
        value_keywords = ['price', 'cost', 'worth', 'value', 'expensive', 'cheap', 'affordable', 
                         'quality', 'premium', 'budget', 'deal', 'overpriced', 'worth it']
        
        query = '''
            SELECT content, sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Categorize by value perception
        value_categories = {
            'positive_value': {'scores': [], 'count': 0, 'keywords': []},
            'negative_value': {'scores': [], 'count': 0, 'keywords': []},
            'neutral_value': {'scores': [], 'count': 0, 'keywords': []}
        }
        
        for row in data:
            content = (row[0] or "").lower()
            sentiment = row[2]
            
            # Check for value keywords
            found_keywords = [kw for kw in value_keywords if kw in content]
            
            if found_keywords:
                if sentiment == 'positive':
                    value_categories['positive_value']['scores'].append(row[1])
                    value_categories['positive_value']['count'] += 1
                    value_categories['positive_value']['keywords'].extend(found_keywords)
                elif sentiment == 'negative':
                    value_categories['negative_value']['scores'].append(row[1])
                    value_categories['negative_value']['count'] += 1
                    value_categories['negative_value']['keywords'].extend(found_keywords)
                else:
                    value_categories['neutral_value']['scores'].append(row[1])
                    value_categories['neutral_value']['count'] += 1
                    value_categories['neutral_value']['keywords'].extend(found_keywords)
        
        # Calculate metrics
        value_analysis = {}
        for category, data_dict in value_categories.items():
            if data_dict['count'] > 0:
                value_analysis[category] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'count': data_dict['count'],
                    'top_keywords': dict(Counter(data_dict['keywords']).most_common(5))
                }
        
        # Calculate overall value perception
        total_value_mentions = sum(cat['count'] for cat in value_analysis.values())
        if total_value_mentions > 0:
            positive_value_pct = (value_analysis.get('positive_value', {}).get('count', 0) / total_value_mentions * 100)
            negative_value_pct = (value_analysis.get('negative_value', {}).get('count', 0) / total_value_mentions * 100)
            
            if positive_value_pct > negative_value_pct + 20:
                perception = 'positive'
            elif negative_value_pct > positive_value_pct + 20:
                perception = 'negative'
            else:
                perception = 'mixed'
        else:
            perception = 'unknown'
        
        return {
            'success': True,
            'value_analysis': value_analysis,
            'overall_perception': perception,
            'total_value_mentions': total_value_mentions
        }
    
    def generate_action_plan(self, product_service: str = "") -> Dict:
        """Generate actionable plan based on all analyses"""
        trends = self.analyze_6_month_trends(product_service)
        risk = self.calculate_risk_score(product_service)
        health = self.calculate_sentiment_health_score(product_service)
        crisis = self.detect_crisis_situations(product_service)
        recommendations = self.generate_automated_recommendations(product_service)
        
        action_plan = {
            'success': True,
            'timestamp': datetime.now().isoformat(),
            'product_service': product_service or 'All Products',
            'priority_actions': [],
            'short_term_actions': [],
            'long_term_actions': [],
            'monitoring_actions': []
        }
        
        # Priority actions based on crisis/risk
        if crisis.get('success') and crisis.get('crisis_level') in ['critical', 'high']:
            action_plan['priority_actions'].extend([
                {
                    'action': 'Activate crisis management protocol',
                    'priority': 'critical',
                    'timeline': 'immediate',
                    'owner': 'Crisis Management Team'
                },
                {
                    'action': 'Review and respond to all critical alerts',
                    'priority': 'critical',
                    'timeline': 'within 24 hours',
                    'owner': 'Customer Service Team'
                }
            ])
        
        # Short-term actions (next 7 days)
        if risk.get('success') and risk.get('risk_level') in ['high', 'critical']:
            action_plan['short_term_actions'].extend([
                {
                    'action': 'Address recurring negative themes',
                    'priority': 'high',
                    'timeline': 'within 7 days',
                    'owner': 'Product/Service Team'
                },
                {
                    'action': 'Increase positive engagement activities',
                    'priority': 'high',
                    'timeline': 'within 7 days',
                    'owner': 'Marketing Team'
                }
            ])
        
        # Long-term actions (next 30-90 days)
        if trends.get('success'):
            if trends.get('emerging_preferences'):
                for pref in trends['emerging_preferences'][:2]:
                    action_plan['long_term_actions'].append({
                        'action': f"Consider enhancing {pref.get('topic', '').replace('_', ' ')} features",
                        'priority': 'medium',
                        'timeline': '30-90 days',
                        'owner': 'Product Development Team'
                    })
            
            if trends.get('overall_trend', {}).get('direction') == 'declining':
                action_plan['long_term_actions'].append({
                    'action': 'Investigate causes of sentiment decline',
                    'priority': 'high',
                    'timeline': '30 days',
                    'owner': 'Analytics Team'
                })
        
        # Monitoring actions
        action_plan['monitoring_actions'].extend([
            {
                'action': 'Monitor sentiment trends daily',
                'frequency': 'daily',
                'owner': 'Analytics Team'
            },
            {
                'action': 'Review alerts and recommendations weekly',
                'frequency': 'weekly',
                'owner': 'Management Team'
            }
        ])
        
        return action_plan
    
    def analyze_sentiment_by_problem_category(self, product_service: str = "") -> Dict:
        """Analyze sentiment by problem/solution categories"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Problem categories and keywords
        problem_categories = {
            'technical_issues': ['bug', 'error', 'crash', 'broken', 'not working', 'glitch', 'freeze', 'slow'],
            'pricing_concerns': ['expensive', 'price', 'cost', 'overpriced', 'cheap', 'affordable', 'worth'],
            'customer_service': ['support', 'service', 'help', 'response', 'contact', 'assistance', 'wait'],
            'feature_requests': ['need', 'want', 'should', 'missing', 'add', 'feature', 'improve', 'better'],
            'usability': ['difficult', 'confusing', 'hard', 'easy', 'simple', 'intuitive', 'complicated'],
            'performance': ['fast', 'slow', 'speed', 'lag', 'performance', 'responsive', 'loading'],
            'reliability': ['reliable', 'stable', 'trust', 'dependable', 'consistent', 'unstable']
        }
        
        category_analysis = defaultdict(lambda: {'scores': [], 'count': 0, 'positive': 0, 'negative': 0, 'engagement': []})
        
        for row in data:
            content = (row[0] or "").lower()
            sentiment = row[1]
            label = row[2]
            engagement = row[4] or 0
            
            # Check for problem categories
            for category, keywords in problem_categories.items():
                if any(keyword in content for keyword in keywords):
                    category_analysis[category]['scores'].append(sentiment)
                    category_analysis[category]['count'] += 1
                    category_analysis[category]['engagement'].append(engagement)
                    if label == 'positive':
                        category_analysis[category]['positive'] += 1
                    elif label == 'negative':
                        category_analysis[category]['negative'] += 1
        
        # Calculate metrics
        problem_analysis = {}
        for category, data_dict in category_analysis.items():
            if data_dict['count'] > 0:
                problem_analysis[category] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'count': data_dict['count'],
                    'positive_pct': round((data_dict['positive'] / data_dict['count'] * 100), 1),
                    'negative_pct': round((data_dict['negative'] / data_dict['count'] * 100), 1),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1)
                }
        
        # Find most problematic category
        if problem_analysis:
            most_problematic = min(problem_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'problem_analysis': problem_analysis,
                'most_problematic_category': {
                    'category': most_problematic[0],
                    'sentiment': most_problematic[1]['avg_sentiment'],
                    'count': most_problematic[1]['count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient problem category data'}
    
    def calculate_influence_score(self, author_id: str, product_service: str = "") -> Dict:
        """Calculate influence score for a specific author"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT COUNT(*) as post_count,
                   AVG(sentiment_score) as avg_sentiment,
                   AVG(engagement) as avg_engagement,
                   MAX(engagement) as max_engagement,
                   SUM(engagement) as total_engagement
            FROM sentiment_data
            WHERE author_id = ? AND date >= ?
        '''
        params = [author_id, six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchone()
        conn.close()
        
        if not data or data[0] == 0:
            return {'success': False, 'error': f'No data found for author {author_id}'}
        
        post_count = data[0]
        avg_sentiment = data[1] or 0
        avg_engagement = data[2] or 0
        max_engagement = data[3] or 0
        total_engagement = data[4] or 0
        
        # Calculate influence score (0-100)
        influence_score = 0.0
        
        # Factor 1: Post frequency (20%)
        frequency_score = min(100, post_count * 2)  # 50 posts = 100
        influence_score += frequency_score * 0.2
        
        # Factor 2: Average engagement (30%)
        engagement_score = min(100, math.log10(avg_engagement + 1) * 20)
        influence_score += engagement_score * 0.3
        
        # Factor 3: Maximum engagement (20%)
        max_engagement_score = min(100, math.log10(max_engagement + 1) * 15)
        influence_score += max_engagement_score * 0.2
        
        # Factor 4: Total engagement (20%)
        total_engagement_score = min(100, math.log10(total_engagement + 1) * 10)
        influence_score += total_engagement_score * 0.2
        
        # Factor 5: Sentiment consistency (10%)
        # Positive sentiment contributes to influence
        sentiment_factor = ((avg_sentiment + 1) / 2) * 100
        influence_score += sentiment_factor * 0.1
        
        influence_score = max(0, min(100, influence_score))
        
        # Determine influence level
        if influence_score >= 80:
            influence_level = 'high'
        elif influence_score >= 60:
            influence_level = 'medium'
        elif influence_score >= 40:
            influence_level = 'low'
        else:
            influence_level = 'minimal'
        
        return {
            'success': True,
            'author_id': author_id,
            'influence_score': round(influence_score, 1),
            'influence_level': influence_level,
            'metrics': {
                'post_count': post_count,
                'avg_sentiment': round(avg_sentiment, 3),
                'avg_engagement': round(avg_engagement, 1),
                'max_engagement': max_engagement,
                'total_engagement': total_engagement
            }
        }
    
    def analyze_sentiment_by_channel(self, product_service: str = "") -> Dict:
        """Analyze sentiment by communication channel"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT platform, source, sentiment_score, sentiment_label, engagement, COUNT(*) as count
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY platform, source'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Group by channel
        channel_analysis = defaultdict(lambda: {'scores': [], 'labels': [], 'engagement': [], 'count': 0})
        
        for row in data:
            platform = (row[0] or "").lower()
            source = (row[1] or "").lower()
            sentiment = row[2]
            label = row[3]
            engagement = row[4] or 0
            count = row[5]
            
            # Determine channel
            if 'twitter' in platform or 'tweet' in source:
                channel = 'twitter'
            elif 'reddit' in platform or 'reddit' in source:
                channel = 'reddit'
            elif 'facebook' in platform or 'facebook' in source:
                channel = 'facebook'
            elif 'instagram' in platform or 'instagram' in source:
                channel = 'instagram'
            elif 'linkedin' in platform or 'linkedin' in source:
                channel = 'linkedin'
            elif 'review' in source or 'review' in platform:
                channel = 'review_sites'
            elif 'email' in source:
                channel = 'email'
            else:
                channel = 'other'
            
            for _ in range(count):
                channel_analysis[channel]['scores'].append(sentiment)
                channel_analysis[channel]['labels'].append(label)
                channel_analysis[channel]['engagement'].append(engagement)
            channel_analysis[channel]['count'] += count
        
        # Calculate metrics
        channel_metrics = {}
        for channel, data_dict in channel_analysis.items():
            if data_dict['count'] > 0:
                channel_metrics[channel] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'positive_pct': round((data_dict['labels'].count('positive') / len(data_dict['labels']) * 100), 1),
                    'negative_pct': round((data_dict['labels'].count('negative') / len(data_dict['labels']) * 100), 1),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1),
                    'total_count': data_dict['count']
                }
        
        # Find best and worst channels
        if channel_metrics:
            best_channel = max(channel_metrics.items(), key=lambda x: x[1]['avg_sentiment'])
            worst_channel = min(channel_metrics.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'channel_analysis': channel_metrics,
                'best_channel': {
                    'channel': best_channel[0],
                    'sentiment': best_channel[1]['avg_sentiment'],
                    'count': best_channel[1]['total_count']
                },
                'worst_channel': {
                    'channel': worst_channel[0],
                    'sentiment': worst_channel[1]['avg_sentiment'],
                    'count': worst_channel[1]['total_count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient channel data'}
    
    def generate_real_time_dashboard_data(self, product_service: str = "", hours: int = 24) -> Dict:
        """Generate real-time dashboard data for the last N hours"""
        hours_ago = datetime.now() - timedelta(hours=hours)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT sentiment_score, sentiment_label, platform, engagement, date
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [hours_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No recent data available'}
        
        # Calculate real-time metrics
        total_count = len(data)
        positive_count = sum(1 for row in data if row[1] == 'positive')
        negative_count = sum(1 for row in data if row[1] == 'negative')
        neutral_count = total_count - positive_count - negative_count
        
        avg_sentiment = statistics.mean([row[0] for row in data])
        total_engagement = sum(row[3] or 0 for row in data)
        avg_engagement = total_engagement / total_count if total_count > 0 else 0
        
        # Hourly breakdown
        hourly_data = defaultdict(lambda: {'count': 0, 'sentiment': [], 'engagement': []})
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[4])
                hour_key = date_obj.strftime('%Y-%m-%d %H:00')
                hourly_data[hour_key]['count'] += 1
                hourly_data[hour_key]['sentiment'].append(row[0])
                hourly_data[hour_key]['engagement'].append(row[3] or 0)
            except:
                continue
        
        hourly_breakdown = {}
        for hour, data_dict in hourly_data.items():
            hourly_breakdown[hour] = {
                'count': data_dict['count'],
                'avg_sentiment': round(statistics.mean(data_dict['sentiment']), 3),
                'avg_engagement': round(statistics.mean(data_dict['engagement']), 1)
            }
        
        # Platform breakdown
        platform_data = defaultdict(lambda: {'count': 0, 'sentiment': []})
        for row in data:
            platform = row[2] or 'unknown'
            platform_data[platform]['count'] += 1
            platform_data[platform]['sentiment'].append(row[0])
        
        platform_breakdown = {}
        for platform, data_dict in platform_data.items():
            platform_breakdown[platform] = {
                'count': data_dict['count'],
                'avg_sentiment': round(statistics.mean(data_dict['sentiment']), 3)
            }
        
        return {
            'success': True,
            'time_period_hours': hours,
            'timestamp': datetime.now().isoformat(),
            'summary': {
                'total_mentions': total_count,
                'positive_count': positive_count,
                'negative_count': negative_count,
                'neutral_count': neutral_count,
                'positive_pct': round((positive_count / total_count * 100), 1) if total_count > 0 else 0,
                'negative_pct': round((negative_count / total_count * 100), 1) if total_count > 0 else 0,
                'avg_sentiment': round(avg_sentiment, 3),
                'total_engagement': total_engagement,
                'avg_engagement': round(avg_engagement, 1)
            },
            'hourly_breakdown': hourly_breakdown,
            'platform_breakdown': platform_breakdown
        }
    
    def compare_multiple_competitors(self, competitor_names: List[str], product_service: str = "") -> Dict:
        """Compare sentiment across multiple competitors"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        competitor_analysis = {}
        
        for competitor in competitor_names:
            query = '''
                SELECT AVG(sentiment_score) as avg_sentiment,
                       COUNT(*) as count,
                       SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                       SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count,
                       AVG(engagement) as avg_engagement
                FROM sentiment_data
                WHERE date >= ? AND competitor_mention = ?
            '''
            params = [six_months_ago.isoformat(), competitor]
            
            if product_service:
                query += ' AND product_service = ?'
                params.append(product_service)
            
            cursor.execute(query, params)
            data = cursor.fetchone()
            
            if data and data[1] > 0:
                competitor_analysis[competitor] = {
                    'avg_sentiment': round(data[0] or 0, 3),
                    'count': data[1],
                    'positive_pct': round((data[2] / data[1] * 100), 1),
                    'negative_pct': round((data[3] / data[1] * 100), 1),
                    'avg_engagement': round(data[4] or 0, 1)
                }
        
        conn.close()
        
        if not competitor_analysis:
            return {'success': False, 'error': 'No competitor data found'}
        
        # Find best and worst competitors
        best_competitor = max(competitor_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
        worst_competitor = min(competitor_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
        
        # Calculate rankings
        rankings = sorted(competitor_analysis.items(), key=lambda x: x[1]['avg_sentiment'], reverse=True)
        
        return {
            'success': True,
            'competitors': competitor_analysis,
            'rankings': [{'competitor': comp, 'sentiment': data['avg_sentiment'], 'rank': idx + 1} 
                         for idx, (comp, data) in enumerate(rankings)],
            'best_competitor': {
                'name': best_competitor[0],
                'sentiment': best_competitor[1]['avg_sentiment']
            },
            'worst_competitor': {
                'name': worst_competitor[0],
                'sentiment': worst_competitor[1]['avg_sentiment']
            }
        }
    
    def analyze_sentiment_trends_by_segment(self, segment_key: str, segment_value: str, 
                                           product_service: str = "") -> Dict:
        """Analyze sentiment trends for a specific segment (e.g., platform, language, region)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Build query based on segment
        query = f'''
            SELECT date, AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count
            FROM sentiment_data
            WHERE date >= ? AND {segment_key} = ?
        '''
        params = [six_months_ago.isoformat(), segment_value]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        query += ' GROUP BY DATE(date) ORDER BY date'
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': f'No data found for segment {segment_key}={segment_value}'}
        
        # Calculate monthly trends
        monthly_trends = defaultdict(lambda: {'scores': [], 'count': 0, 'positive': 0, 'negative': 0})
        
        for row in data:
            try:
                date_obj = datetime.fromisoformat(row[0])
                month_key = date_obj.strftime('%Y-%m')
                monthly_trends[month_key]['scores'].append(row[1])
                monthly_trends[month_key]['count'] += row[2]
                monthly_trends[month_key]['positive'] += row[3]
                monthly_trends[month_key]['negative'] += row[4]
            except:
                continue
        
        # Calculate monthly metrics
        monthly_metrics = {}
        for month, data_dict in monthly_trends.items():
            monthly_metrics[month] = {
                'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                'count': data_dict['count'],
                'positive_pct': round((data_dict['positive'] / data_dict['count'] * 100), 1) if data_dict['count'] > 0 else 0,
                'negative_pct': round((data_dict['negative'] / data_dict['count'] * 100), 1) if data_dict['count'] > 0 else 0
            }
        
        # Calculate overall trend
        if len(monthly_metrics) >= 2:
            months = sorted(monthly_metrics.keys())
            first_month_sentiment = monthly_metrics[months[0]]['avg_sentiment']
            last_month_sentiment = monthly_metrics[months[-1]]['avg_sentiment']
            trend_direction = 'improving' if last_month_sentiment > first_month_sentiment else 'declining' if last_month_sentiment < first_month_sentiment else 'stable'
            trend_magnitude = abs(last_month_sentiment - first_month_sentiment)
        else:
            trend_direction = 'insufficient_data'
            trend_magnitude = 0
        
        return {
            'success': True,
            'segment_key': segment_key,
            'segment_value': segment_value,
            'monthly_trends': monthly_metrics,
            'overall_trend': {
                'direction': trend_direction,
                'magnitude': round(trend_magnitude, 3)
            }
        }
    
    def calculate_customer_satisfaction_score(self, product_service: str = "") -> Dict:
        """Calculate customer satisfaction score (similar to NPS)"""
        trends = self.analyze_6_month_trends(product_service)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for satisfaction score'}
        
        # Get recent metrics
        if trends.get('monthly_metrics'):
            last_month = sorted(trends['monthly_metrics'].keys())[-1]
            metrics = trends['monthly_metrics'][last_month]
            
            # Calculate satisfaction score (0-100)
            # Based on positive sentiment percentage and overall sentiment
            positive_pct = (metrics['positive_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            negative_pct = (metrics['negative_count'] / metrics['total_count'] * 100) if metrics['total_count'] > 0 else 0
            
            # Convert sentiment score (-1 to 1) to 0-100 scale
            sentiment_score = ((metrics['avg_sentiment'] + 1) / 2) * 100
            
            # Calculate satisfaction: weighted average
            satisfaction_score = (positive_pct * 0.6) + (sentiment_score * 0.4)
            
            # Calculate NPS-like score (-100 to 100)
            # Promoters: positive sentiment > 0.3
            # Detractors: negative sentiment < -0.3
            # Passives: everything else
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            query = '''
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN sentiment_score > 0.3 THEN 1 ELSE 0 END) as promoters,
                       SUM(CASE WHEN sentiment_score < -0.3 THEN 1 ELSE 0 END) as detractors
                FROM sentiment_data
                WHERE date >= ?
            '''
            params = [(datetime.now() - timedelta(days=30)).isoformat()]
            
            if product_service:
                query += ' AND product_service = ?'
                params.append(product_service)
            
            cursor.execute(query, params)
            nps_data = cursor.fetchone()
            conn.close()
            
            if nps_data and nps_data[0] > 0:
                promoter_pct = (nps_data[1] / nps_data[0] * 100)
                detractor_pct = (nps_data[2] / nps_data[0] * 100)
                nps_score = promoter_pct - detractor_pct
            else:
                nps_score = 0
                promoter_pct = 0
                detractor_pct = 0
            
            # Determine satisfaction level
            if satisfaction_score >= 80:
                satisfaction_level = 'excellent'
            elif satisfaction_score >= 65:
                satisfaction_level = 'good'
            elif satisfaction_score >= 50:
                satisfaction_level = 'fair'
            elif satisfaction_score >= 35:
                satisfaction_level = 'poor'
            else:
                satisfaction_level = 'critical'
            
            return {
                'success': True,
                'satisfaction_score': round(satisfaction_score, 1),
                'satisfaction_level': satisfaction_level,
                'nps_score': round(nps_score, 1),
                'promoter_pct': round(promoter_pct, 1),
                'detractor_pct': round(detractor_pct, 1),
                'positive_pct': round(positive_pct, 1),
                'negative_pct': round(negative_pct, 1),
                'avg_sentiment': round(metrics['avg_sentiment'], 3)
            }
        
        return {'success': False, 'error': 'Could not calculate satisfaction score'}
    
    def analyze_sentiment_by_emotion_category(self, product_service: str = "") -> Dict:
        """Analyze sentiment by specific emotion categories"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, sentiment_score, sentiment_label, emotions
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Emotion categories
        emotion_categories = {
            'joy': ['happy', 'excited', 'love', 'great', 'amazing', 'wonderful', 'fantastic'],
            'anger': ['angry', 'furious', 'mad', 'frustrated', 'annoyed', 'irritated'],
            'sadness': ['sad', 'disappointed', 'unhappy', 'depressed', 'upset', 'let down'],
            'fear': ['worried', 'concerned', 'anxious', 'scared', 'nervous', 'afraid'],
            'surprise': ['surprised', 'shocked', 'amazed', 'unexpected', 'wow'],
            'trust': ['trust', 'confident', 'reliable', 'dependable', 'secure'],
            'disgust': ['disgusted', 'revolted', 'sick', 'gross', 'horrible']
        }
        
        emotion_analysis = defaultdict(lambda: {'count': 0, 'scores': [], 'positive': 0, 'negative': 0})
        
        for row in data:
            content = (row[0] or "").lower()
            sentiment = row[1]
            label = row[2]
            emotions_str = row[3] or ""
            
            # Check for emotion keywords
            for emotion, keywords in emotion_categories.items():
                if any(keyword in content for keyword in keywords):
                    emotion_analysis[emotion]['count'] += 1
                    emotion_analysis[emotion]['scores'].append(sentiment)
                    if label == 'positive':
                        emotion_analysis[emotion]['positive'] += 1
                    elif label == 'negative':
                        emotion_analysis[emotion]['negative'] += 1
        
        # Calculate metrics
        emotion_metrics = {}
        for emotion, data_dict in emotion_analysis.items():
            if data_dict['count'] > 0:
                emotion_metrics[emotion] = {
                    'count': data_dict['count'],
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'positive_pct': round((data_dict['positive'] / data_dict['count'] * 100), 1),
                    'negative_pct': round((data_dict['negative'] / data_dict['count'] * 100), 1)
                }
        
        # Find dominant emotion
        if emotion_metrics:
            dominant_emotion = max(emotion_metrics.items(), key=lambda x: x[1]['count'])
            
            return {
                'success': True,
                'emotion_analysis': emotion_metrics,
                'dominant_emotion': {
                    'emotion': dominant_emotion[0],
                    'count': dominant_emotion[1]['count'],
                    'avg_sentiment': dominant_emotion[1]['avg_sentiment']
                }
            }
        
        return {'success': False, 'error': 'Insufficient emotion data'}
    
    def generate_year_over_year_comparison(self, product_service: str = "") -> Dict:
        """Compare sentiment year over year"""
        now = datetime.now()
        current_year_start = datetime(now.year, 1, 1)
        previous_year_start = datetime(now.year - 1, 1, 1)
        previous_year_end = datetime(now.year, 1, 1)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Current year
        query_current = '''
            SELECT AVG(sentiment_score) as avg_sentiment,
                   COUNT(*) as count,
                   SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                   SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count,
                   AVG(engagement) as avg_engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params_current = [current_year_start.isoformat()]
        
        if product_service:
            query_current += ' AND product_service = ?'
            params_current.append(product_service)
        
        cursor.execute(query_current, params_current)
        current_data = cursor.fetchone()
        
        # Previous year
        query_previous = query_current.replace('date >= ?', 'date >= ? AND date < ?')
        params_previous = [previous_year_start.isoformat(), previous_year_end.isoformat()]
        
        if product_service:
            query_previous += ' AND product_service = ?'
            params_previous.append(product_service)
        
        cursor.execute(query_previous, params_previous)
        previous_data = cursor.fetchone()
        conn.close()
        
        if not current_data or not previous_data or current_data[1] == 0 or previous_data[1] == 0:
            return {'success': False, 'error': 'Insufficient data for year-over-year comparison'}
        
        current_year = {
            'year': now.year,
            'avg_sentiment': round(current_data[0] or 0, 3),
            'count': current_data[1],
            'positive_pct': round((current_data[2] / current_data[1] * 100), 1),
            'negative_pct': round((current_data[3] / current_data[1] * 100), 1),
            'avg_engagement': round(current_data[4] or 0, 1)
        }
        
        previous_year = {
            'year': now.year - 1,
            'avg_sentiment': round(previous_data[0] or 0, 3),
            'count': previous_data[1],
            'positive_pct': round((previous_data[2] / previous_data[1] * 100), 1),
            'negative_pct': round((previous_data[3] / previous_data[1] * 100), 1),
            'avg_engagement': round(previous_data[4] or 0, 1)
        }
        
        # Calculate changes
        changes = {
            'sentiment_change': round(current_year['avg_sentiment'] - previous_year['avg_sentiment'], 3),
            'count_change': current_year['count'] - previous_year['count'],
            'count_change_pct': round(((current_year['count'] - previous_year['count']) / previous_year['count'] * 100), 1) if previous_year['count'] > 0 else 0,
            'positive_pct_change': round(current_year['positive_pct'] - previous_year['positive_pct'], 1),
            'negative_pct_change': round(current_year['negative_pct'] - previous_year['negative_pct'], 1),
            'engagement_change': round(current_year['avg_engagement'] - previous_year['avg_engagement'], 1)
        }
        
        # Determine trend
        if changes['sentiment_change'] > 0.1:
            trend = 'significantly improved'
        elif changes['sentiment_change'] > 0.05:
            trend = 'improved'
        elif changes['sentiment_change'] < -0.1:
            trend = 'significantly declined'
        elif changes['sentiment_change'] < -0.05:
            trend = 'declined'
        else:
            trend = 'remained stable'
        
        return {
            'success': True,
            'current_year': current_year,
            'previous_year': previous_year,
            'changes': changes,
            'trend': trend,
            'summary': f"Sentiment has {trend} from {previous_year['avg_sentiment']:.3f} to {current_year['avg_sentiment']:.3f}"
        }
    
    def analyze_sentiment_by_content_type(self, product_service: str = "") -> Dict:
        """Analyze sentiment by content type (text, image, video, link)"""
        six_months_ago = datetime.now() - timedelta(days=180)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        query = '''
            SELECT content, source, sentiment_score, sentiment_label, engagement
            FROM sentiment_data
            WHERE date >= ?
        '''
        params = [six_months_ago.isoformat()]
        
        if product_service:
            query += ' AND product_service = ?'
            params.append(product_service)
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        if not data:
            return {'success': False, 'error': 'No data available'}
        
        # Infer content type from content and source
        content_types = defaultdict(lambda: {'scores': [], 'count': 0, 'positive': 0, 'negative': 0, 'engagement': []})
        
        for row in data:
            content = (row[0] or "").lower()
            source = (row[1] or "").lower()
            sentiment = row[2]
            label = row[3]
            engagement = row[4] or 0
            
            # Infer content type
            if 'http' in content or 'www.' in content or 'link' in source:
                content_type = 'link'
            elif 'image' in source or 'photo' in source or 'picture' in source:
                content_type = 'image'
            elif 'video' in source or 'youtube' in source or 'vimeo' in source:
                content_type = 'video'
            else:
                content_type = 'text'
            
            content_types[content_type]['scores'].append(sentiment)
            content_types[content_type]['count'] += 1
            content_types[content_type]['engagement'].append(engagement)
            if label == 'positive':
                content_types[content_type]['positive'] += 1
            elif label == 'negative':
                content_types[content_type]['negative'] += 1
        
        # Calculate metrics
        type_analysis = {}
        for content_type, data_dict in content_types.items():
            if data_dict['count'] > 0:
                type_analysis[content_type] = {
                    'avg_sentiment': round(statistics.mean(data_dict['scores']), 3),
                    'count': data_dict['count'],
                    'positive_pct': round((data_dict['positive'] / data_dict['count'] * 100), 1),
                    'negative_pct': round((data_dict['negative'] / data_dict['count'] * 100), 1),
                    'avg_engagement': round(statistics.mean(data_dict['engagement']), 1)
                }
        
        # Find best content type
        if type_analysis:
            best_type = max(type_analysis.items(), key=lambda x: x[1]['avg_sentiment'])
            
            return {
                'success': True,
                'content_type_analysis': type_analysis,
                'best_content_type': {
                    'type': best_type[0],
                    'sentiment': best_type[1]['avg_sentiment'],
                    'count': best_type[1]['count']
                }
            }
        
        return {'success': False, 'error': 'Insufficient content type data'}
    
    def generate_proactive_alerts(self, product_service: str = "", threshold: float = 0.2) -> Dict:
        """Generate proactive alerts based on pattern detection"""
        trends = self.analyze_6_month_trends(product_service)
        velocity = self.calculate_sentiment_velocity(product_service, days=7)
        
        if not trends.get('success'):
            return {'success': False, 'error': 'Insufficient data for proactive alerts'}
        
        alerts = []
        
        # Alert 1: Rapid sentiment decline
        if velocity.get('success') and velocity.get('velocity', 0) < -threshold:
            alerts.append({
                'type': 'rapid_decline',
                'severity': 'high',
                'title': 'Rapid Sentiment Decline Detected',
                'description': f"Sentiment declining at {abs(velocity['velocity']):.4f} per day",
                'recommended_action': 'Investigate root causes immediately'
            })
        
        # Alert 2: Sudden spike in negative sentiment
        if trends.get('monthly_metrics'):
            months = sorted(trends['monthly_metrics'].keys())
            if len(months) >= 2:
                last_month = trends['monthly_metrics'][months[-1]]
                prev_month = trends['monthly_metrics'][months[-2]]
                
                negative_increase = last_month.get('negative_pct', 0) - prev_month.get('negative_pct', 0)
                if negative_increase > 15:
                    alerts.append({
                        'type': 'negative_spike',
                        'severity': 'high',
                        'title': 'Sudden Increase in Negative Sentiment',
                        'description': f"Negative sentiment increased by {negative_increase:.1f}%",
                        'recommended_action': 'Review recent negative feedback and address concerns'
                    })
        
        # Alert 3: Declining engagement
        if trends.get('monthly_metrics'):
            months = sorted(trends['monthly_metrics'].keys())
            if len(months) >= 2:
                last_month = trends['monthly_metrics'][months[-1]]
                prev_month = trends['monthly_metrics'][months[-2]]
                
                engagement_decrease = prev_month.get('avg_engagement', 0) - last_month.get('avg_engagement', 0)
                if engagement_decrease > (prev_month.get('avg_engagement', 0) * 0.2):
                    alerts.append({
                        'type': 'engagement_decline',
                        'severity': 'medium',
                        'title': 'Declining Engagement Detected',
                        'description': f"Average engagement decreased by {engagement_decrease:.1f}",
                        'recommended_action': 'Review content strategy and engagement tactics'
                    })
        
        # Alert 4: Emerging negative trend
        if trends.get('overall_trend', {}).get('direction') == 'declining':
            decline_magnitude = trends.get('overall_trend', {}).get('magnitude', 0)
            if decline_magnitude > 0.15:
                alerts.append({
                    'type': 'negative_trend',
                    'severity': 'medium',
                    'title': 'Emerging Negative Trend',
                    'description': f"Overall sentiment declining with magnitude {decline_magnitude:.3f}",
                    'recommended_action': 'Monitor closely and take proactive measures'
                })
        
        return {
            'success': True,
            'alerts': alerts,
            'total_alerts': len(alerts),
            'high_severity_count': sum(1 for a in alerts if a['severity'] == 'high'),
            'medium_severity_count': sum(1 for a in alerts if a['severity'] == 'medium')
        }


def main():
    """Main function for testing and demonstration"""
    print("📊 Social Media Sentiment Analyzer - Enhanced Version")
    print("=" * 70)
    
    # Ask for API integration
    use_apis = input("Enable API integrations? (y/n): ").strip().lower() == 'y'
    api_credentials = {}
    
    if use_apis:
        print("\nAPI Configuration (press Enter to skip):")
        twitter_key = input("Twitter Consumer Key: ").strip()
        if twitter_key:
            api_credentials['twitter'] = {
                'consumer_key': twitter_key,
                'consumer_secret': input("Twitter Consumer Secret: ").strip(),
                'access_token': input("Twitter Access Token: ").strip(),
                'access_token_secret': input("Twitter Access Token Secret: ").strip()
            }
        
        reddit_id = input("Reddit Client ID: ").strip()
        if reddit_id:
            api_credentials['reddit'] = {
                'client_id': reddit_id,
                'client_secret': input("Reddit Client Secret: ").strip(),
                'user_agent': input("Reddit User Agent (default: SentimentAnalyzer/1.0): ").strip() or "SentimentAnalyzer/1.0"
            }
    
    analyzer = SocialMediaSentimentAnalyzer(
        enable_api_integrations=use_apis,
        api_credentials=api_credentials if api_credentials else None
    )
    
    print("\n" + "=" * 70)
    print("1. Analyze single text")
    print("2. Import sample data")
    print("3. Analyze 6-month trends")
    print("4. Generate summary report")
    print("5. Fetch Twitter data")
    print("6. Fetch Reddit data")
    print("7. Export to CSV")
    print("8. Export to JSON")
    print("9. Create visualizations")
    print("10. View active alerts")
    print("11. Competitor comparison")
    print("12. Predict future trends")
    print("13. Analyze by time of day")
    print("14. Analyze hashtags and mentions")
    print("15. Analyze influencer impact")
    print("16. Calculate risk score")
    print("17. Export to Excel")
    print("18. Analyze by day of week")
    print("19. Analyze by language")
    print("20. Detect bot patterns")
    print("21. Generate automated recommendations")
    print("22. Analyze emotion trends")
    print("23. Create comprehensive dashboard data")
    print("24. Analyze by geographic region")
    print("25. Analyze sentiment by topic")
    print("26. Platform comparison analysis")
    print("27. Analyze by engagement level")
    print("28. Analyze by author type")
    print("29. Create sentiment heatmap data")
    print("30. Generate executive summary")
    print("31. Analyze by content length")
    print("32. Analyze seasonal patterns")
    print("33. Analyze keyword frequency")
    print("34. Calculate sentiment velocity")
    print("35. Analyze sentiment clusters")
    print("36. Generate comparative report")
    print("37. Analyze sentiment by event/campaign")
    print("38. Calculate sentiment health score")
    print("39. Analyze sentiment with competitor mention")
    print("40. Generate benchmark report")
    print("41. Analyze sentiment momentum")
    print("42. Generate comprehensive insights summary")
    print("43. Detect crisis situations")
    print("44. Analyze sentiment by posting frequency")
    print("45. Calculate content quality score")
    print("46. Analyze sentiment by interaction type")
    print("47. Predict churn risk")
    print("48. Analyze sentiment by value perception")
    print("49. Generate action plan")
    print("50. Analyze sentiment by problem category")
    print("51. Calculate influence score for author")
    print("52. Analyze sentiment by channel")
    print("53. Generate real-time dashboard data")
    print("54. Compare multiple competitors")
    print("55. Analyze sentiment trends by segment")
    print("56. Calculate customer satisfaction score")
    print("57. Analyze sentiment by emotion category")
    print("58. Generate year-over-year comparison")
    print("59. Analyze sentiment by content type")
    print("60. Generate proactive alerts")
    print("61. Exit")
    
    while True:
        choice = input("\nSelect option (1-61): ").strip()
        
        if choice == '1':
            text = input("Enter text to analyze: ").strip()
            source = input("Source (optional): ").strip() or "manual"
            platform = input("Platform (optional): ").strip() or "unknown"
            
            result = analyzer.analyze_sentiment(text, source, platform)
            print(f"\n📊 Analysis Results:")
            print(f"   Sentiment: {result['sentiment_label']} ({result['sentiment_score']:.3f})")
            print(f"   Confidence: {result['confidence']:.3f}")
            print(f"   Emotions: {result['emotions']}")
            print(f"   Topics: {result['topics']}")
        
        elif choice == '2':
            print("\n📥 Importing sample data...")
            # Generate sample data for the past 6 months
            base_date = datetime.now() - timedelta(days=180)
            
            sample_texts = [
                ("This product is amazing! Love it!", "positive", "twitter"),
                ("Terrible quality, very disappointed", "negative", "reddit"),
                ("Good value for money, works well", "positive", "amazon"),
                ("Customer service was unhelpful", "negative", "trustpilot"),
                ("Great features, easy to use", "positive", "facebook"),
                ("Overpriced and doesn't work as advertised", "negative", "amazon"),
                ("Excellent product, highly recommend", "positive", "twitter"),
                ("Poor quality, broke after a week", "negative", "reddit"),
            ]
            
            for i in range(30):  # 30 sample entries
                text, label, platform = sample_texts[i % len(sample_texts)]
                date = base_date + timedelta(days=i * 6)
                
                analysis = analyzer.analyze_sentiment(text, f"sample_{i}", platform)
                
                data_point = SentimentDataPoint(
                    date=date.isoformat(),
                    sentiment_score=analysis['sentiment_score'],
                    sentiment_label=analysis['sentiment_label'],
                    source=f"sample_{i}",
                    platform=platform,
                    content=text,
                    emotions=analysis['emotions'],
                    topics=analysis['topics'],
                    engagement=10 + i * 2
                )
                
                analyzer.save_sentiment_data(data_point)
            
            print("✅ Sample data imported successfully!")
        
        elif choice == '3':
            product = input("Product/Service name (optional): ").strip()
            print("\n📈 Analyzing 6-month trends...")
            
            trends = analyzer.analyze_6_month_trends(product)
            if trends.get('success'):
                print(f"\n📊 Trend Analysis:")
                print(f"   Overall Trend: {trends['overall_trend']['direction']}")
                print(f"   Sentiment Change: {trends['overall_trend']['first_month_avg']} → {trends['overall_trend']['last_month_avg']}")
                print(f"\n   Top Insights:")
                for insight in trends['top_insights']:
                    print(f"     • {insight}")
            else:
                print(f"❌ {trends.get('error')}")
        
        elif choice == '4':
            product = input("Product/Service name (optional): ").strip()
            print("\n📄 Generating summary report...")
            
            report = analyzer.generate_summary_report(product)
            print("\n" + report)
            
            # Option to save
            save = input("\nSave report to file? (y/n): ").strip().lower()
            if save == 'y':
                filename = f"sentiment_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                with open(filename, 'w') as f:
                    f.write(report)
                print(f"✅ Report saved to {filename}")
        
        elif choice == '5':
            if use_apis and 'twitter' in analyzer.api_clients:
                query = input("Twitter search query: ").strip()
                product = input("Product/Service name (optional): ").strip()
                count = int(input("Number of tweets (default 100): ").strip() or "100")
                
                print("\n📥 Fetching Twitter data...")
                tweets = analyzer.fetch_twitter_data(query, count, product)
                print(f"✅ Fetched and analyzed {len(tweets)} tweets")
            else:
                print("❌ Twitter API not configured")
        
        elif choice == '6':
            if use_apis and 'reddit' in analyzer.api_clients:
                subreddit = input("Subreddit name: ").strip()
                product = input("Product/Service name (optional): ").strip()
                limit = int(input("Number of posts (default 100): ").strip() or "100")
                
                print("\n📥 Fetching Reddit data...")
                posts = analyzer.fetch_reddit_data(subreddit, limit, product)
                print(f"✅ Fetched and analyzed {len(posts)} posts")
            else:
                print("❌ Reddit API not configured")
        
        elif choice == '7':
            product = input("Product/Service name (optional): ").strip()
            print("\n📊 Exporting to CSV...")
            filename = analyzer.export_to_csv(product)
            print(f"✅ Exported to {filename}")
        
        elif choice == '8':
            product = input("Product/Service name (optional): ").strip()
            print("\n📊 Exporting to JSON...")
            filename = analyzer.export_to_json(product)
            print(f"✅ Exported to {filename}")
        
        elif choice == '9':
            product = input("Product/Service name (optional): ").strip()
            print("\n📈 Creating visualizations...")
            filename = analyzer.create_visualizations(product)
            if filename:
                print(f"✅ Visualization saved to {filename}")
        
        elif choice == '10':
            severity = input("Filter by severity (high/critical/low, or press Enter for all): ").strip() or None
            print("\n🚨 Active Alerts:")
            alerts = analyzer.get_active_alerts(severity)
            if alerts:
                for alert in alerts:
                    print(f"\n  [{alert['severity'].upper()}] {alert['alert_type']}")
                    print(f"  {alert['message']}")
                    print(f"  Created: {alert['created_at']}")
            else:
                print("  No active alerts")
        
        elif choice == '11':
            competitors_input = input("Enter competitor names (comma-separated): ").strip()
            competitors = [c.strip() for c in competitors_input.split(',') if c.strip()]
            product = input("Product/Service name (optional): ").strip()
            
            # Set competitor keywords
            competitor_keywords = {}
            for comp in competitors:
                keywords_input = input(f"Keywords for {comp} (comma-separated): ").strip()
                competitor_keywords[comp] = [k.strip() for k in keywords_input.split(',') if k.strip()]
            
            analyzer.set_competitor_keywords(competitor_keywords)
            
            print("\n📊 Analyzing competitor comparison...")
            comparison = analyzer.analyze_competitor_comparison(competitors, product)
            
            if comparison:
                print("\nCompetitor Comparison Results:")
                for comp, data in comparison.items():
                    print(f"\n  {comp}:")
                    print(f"    Average Sentiment: {data['avg_sentiment']}")
                    print(f"    Mentions: {data['mention_count']}")
                    if 'comparison_score' in data:
                        print(f"    Comparison Score: {data['comparison_score']}")
                        print(f"    Advantage: {data['advantage']}")
            else:
                print("  No competitor data found")
        
        elif choice == '12':
            product = input("Product/Service name (optional): ").strip()
            days = int(input("Days ahead to predict (default 30): ").strip() or "30")
            
            print("\n🔮 Predicting future trends...")
            prediction = analyzer.predict_future_trends(product, days)
            if prediction.get('success'):
                print(f"\n📊 Prediction Results:")
                print(f"   Current Trend: {prediction['current_trend']}")
                print(f"   Predicted Sentiment: {prediction['predicted_sentiment']}")
                print(f"   Predicted Change: {prediction['predicted_change']:+.3f}")
                print(f"   Confidence: {prediction['confidence']:.1%}")
                print(f"   Forecast: {prediction['prediction']}")
            else:
                print(f"❌ {prediction.get('error')}")
        
        elif choice == '13':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n⏰ Analyzing by time of day...")
            time_analysis = analyzer.analyze_by_time_of_day(product)
            if time_analysis.get('success'):
                print(f"\n📊 Time Analysis:")
                if 'best_hour' in time_analysis:
                    print(f"   Best Hour: {time_analysis['best_hour']['hour']}:00 (sentiment: {time_analysis['best_hour']['sentiment']})")
                if 'worst_hour' in time_analysis:
                    print(f"   Worst Hour: {time_analysis['worst_hour']['hour']}:00 (sentiment: {time_analysis['worst_hour']['sentiment']})")
                print(f"\n   Insights:")
                for insight in time_analysis.get('insights', []):
                    print(f"     • {insight}")
            else:
                print(f"❌ {time_analysis.get('error')}")
        
        elif choice == '14':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n#️⃣ Analyzing hashtags and mentions...")
            hashtag_analysis = analyzer.analyze_hashtags_and_mentions(product)
            if hashtag_analysis.get('success'):
                print(f"\n📊 Hashtag & Mention Analysis:")
                print(f"   Total Hashtags Found: {hashtag_analysis['total_hashtags']}")
                print(f"   Total Mentions Found: {hashtag_analysis['total_mentions']}")
                
                if hashtag_analysis['top_hashtags']:
                    print(f"\n   Top Hashtags:")
                    for tag, data in list(hashtag_analysis['top_hashtags'].items())[:5]:
                        print(f"     #{tag}: {data['count']} uses, sentiment: {data['avg_sentiment']:.2f}")
                
                if hashtag_analysis['top_mentions']:
                    print(f"\n   Top Mentions:")
                    for mention, data in list(hashtag_analysis['top_mentions'].items())[:5]:
                        print(f"     @{mention}: {data['count']} mentions, sentiment: {data['avg_sentiment']:.2f}")
            else:
                print(f"❌ {hashtag_analysis.get('error')}")
        
        elif choice == '15':
            product = input("Product/Service name (optional): ").strip()
            min_followers = int(input("Minimum followers for influencer (default 10000): ").strip() or "10000")
            
            print("\n👥 Analyzing influencer impact...")
            impact = analyzer.analyze_influencer_impact(min_followers, product)
            if impact.get('success'):
                print(f"\n📊 Influencer Impact Analysis:")
                print(f"   Influencers: {impact['influencer_count']}")
                print(f"   Regular Users: {impact['regular_user_count']}")
                print(f"\n   Influencer Metrics:")
                print(f"     Avg Sentiment: {impact['influencer_metrics']['avg_sentiment']}")
                print(f"     Avg Engagement: {impact['influencer_metrics']['avg_engagement']:.0f}")
                print(f"\n   Regular User Metrics:")
                print(f"     Avg Sentiment: {impact['regular_user_metrics']['avg_sentiment']}")
                print(f"     Avg Engagement: {impact['regular_user_metrics']['avg_engagement']:.0f}")
                print(f"\n   Impact:")
                print(f"     Sentiment Difference: {impact['impact_analysis']['sentiment_difference']:+.3f}")
                print(f"     Engagement Multiplier: {impact['impact_analysis']['engagement_multiplier']:.2f}x")
            else:
                print(f"❌ {impact.get('error')}")
        
        elif choice == '16':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n⚠️ Calculating risk score...")
            risk = analyzer.calculate_risk_score(product)
            if risk.get('success'):
                print(f"\n📊 Risk Analysis:")
                print(f"   Risk Score: {risk['risk_score']}/100")
                print(f"   Risk Level: {risk['risk_level'].upper()}")
                
                if risk['risk_factors']:
                    print(f"\n   Risk Factors:")
                    for factor in risk['risk_factors']:
                        print(f"     • {factor['factor']}: {factor['score']:.2f} (weight: {factor['weight']})")
                
                if risk['recommendations']:
                    print(f"\n   Recommendations:")
                    for rec in risk['recommendations']:
                        print(f"     • {rec}")
            else:
                print(f"❌ {risk.get('error')}")
        
        elif choice == '17':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📊 Exporting to Excel...")
            filename = analyzer.export_to_excel(product)
            if filename:
                print(f"✅ Exported to {filename}")
            else:
                print("❌ Excel export failed (install openpyxl: pip install openpyxl)")
        
        elif choice == '18':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📅 Analyzing by day of week...")
            day_analysis = analyzer.analyze_by_day_of_week(product)
            if day_analysis.get('success'):
                print(f"\n📊 Day of Week Analysis:")
                if 'best_day' in day_analysis:
                    print(f"   Best Day: {day_analysis['best_day']['day']} (sentiment: {day_analysis['best_day']['sentiment']})")
                if 'worst_day' in day_analysis:
                    print(f"   Worst Day: {day_analysis['worst_day']['day']} (sentiment: {day_analysis['worst_day']['sentiment']})")
                if 'most_active_day' in day_analysis:
                    print(f"   Most Active: {day_analysis['most_active_day']['day']} ({day_analysis['most_active_day']['count']} posts)")
                print(f"\n   Daily Breakdown:")
                for day, metrics in day_analysis.get('daily_breakdown', {}).items():
                    print(f"     {day}: {metrics['avg_sentiment']:.3f} ({metrics['total_count']} posts)")
            else:
                print(f"❌ {day_analysis.get('error')}")
        
        elif choice == '19':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n🌐 Analyzing by language...")
            lang_analysis = analyzer.analyze_by_language(product)
            if lang_analysis.get('success'):
                print(f"\n📊 Language Analysis:")
                print(f"   Total Languages: {lang_analysis['total_languages']}")
                if 'dominant_language' in lang_analysis:
                    dom = lang_analysis['dominant_language']
                    print(f"   Dominant Language: {dom['language']} ({dom['percentage']}%)")
                print(f"\n   Language Breakdown:")
                for lang, data in lang_analysis.get('language_breakdown', {}).items():
                    print(f"     {lang}: {data['count']} posts ({data['percentage']}%), sentiment: {data['avg_sentiment']:.3f}")
            else:
                print(f"❌ {lang_analysis.get('error')}")
        
        elif choice == '20':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n🤖 Detecting bot patterns...")
            bot_analysis = analyzer.detect_bot_patterns(product)
            if bot_analysis.get('success'):
                print(f"\n📊 Bot Detection Analysis:")
                print(f"   Total Accounts Analyzed: {bot_analysis['total_accounts_analyzed']}")
                print(f"   Suspicious Accounts Found: {bot_analysis['total_suspicious']}")
                
                if bot_analysis['suspicious_accounts']:
                    print(f"\n   Top Suspicious Accounts:")
                    for i, account in enumerate(bot_analysis['suspicious_accounts'][:10], 1):
                        print(f"     {i}. Author ID: {account['author_id']}")
                        print(f"        Bot Score: {account['bot_score']}/100")
                        print(f"        Posts: {account['post_count']}")
                        print(f"        Indicators: {', '.join(account['indicators'])}")
            else:
                print(f"❌ {bot_analysis.get('error')}")
        
        elif choice == '21':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n💡 Generating automated recommendations...")
            recommendations = analyzer.generate_automated_recommendations(product)
            if recommendations.get('success'):
                print(f"\n📋 Recommendations Summary:")
                print(f"   Total: {recommendations['total_recommendations']}")
                
                if recommendations.get('high_priority'):
                    print(f"\n   🔴 HIGH PRIORITY:")
                    for rec in recommendations['high_priority']:
                        print(f"     [{rec['category']}] {rec['recommendation']}")
                        for action in rec.get('action_items', []):
                            print(f"       • {action}")
                
                if recommendations.get('recommendations'):
                    print(f"\n   📝 RECOMMENDATIONS:")
                    for rec in recommendations['recommendations']:
                        priority_icon = '🟡' if rec['priority'] == 'MEDIUM' else '🟢'
                        print(f"     {priority_icon} [{rec['category']}] {rec['recommendation']}")
                        for action in rec.get('action_items', []):
                            print(f"       • {action}")
            else:
                print(f"❌ {recommendations.get('error')}")
        
        elif choice == '22':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n😊 Analyzing emotion trends...")
            emotion_analysis = analyzer.analyze_emotion_trends(product)
            if emotion_analysis.get('success'):
                print(f"\n📊 Emotion Trends Analysis:")
                print(f"   Emotions Tracked: {emotion_analysis['total_emotions_tracked']}")
                
                if emotion_analysis.get('trending_emotions'):
                    print(f"\n   Trending Emotions (Increasing):")
                    for emotion in emotion_analysis['trending_emotions']:
                        print(f"     • {emotion['emotion']}: +{emotion['change']:.3f}")
                
                print(f"\n   Emotion Trends by Month:")
                for emotion, trend_data in list(emotion_analysis.get('emotion_trends', {}).items())[:5]:
                    if trend_data:
                        print(f"     {emotion}: {trend_data[-1]['avg_score']:.3f} (latest)")
            else:
                print(f"❌ {emotion_analysis.get('error')}")
        
        elif choice == '23':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📊 Creating comprehensive dashboard data...")
            dashboard = analyzer.create_comprehensive_dashboard_data(product)
            if dashboard.get('success'):
                # Save to JSON
                filename = analyzer.output_dir / "exports" / f"dashboard_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(dashboard, f, indent=2, ensure_ascii=False)
                
                print(f"\n✅ Dashboard Data Created:")
                print(f"   Overview: Trend {dashboard['overview']['overall_trend'].get('direction', 'N/A')}, Risk: {dashboard['overview']['risk_score']}/100")
                print(f"   Recommendations: {len(dashboard['recommendations']['all']) + len(dashboard['recommendations']['high_priority'])}")
                print(f"   Critical Alerts: {dashboard['alerts']['critical']}")
                print(f"   Saved to: {filename}")
            else:
                print(f"❌ Failed to create dashboard data")
        
        elif choice == '24':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n🌍 Analyzing by geographic region...")
            geo_analysis = analyzer.analyze_by_geographic_region(product)
            if geo_analysis.get('success'):
                print(f"\n📊 Geographic Analysis:")
                print(f"   Total Regions: {geo_analysis['total_regions']}")
                if 'best_region' in geo_analysis:
                    print(f"   Best Region: {geo_analysis['best_region']['region']} (sentiment: {geo_analysis['best_region']['sentiment']})")
                if 'worst_region' in geo_analysis:
                    print(f"   Worst Region: {geo_analysis['worst_region']['region']} (sentiment: {geo_analysis['worst_region']['sentiment']})")
                print(f"\n   Region Breakdown:")
                for region, data in list(geo_analysis.get('region_breakdown', {}).items())[:10]:
                    print(f"     {region}: {data['avg_sentiment']:.3f} ({data['count']} posts, {data['percentage']}%)")
            else:
                print(f"❌ {geo_analysis.get('error')}")
        
        elif choice == '25':
            topic = input("Enter topic to analyze: ").strip()
            product = input("Product/Service name (optional): ").strip()
            
            print(f"\n🔍 Analyzing sentiment for topic: {topic}...")
            topic_analysis = analyzer.analyze_sentiment_by_topic(topic, product)
            if topic_analysis.get('success'):
                print(f"\n📊 Topic Analysis: {topic}")
                print(f"   Total Mentions: {topic_analysis['total_mentions']}")
                print(f"   Average Sentiment: {topic_analysis['avg_sentiment']:.3f}")
                print(f"   Positive: {topic_analysis['positive_pct']}%")
                print(f"   Negative: {topic_analysis['negative_pct']}%")
                print(f"   Total Engagement: {topic_analysis['total_engagement']:.0f}")
                print(f"   Avg Engagement: {topic_analysis['avg_engagement']:.1f}")
                if topic_analysis.get('platform_distribution'):
                    print(f"\n   Platform Distribution:")
                    for platform, count in list(topic_analysis['platform_distribution'].items())[:5]:
                        print(f"     {platform}: {count}")
            else:
                print(f"❌ {topic_analysis.get('error')}")
        
        elif choice == '26':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📱 Analyzing platform comparison...")
            platform_comp = analyzer.analyze_sentiment_by_platform_comparison(product)
            if platform_comp.get('success'):
                print(f"\n📊 Platform Comparison:")
                if 'best_platform' in platform_comp:
                    print(f"   Best Platform: {platform_comp['best_platform']['platform']} (sentiment: {platform_comp['best_platform']['sentiment']})")
                if 'worst_platform' in platform_comp:
                    print(f"   Worst Platform: {platform_comp['worst_platform']['platform']} (sentiment: {platform_comp['worst_platform']['sentiment']})")
                print(f"\n   Platform Details:")
                for platform, data in platform_comp.get('platform_comparison', {}).items():
                    print(f"     {platform}:")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Mentions: {data['total_count']}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
            else:
                print(f"❌ {platform_comp.get('error')}")
        
        elif choice == '27':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📈 Analyzing by engagement level...")
            engagement_analysis = analyzer.analyze_sentiment_by_engagement_level(product)
            if engagement_analysis.get('success'):
                print(f"\n📊 Engagement Level Analysis:")
                print(f"   Correlation: {engagement_analysis['correlation']:.3f}")
                print(f"   Interpretation: {engagement_analysis['correlation_interpretation']}")
                print(f"\n   Engagement Levels:")
                for level, data in engagement_analysis.get('engagement_level_analysis', {}).items():
                    print(f"     {level.upper()}:")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Count: {data['count']} ({data['percentage']}%)")
                if engagement_analysis.get('insights'):
                    print(f"\n   Insights:")
                    for insight in engagement_analysis['insights']:
                        print(f"     • {insight}")
            else:
                print(f"❌ {engagement_analysis.get('error')}")
        
        elif choice == '28':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n👤 Analyzing by author type...")
            author_analysis = analyzer.analyze_sentiment_by_author_type(product)
            if author_analysis.get('success'):
                print(f"\n📊 Author Type Analysis:")
                if 'best_author_type' in author_analysis:
                    best = author_analysis['best_author_type']
                    print(f"   Best Author Type: {best['type']} (sentiment: {best['sentiment']}, count: {best['count']})")
                print(f"\n   Author Type Breakdown:")
                for author_type, data in author_analysis.get('author_type_analysis', {}).items():
                    print(f"     {author_type}:")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Count: {data['count']}")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
            else:
                print(f"❌ {author_analysis.get('error')}")
        
        elif choice == '29':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n🔥 Creating sentiment heatmap data...")
            heatmap = analyzer.create_sentiment_heatmap_data(product)
            if heatmap.get('success'):
                filename = analyzer.output_dir / "exports" / f"heatmap_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(heatmap, f, indent=2, ensure_ascii=False)
                
                print(f"\n✅ Heatmap Data Created:")
                print(f"   Platforms: {len(heatmap['platforms'])}")
                print(f"   Months: {len(heatmap['months'])}")
                print(f"   Saved to: {filename}")
            else:
                print(f"❌ {heatmap.get('error')}")
        
        elif choice == '30':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📋 Generating executive summary...")
            summary = analyzer.generate_executive_summary(product)
            print("\n" + summary)
            
            save = input("\nSave summary to file? (y/n): ").strip().lower()
            if save == 'y':
                filename = analyzer.output_dir / "reports" / f"executive_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(summary)
                print(f"✅ Summary saved to {filename}")
        
        elif choice == '31':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📏 Analyzing by content length...")
            length_analysis = analyzer.analyze_sentiment_by_content_length(product)
            if length_analysis.get('success'):
                print(f"\n📊 Content Length Analysis:")
                print(f"   Optimal Length: {length_analysis['optimal_length_category']} (sentiment: {length_analysis['optimal_sentiment']:.3f})")
                print(f"\n   Length Categories:")
                for length_cat, data in length_analysis.get('length_analysis', {}).items():
                    print(f"     {length_cat.replace('_', ' ').title()}:")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
                    print(f"       Count: {data['count']} ({data['percentage']}%)")
                if length_analysis.get('insights'):
                    print(f"\n   Insights:")
                    for insight in length_analysis['insights']:
                        print(f"     • {insight}")
            else:
                print(f"❌ {length_analysis.get('error')}")
        
        elif choice == '32':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📅 Analyzing seasonal patterns...")
            seasonal = analyzer.analyze_seasonal_patterns(product)
            if seasonal.get('success'):
                print(f"\n📊 Seasonal Analysis:")
                if 'best_month' in seasonal:
                    print(f"   Best Month: {seasonal['best_month']['month']} (sentiment: {seasonal['best_month']['sentiment']})")
                if 'worst_month' in seasonal:
                    print(f"   Worst Month: {seasonal['worst_month']['month']} (sentiment: {seasonal['worst_month']['sentiment']})")
                print(f"\n   Monthly Breakdown:")
                for month, data in seasonal.get('seasonal_patterns', {}).items():
                    print(f"     {month}: {data['avg_sentiment']:.3f} ({data['count']} posts, {data['positive_pct']}% positive)")
                if seasonal.get('insights'):
                    print(f"\n   Insights:")
                    for insight in seasonal['insights']:
                        print(f"     • {insight}")
            else:
                print(f"❌ {seasonal.get('error')}")
        
        elif choice == '33':
            product = input("Product/Service name (optional): ").strip()
            top_n = int(input("Number of top keywords (default 20): ").strip() or "20")
            
            print("\n🔑 Analyzing keyword frequency...")
            keywords = analyzer.analyze_keyword_frequency(product, top_n)
            if keywords.get('success'):
                print(f"\n📊 Keyword Analysis:")
                print(f"   Total Unique Keywords: {keywords['total_unique_keywords']}")
                print(f"\n   Top Keywords:")
                for word, data in list(keywords.get('top_keywords', {}).items())[:10]:
                    print(f"     {word}: {data['frequency']} mentions, sentiment: {data['avg_sentiment']:.3f}")
                if keywords.get('positive_keywords'):
                    print(f"\n   Positive Keywords:")
                    for word in keywords['positive_keywords'][:5]:
                        print(f"     • {word}")
                if keywords.get('negative_keywords'):
                    print(f"\n   Negative Keywords:")
                    for word in keywords['negative_keywords'][:5]:
                        print(f"     • {word}")
            else:
                print(f"❌ {keywords.get('error')}")
        
        elif choice == '34':
            product = input("Product/Service name (optional): ").strip()
            days = int(input("Days to analyze (default 7): ").strip() or "7")
            
            print(f"\n⚡ Calculating sentiment velocity (last {days} days)...")
            velocity = analyzer.calculate_sentiment_velocity(product, days)
            if velocity.get('success'):
                print(f"\n📊 Sentiment Velocity:")
                print(f"   Velocity: {velocity['velocity']:+.4f} per day")
                print(f"   Direction: {velocity['direction']}")
                print(f"   Interpretation: {velocity['interpretation']}")
                print(f"   Data Points: {velocity['data_points']}")
            else:
                print(f"❌ {velocity.get('error')}")
        
        elif choice == '35':
            product = input("Product/Service name (optional): ").strip()
            n_clusters = int(input("Number of clusters (default 5): ").strip() or "5")
            
            print(f"\n🔗 Analyzing sentiment clusters ({n_clusters} clusters)...")
            clusters = analyzer.analyze_sentiment_clusters(product, n_clusters)
            if clusters.get('success'):
                print(f"\n📊 Cluster Analysis:")
                print(f"   Total Items: {clusters['total_items']}")
                print(f"   Clusters Found: {len(clusters['clusters'])}")
                print(f"\n   Cluster Details:")
                for cluster_id, data in clusters.get('clusters', {}).items():
                    label = clusters.get('cluster_labels', {}).get(cluster_id, 'Unknown')
                    print(f"     Cluster {cluster_id} ({label}):")
                    print(f"       Size: {data['size']} ({data['percentage']}%)")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
            else:
                print(f"❌ {clusters.get('error')}")
        
        elif choice == '36':
            product = input("Product/Service name (optional): ").strip()
            period1 = int(input("First period days (default 90): ").strip() or "90")
            period2 = int(input("Second period days (default 90): ").strip() or "90")
            
            print(f"\n📊 Generating comparative report...")
            comparison = analyzer.generate_comparative_report(period1, period2, product)
            if comparison.get('success'):
                print(f"\n📈 Comparative Analysis:")
                print(f"   Trend: {comparison['trend']}")
                print(f"   Summary: {comparison['summary']}")
                print(f"\n   {comparison['period1']['label']}:")
                p1 = comparison['period1']['data']
                print(f"     Sentiment: {p1['avg_sentiment']:.3f}")
                print(f"     Mentions: {p1['count']}")
                print(f"     Positive: {p1['positive_pct']}% | Negative: {p1['negative_pct']}%")
                print(f"\n   {comparison['period2']['label']}:")
                p2 = comparison['period2']['data']
                print(f"     Sentiment: {p2['avg_sentiment']:.3f}")
                print(f"     Mentions: {p2['count']}")
                print(f"     Positive: {p2['positive_pct']}% | Negative: {p2['negative_pct']}%")
                print(f"\n   Changes:")
                ch = comparison['changes']
                print(f"     Sentiment: {ch['sentiment_change']:+.3f}")
                print(f"     Mentions: {ch['count_change']:+d} ({ch['count_change_pct']:+.1f}%)")
                print(f"     Positive: {ch['positive_pct_change']:+.1f}%")
                print(f"     Negative: {ch['negative_pct_change']:+.1f}%")
            else:
                print(f"❌ {comparison.get('error')}")
        
        elif choice == '37':
            event_name = input("Event/Campaign name: ").strip()
            product = input("Product/Service name (optional): ").strip()
            days_before = int(input("Days before event (default 7): ").strip() or "7")
            days_after = int(input("Days after event (default 7): ").strip() or "7")
            
            print(f"\n📅 Analyzing sentiment around event: {event_name}...")
            event_analysis = analyzer.analyze_sentiment_by_event_campaign(event_name, days_before, days_after, product)
            if event_analysis.get('success'):
                print(f"\n📊 Event Impact Analysis:")
                print(f"   Event Date: {event_analysis['event_date']}")
                print(f"\n   Before Event ({event_analysis['before_event']['period']}):")
                print(f"     Sentiment: {event_analysis['before_event']['avg_sentiment']:.3f}")
                print(f"     Mentions: {event_analysis['before_event']['count']}")
                print(f"\n   During Event ({event_analysis['during_event']['period']}):")
                print(f"     Sentiment: {event_analysis['during_event']['avg_sentiment']:.3f}")
                print(f"     Mentions: {event_analysis['during_event']['count']}")
                print(f"\n   After Event ({event_analysis['after_event']['period']}):")
                print(f"     Sentiment: {event_analysis['after_event']['avg_sentiment']:.3f}")
                print(f"     Mentions: {event_analysis['after_event']['count']}")
                print(f"\n   Impact: {event_analysis['impact']['sentiment_change']:+.3f} ({event_analysis['impact']['interpretation']})")
            else:
                print(f"❌ {event_analysis.get('error')}")
        
        elif choice == '38':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n💚 Calculating sentiment health score...")
            health = analyzer.calculate_sentiment_health_score(product)
            if health.get('success'):
                print(f"\n📊 Sentiment Health Score:")
                print(f"   Health Score: {health['health_score']}/100")
                print(f"   Health Level: {health['health_level'].upper()}")
                print(f"\n   Contributing Factors:")
                for factor in health.get('factors', []):
                    print(f"     • {factor['factor']}: {factor['score']:.1f} (weight: {factor['weight']})")
                if health.get('recommendations'):
                    print(f"\n   Recommendations:")
                    for rec in health['recommendations']:
                        print(f"     • {rec}")
            else:
                print(f"❌ {health.get('error')}")
        
        elif choice == '39':
            competitor = input("Competitor name: ").strip()
            product = input("Product/Service name (optional): ").strip()
            
            print(f"\n🔍 Analyzing sentiment with competitor mention: {competitor}...")
            comp_analysis = analyzer.analyze_sentiment_by_competitor_mention(competitor, product)
            if comp_analysis.get('success'):
                print(f"\n📊 Competitor Mention Analysis:")
                print(f"   With {competitor} Mention:")
                wc = comp_analysis['with_competitor_mention']
                print(f"     Sentiment: {wc['avg_sentiment']:.3f}")
                print(f"     Mentions: {wc['count']}")
                print(f"     Positive: {wc['positive_pct']}% | Negative: {wc['negative_pct']}%")
                print(f"\n   Without {competitor} Mention:")
                woc = comp_analysis['without_competitor_mention']
                print(f"     Sentiment: {woc['avg_sentiment']:.3f}")
                print(f"     Mentions: {woc['count']}")
                print(f"     Positive: {woc['positive_pct']}% | Negative: {woc['negative_pct']}%")
                print(f"\n   Difference: {comp_analysis['difference']['sentiment_diff']:+.3f} ({comp_analysis['difference']['interpretation']})")
            else:
                print(f"❌ {comp_analysis.get('error')}")
        
        elif choice == '40':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📊 Generating benchmark report...")
            benchmark = analyzer.generate_benchmark_report(product)
            if benchmark.get('success'):
                print(f"\n📈 Benchmark Analysis:")
                print(f"   Benchmark Score: {benchmark['benchmark_score']}/100")
                print(f"   Overall Status: {benchmark['overall_status'].upper()}")
                print(f"\n   Benchmark Comparisons:")
                for metric, data in benchmark.get('benchmarks', {}).items():
                    print(f"     {metric.replace('_', ' ').title()}:")
                    print(f"       Current: {data['current']}")
                    print(f"       Industry Avg: {data['industry_avg']}")
                    print(f"       Difference: {data['difference']:+.2f} ({data['status']})")
                if benchmark.get('recommendations'):
                    print(f"\n   Recommendations:")
                    for rec in benchmark['recommendations']:
                        print(f"     • {rec}")
            else:
                print(f"❌ {benchmark.get('error')}")
        
        elif choice == '41':
            product = input("Product/Service name (optional): ").strip()
            window = int(input("Analysis window in days (default 30): ").strip() or "30")
            
            print(f"\n⚡ Analyzing sentiment momentum ({window} days)...")
            momentum = analyzer.analyze_sentiment_momentum(product, window)
            if momentum.get('success'):
                print(f"\n📊 Momentum Analysis:")
                print(f"   Velocity: {momentum['velocity']:+.4f} per day")
                print(f"   Momentum: {momentum['momentum']:+.4f}")
                print(f"   Direction: {momentum['momentum_direction']}")
                print(f"   Interpretation: {momentum['interpretation']}")
            else:
                print(f"❌ {momentum.get('error')}")
        
        elif choice == '42':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📋 Generating comprehensive insights summary...")
            insights = analyzer.generate_insights_summary(product)
            print("\n" + insights)
            
            save = input("\nSave insights to file? (y/n): ").strip().lower()
            if save == 'y':
                filename = analyzer.output_dir / "reports" / f"insights_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(insights)
                print(f"✅ Insights saved to {filename}")
        
        elif choice == '43':
            product = input("Product/Service name (optional): ").strip()
            threshold = float(input("Negative sentiment threshold (default 0.3): ").strip() or "0.3")
            
            print("\n🚨 Detecting crisis situations...")
            crisis = analyzer.detect_crisis_situations(product, threshold)
            if crisis.get('success'):
                print(f"\n📊 Crisis Detection:")
                print(f"   Crisis Score: {crisis['crisis_score']}/100")
                print(f"   Crisis Level: {crisis['crisis_level'].upper()}")
                print(f"   Action Required: {crisis['action_required'].upper()}")
                if crisis.get('indicators'):
                    print(f"\n   Crisis Indicators:")
                    for indicator in crisis['indicators']:
                        print(f"     [{indicator['severity'].upper()}] {indicator['description']} (score: {indicator['score']})")
                if crisis.get('recommendations'):
                    print(f"\n   Recommendations:")
                    for rec in crisis['recommendations']:
                        print(f"     • {rec}")
            else:
                print(f"❌ {crisis.get('error')}")
        
        elif choice == '44':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📊 Analyzing sentiment by posting frequency...")
            frequency = analyzer.analyze_sentiment_by_frequency(product)
            if frequency.get('success'):
                print(f"\n📊 Frequency Analysis:")
                print(f"   Frequency Categories:")
                for freq_cat, data in frequency.get('frequency_analysis', {}).items():
                    print(f"     {freq_cat.replace('_', ' ').title()}:")
                    print(f"       Authors: {data['author_count']}")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
                if frequency.get('insights'):
                    print(f"\n   Insights:")
                    for insight in frequency['insights']:
                        print(f"     • {insight}")
            else:
                print(f"❌ {frequency.get('error')}")
        
        elif choice == '45':
            content = input("Enter content to score: ").strip()
            sentiment_input = input("Sentiment score (-1 to 1, or press Enter to analyze): ").strip()
            
            if sentiment_input:
                sentiment_score = float(sentiment_input)
            else:
                analysis = analyzer.analyze_sentiment(content)
                sentiment_score = analysis.get('sentiment_score', 0)
            
            engagement = int(input("Engagement count: ").strip() or "0")
            followers = int(input("Author followers: ").strip() or "0")
            
            print("\n⭐ Calculating content quality score...")
            quality = analyzer.calculate_content_quality_score(content, sentiment_score, engagement, followers)
            print(f"\n📊 Content Quality Score:")
            print(f"   Quality Score: {quality['quality_score']}/100")
            print(f"   Quality Level: {quality['quality_level'].upper()}")
            print(f"\n   Contributing Factors:")
            for factor in quality.get('factors', []):
                print(f"     • {factor['factor']}: {factor['score']:.1f} (weight: {factor['weight']})")
        
        elif choice == '46':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n💬 Analyzing sentiment by interaction type...")
            interaction = analyzer.analyze_sentiment_by_interaction_type(product)
            if interaction.get('success'):
                print(f"\n📊 Interaction Type Analysis:")
                if 'best_interaction_type' in interaction:
                    best = interaction['best_interaction_type']
                    print(f"   Best Type: {best['type']} (sentiment: {best['sentiment']}, count: {best['count']})")
                print(f"\n   Interaction Type Breakdown:")
                for int_type, data in interaction.get('interaction_type_analysis', {}).items():
                    print(f"     {int_type.title()}:")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                    print(f"       Engagement: {data['avg_engagement']:.1f}")
                    print(f"       Count: {data['count']}")
            else:
                print(f"❌ {interaction.get('error')}")
        
        elif choice == '47':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n⚠️ Predicting churn risk...")
            churn = analyzer.predict_churn_risk(product)
            if churn.get('success'):
                print(f"\n📊 Churn Risk Prediction:")
                print(f"   Churn Risk Score: {churn['churn_risk_score']}/100")
                print(f"   Churn Risk Level: {churn['churn_risk_level'].upper()}")
                print(f"   Recommended Action: {churn['recommended_action'].replace('_', ' ').title()}")
                if churn.get('risk_factors'):
                    print(f"\n   Risk Factors:")
                    for factor in churn['risk_factors']:
                        print(f"     • {factor['factor']}: contributes {factor['contribution']:.1f} (weight: {factor['weight']})")
                if churn.get('recommendations'):
                    print(f"\n   Recommendations:")
                    for rec in churn['recommendations']:
                        print(f"     • {rec}")
            else:
                print(f"❌ {churn.get('error')}")
        
        elif choice == '48':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n💰 Analyzing sentiment by value perception...")
            value = analyzer.analyze_sentiment_by_value_perception(product)
            if value.get('success'):
                print(f"\n📊 Value Perception Analysis:")
                print(f"   Overall Perception: {value['overall_perception'].upper()}")
                print(f"   Total Value Mentions: {value['total_value_mentions']}")
                print(f"\n   Value Categories:")
                for category, data in value.get('value_analysis', {}).items():
                    print(f"     {category.replace('_', ' ').title()}:")
                    print(f"       Count: {data['count']}")
                    print(f"       Avg Sentiment: {data['avg_sentiment']:.3f}")
                    if data.get('top_keywords'):
                        print(f"       Top Keywords: {', '.join(list(data['top_keywords'].keys())[:3])}")
            else:
                print(f"❌ {value.get('error')}")
        
        elif choice == '49':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📋 Generating action plan...")
            action_plan = analyzer.generate_action_plan(product)
            if action_plan.get('success'):
                print(f"\n📊 Action Plan Generated:")
                
                if action_plan.get('priority_actions'):
                    print(f"\n   🔴 PRIORITY ACTIONS:")
                    for action in action_plan['priority_actions']:
                        print(f"     • {action['action']}")
                        print(f"       Priority: {action['priority']} | Timeline: {action['timeline']} | Owner: {action['owner']}")
                
                if action_plan.get('short_term_actions'):
                    print(f"\n   ⚡ SHORT-TERM ACTIONS (Next 7 days):")
                    for action in action_plan['short_term_actions']:
                        print(f"     • {action['action']}")
                        print(f"       Priority: {action['priority']} | Timeline: {action['timeline']} | Owner: {action['owner']}")
                
                if action_plan.get('long_term_actions'):
                    print(f"\n   📅 LONG-TERM ACTIONS (30-90 days):")
                    for action in action_plan['long_term_actions']:
                        print(f"     • {action['action']}")
                        print(f"       Priority: {action['priority']} | Timeline: {action['timeline']} | Owner: {action['owner']}")
                
                if action_plan.get('monitoring_actions'):
                    print(f"\n   👁️ MONITORING ACTIONS:")
                    for action in action_plan['monitoring_actions']:
                        print(f"     • {action['action']}")
                        print(f"       Frequency: {action['frequency']} | Owner: {action['owner']}")
                
                # Save to JSON
                save = input("\nSave action plan to file? (y/n): ").strip().lower()
                if save == 'y':
                    filename = analyzer.output_dir / "reports" / f"action_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(action_plan, f, indent=2, ensure_ascii=False)
                    print(f"✅ Action plan saved to {filename}")
            else:
                print(f"❌ Failed to generate action plan")
        
        elif choice == '50':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n🔍 Analyzing sentiment by problem category...")
            problems = analyzer.analyze_sentiment_by_problem_category(product)
            if problems.get('success'):
                print(f"\n📊 Problem Category Analysis:")
                if 'most_problematic_category' in problems:
                    mp = problems['most_problematic_category']
                    print(f"   Most Problematic: {mp['category'].replace('_', ' ').title()} (sentiment: {mp['sentiment']:.3f}, mentions: {mp['count']})")
                print(f"\n   Category Breakdown:")
                for category, data in problems.get('problem_analysis', {}).items():
                    print(f"     {category.replace('_', ' ').title()}:")
                    print(f"       Count: {data['count']}")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
            else:
                print(f"❌ {problems.get('error')}")
        
        elif choice == '51':
            author_id = input("Author ID: ").strip()
            product = input("Product/Service name (optional): ").strip()
            
            print(f"\n⭐ Calculating influence score for author: {author_id}...")
            influence = analyzer.calculate_influence_score(author_id, product)
            if influence.get('success'):
                print(f"\n📊 Influence Analysis:")
                print(f"   Influence Score: {influence['influence_score']}/100")
                print(f"   Influence Level: {influence['influence_level'].upper()}")
                print(f"\n   Metrics:")
                metrics = influence.get('metrics', {})
                print(f"     Post Count: {metrics.get('post_count', 0)}")
                print(f"     Avg Sentiment: {metrics.get('avg_sentiment', 0):.3f}")
                print(f"     Avg Engagement: {metrics.get('avg_engagement', 0):.1f}")
                print(f"     Max Engagement: {metrics.get('max_engagement', 0)}")
                print(f"     Total Engagement: {metrics.get('total_engagement', 0)}")
            else:
                print(f"❌ {influence.get('error')}")
        
        elif choice == '52':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📡 Analyzing sentiment by channel...")
            channels = analyzer.analyze_sentiment_by_channel(product)
            if channels.get('success'):
                print(f"\n📊 Channel Analysis:")
                if 'best_channel' in channels:
                    best = channels['best_channel']
                    print(f"   Best Channel: {best['channel']} (sentiment: {best['sentiment']:.3f}, mentions: {best['count']})")
                if 'worst_channel' in channels:
                    worst = channels['worst_channel']
                    print(f"   Worst Channel: {worst['channel']} (sentiment: {worst['sentiment']:.3f}, mentions: {worst['count']})")
                print(f"\n   Channel Breakdown:")
                for channel, data in channels.get('channel_analysis', {}).items():
                    print(f"     {channel.replace('_', ' ').title()}:")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                    print(f"       Engagement: {data['avg_engagement']:.1f}")
                    print(f"       Total Mentions: {data['total_count']}")
            else:
                print(f"❌ {channels.get('error')}")
        
        elif choice == '53':
            product = input("Product/Service name (optional): ").strip()
            hours = int(input("Hours to analyze (default 24): ").strip() or "24")
            
            print(f"\n⚡ Generating real-time dashboard data (last {hours} hours)...")
            dashboard = analyzer.generate_real_time_dashboard_data(product, hours)
            if dashboard.get('success'):
                print(f"\n📊 Real-Time Dashboard:")
                summary = dashboard.get('summary', {})
                print(f"   Total Mentions: {summary.get('total_mentions', 0)}")
                print(f"   Positive: {summary.get('positive_count', 0)} ({summary.get('positive_pct', 0)}%)")
                print(f"   Negative: {summary.get('negative_count', 0)} ({summary.get('negative_pct', 0)}%)")
                print(f"   Avg Sentiment: {summary.get('avg_sentiment', 0):.3f}")
                print(f"   Total Engagement: {summary.get('total_engagement', 0)}")
                print(f"   Avg Engagement: {summary.get('avg_engagement', 0):.1f}")
                
                if dashboard.get('hourly_breakdown'):
                    print(f"\n   Recent Hours (last 5):")
                    for hour, data in list(dashboard['hourly_breakdown'].items())[-5:]:
                        print(f"     {hour}: {data['count']} mentions, sentiment: {data['avg_sentiment']:.3f}")
                
                if dashboard.get('platform_breakdown'):
                    print(f"\n   Platform Breakdown:")
                    for platform, data in dashboard['platform_breakdown'].items():
                        print(f"     {platform}: {data['count']} mentions, sentiment: {data['avg_sentiment']:.3f}")
                
                # Save to JSON
                save = input("\nSave dashboard data to file? (y/n): ").strip().lower()
                if save == 'y':
                    filename = analyzer.output_dir / "reports" / f"realtime_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(dashboard, f, indent=2, ensure_ascii=False)
                    print(f"✅ Dashboard data saved to {filename}")
            else:
                print(f"❌ {dashboard.get('error')}")
        
        elif choice == '54':
            competitors_input = input("Enter competitor names (comma-separated): ").strip()
            competitors = [c.strip() for c in competitors_input.split(',') if c.strip()]
            product = input("Product/Service name (optional): ").strip()
            
            if not competitors:
                print("❌ Please enter at least one competitor name")
            else:
                print(f"\n🔍 Comparing {len(competitors)} competitors...")
                comparison = analyzer.compare_multiple_competitors(competitors, product)
                if comparison.get('success'):
                    print(f"\n📊 Multi-Competitor Comparison:")
                    print(f"\n   Rankings:")
                    for rank_info in comparison.get('rankings', []):
                        print(f"     {rank_info['rank']}. {rank_info['competitor']}: {rank_info['sentiment']:.3f}")
                    
                    if 'best_competitor' in comparison:
                        best = comparison['best_competitor']
                        print(f"\n   Best: {best['name']} (sentiment: {best['sentiment']:.3f})")
                    
                    if 'worst_competitor' in comparison:
                        worst = comparison['worst_competitor']
                        print(f"   Worst: {worst['name']} (sentiment: {worst['sentiment']:.3f})")
                    
                    print(f"\n   Detailed Comparison:")
                    for comp, data in comparison.get('competitors', {}).items():
                        print(f"     {comp}:")
                        print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                        print(f"       Mentions: {data['count']}")
                        print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                        print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
                else:
                    print(f"❌ {comparison.get('error')}")
        
        elif choice == '55':
            print("\nAvailable segment keys: platform, language, region, product_service")
            segment_key = input("Segment key: ").strip()
            segment_value = input("Segment value: ").strip()
            product = input("Product/Service name (optional): ").strip()
            
            print(f"\n📊 Analyzing sentiment trends for {segment_key}={segment_value}...")
            segment_trends = analyzer.analyze_sentiment_trends_by_segment(segment_key, segment_value, product)
            if segment_trends.get('success'):
                print(f"\n📈 Segment Trend Analysis:")
                print(f"   Segment: {segment_trends['segment_key']} = {segment_trends['segment_value']}")
                overall = segment_trends.get('overall_trend', {})
                print(f"   Overall Trend: {overall.get('direction', 'N/A').upper()} (magnitude: {overall.get('magnitude', 0):.3f})")
                print(f"\n   Monthly Trends:")
                for month, data in segment_trends.get('monthly_trends', {}).items():
                    print(f"     {month}:")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Mentions: {data['count']}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
            else:
                print(f"❌ {segment_trends.get('error')}")
        
        elif choice == '56':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n😊 Calculating customer satisfaction score...")
            satisfaction = analyzer.calculate_customer_satisfaction_score(product)
            if satisfaction.get('success'):
                print(f"\n📊 Customer Satisfaction Analysis:")
                print(f"   Satisfaction Score: {satisfaction['satisfaction_score']}/100")
                print(f"   Satisfaction Level: {satisfaction['satisfaction_level'].upper()}")
                print(f"   NPS Score: {satisfaction['nps_score']:+.1f}")
                print(f"   Promoters: {satisfaction['promoter_pct']}%")
                print(f"   Detractors: {satisfaction['detractor_pct']}%")
                print(f"   Positive: {satisfaction['positive_pct']}% | Negative: {satisfaction['negative_pct']}%")
                print(f"   Avg Sentiment: {satisfaction['avg_sentiment']:.3f}")
            else:
                print(f"❌ {satisfaction.get('error')}")
        
        elif choice == '57':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n😊 Analyzing sentiment by emotion category...")
            emotions = analyzer.analyze_sentiment_by_emotion_category(product)
            if emotions.get('success'):
                print(f"\n📊 Emotion Analysis:")
                if 'dominant_emotion' in emotions:
                    dom = emotions['dominant_emotion']
                    print(f"   Dominant Emotion: {dom['emotion'].upper()} (count: {dom['count']}, sentiment: {dom['avg_sentiment']:.3f})")
                print(f"\n   Emotion Breakdown:")
                for emotion, data in emotions.get('emotion_analysis', {}).items():
                    print(f"     {emotion.title()}:")
                    print(f"       Count: {data['count']}")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
            else:
                print(f"❌ {emotions.get('error')}")
        
        elif choice == '58':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📅 Generating year-over-year comparison...")
            yoy = analyzer.generate_year_over_year_comparison(product)
            if yoy.get('success'):
                print(f"\n📊 Year-Over-Year Comparison:")
                print(f"   Trend: {yoy['trend']}")
                print(f"   Summary: {yoy['summary']}")
                print(f"\n   Current Year ({yoy['current_year']['year']}):")
                cy = yoy['current_year']
                print(f"     Sentiment: {cy['avg_sentiment']:.3f}")
                print(f"     Mentions: {cy['count']}")
                print(f"     Positive: {cy['positive_pct']}% | Negative: {cy['negative_pct']}%")
                print(f"\n   Previous Year ({yoy['previous_year']['year']}):")
                py = yoy['previous_year']
                print(f"     Sentiment: {py['avg_sentiment']:.3f}")
                print(f"     Mentions: {py['count']}")
                print(f"     Positive: {py['positive_pct']}% | Negative: {py['negative_pct']}%")
                print(f"\n   Changes:")
                ch = yoy['changes']
                print(f"     Sentiment: {ch['sentiment_change']:+.3f}")
                print(f"     Mentions: {ch['count_change']:+d} ({ch['count_change_pct']:+.1f}%)")
                print(f"     Positive: {ch['positive_pct_change']:+.1f}%")
                print(f"     Negative: {ch['negative_pct_change']:+.1f}%")
            else:
                print(f"❌ {yoy.get('error')}")
        
        elif choice == '59':
            product = input("Product/Service name (optional): ").strip()
            
            print("\n📄 Analyzing sentiment by content type...")
            content_types = analyzer.analyze_sentiment_by_content_type(product)
            if content_types.get('success'):
                print(f"\n📊 Content Type Analysis:")
                if 'best_content_type' in content_types:
                    best = content_types['best_content_type']
                    print(f"   Best Type: {best['type']} (sentiment: {best['sentiment']:.3f}, count: {best['count']})")
                print(f"\n   Content Type Breakdown:")
                for ctype, data in content_types.get('content_type_analysis', {}).items():
                    print(f"     {ctype.title()}:")
                    print(f"       Sentiment: {data['avg_sentiment']:.3f}")
                    print(f"       Count: {data['count']}")
                    print(f"       Positive: {data['positive_pct']}% | Negative: {data['negative_pct']}%")
                    print(f"       Avg Engagement: {data['avg_engagement']:.1f}")
            else:
                print(f"❌ {content_types.get('error')}")
        
        elif choice == '60':
            product = input("Product/Service name (optional): ").strip()
            threshold = float(input("Velocity threshold (default 0.2): ").strip() or "0.2")
            
            print("\n🚨 Generating proactive alerts...")
            alerts = analyzer.generate_proactive_alerts(product, threshold)
            if alerts.get('success'):
                print(f"\n📊 Proactive Alerts:")
                print(f"   Total Alerts: {alerts['total_alerts']}")
                print(f"   High Severity: {alerts['high_severity_count']}")
                print(f"   Medium Severity: {alerts['medium_severity_count']}")
                
                if alerts.get('alerts'):
                    print(f"\n   Alert Details:")
                    for alert in alerts['alerts']:
                        print(f"     [{alert['severity'].upper()}] {alert['title']}")
                        print(f"       Description: {alert['description']}")
                        print(f"       Action: {alert['recommended_action']}")
                        print()
                else:
                    print("\n   ✅ No proactive alerts at this time")
            else:
                print(f"❌ {alerts.get('error')}")
        
        elif choice == '61':
            print("👋 Goodbye!")
            break
        
        else:
            print("❌ Invalid option")


if __name__ == "__main__":
    main()





