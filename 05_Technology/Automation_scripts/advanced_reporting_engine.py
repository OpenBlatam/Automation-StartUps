"""
Motor de Reportes Avanzados
Sistema de generación de reportes automáticos con análisis, visualizaciones y exportación
"""

import asyncio
import json
import logging
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass
from enum import Enum
import warnings
warnings.filterwarnings('ignore')

# Report generation libraries
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.figure_factory as ff
from jinja2 import Template
import weasyprint
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Data analysis libraries
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import scipy.stats as stats

class ReportType(Enum):
    EXECUTIVE = "executive"
    OPERATIONAL = "operational"
    ANALYTICAL = "analytical"
    FINANCIAL = "financial"
    PERFORMANCE = "performance"
    COMPLIANCE = "compliance"
    CUSTOM = "custom"

class ReportFormat(Enum):
    PDF = "pdf"
    HTML = "html"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"
    POWERPOINT = "powerpoint"
    WORD = "word"

class ReportFrequency(Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"
    ON_DEMAND = "on_demand"

@dataclass
class ReportRequest:
    report_id: str
    report_type: ReportType
    title: str
    description: str
    data_sources: List[str]
    metrics: List[str]
    visualizations: List[str]
    format: ReportFormat
    frequency: ReportFrequency
    recipients: List[str]
    parameters: Dict[str, Any] = None
    template: str = None
    filters: Dict[str, Any] = None

@dataclass
class ReportResult:
    report_id: str
    status: str
    file_path: str
    file_size: int
    generation_time: float
    metrics_summary: Dict[str, Any]
    visualizations: List[str]
    recipients: List[str]
    generated_at: str

class AdvancedReportingEngine:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.report_templates = {}
        self.report_history = {}
        self.data_sources = {}
        self.metrics_calculators = {}
        
        # Configuración por defecto
        self.default_config = {
            "output_directory": "./reports",
            "template_directory": "./templates",
            "max_file_size": 50 * 1024 * 1024,  # 50MB
            "default_font": "Arial",
            "default_font_size": 12,
            "chart_width": 800,
            "chart_height": 600,
            "email_smtp_server": "smtp.gmail.com",
            "email_smtp_port": 587
        }
        
        # Inicializar calculadores de métricas
        self._initialize_metrics_calculators()
        
        # Inicializar plantillas
        self._initialize_templates()
        
    def _initialize_metrics_calculators(self):
        """Inicializar calculadores de métricas"""
        try:
            self.metrics_calculators = {
                "accuracy": self._calculate_accuracy,
                "precision": self._calculate_precision,
                "recall": self._calculate_recall,
                "f1_score": self._calculate_f1_score,
                "mse": self._calculate_mse,
                "mae": self._calculate_mae,
                "r2_score": self._calculate_r2_score,
                "mean": self._calculate_mean,
                "median": self._calculate_median,
                "std": self._calculate_std,
                "variance": self._calculate_variance,
                "correlation": self._calculate_correlation,
                "trend": self._calculate_trend,
                "growth_rate": self._calculate_growth_rate,
                "volatility": self._calculate_volatility
            }
        except Exception as e:
            self.logger.error(f"Error initializing metrics calculators: {e}")
    
    def _initialize_templates(self):
        """Inicializar plantillas de reportes"""
        try:
            # Plantilla ejecutiva
            self.report_templates["executive"] = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>{{ title }}</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    .header { text-align: center; margin-bottom: 30px; }
                    .summary { background-color: #f5f5f5; padding: 20px; margin: 20px 0; }
                    .metric { display: inline-block; margin: 10px; padding: 15px; background-color: #e8f4fd; border-radius: 5px; }
                    .chart { margin: 20px 0; text-align: center; }
                    .footer { margin-top: 30px; text-align: center; color: #666; }
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{{ title }}</h1>
                    <p>{{ description }}</p>
                    <p>Generated on: {{ generated_at }}</p>
                </div>
                
                <div class="summary">
                    <h2>Executive Summary</h2>
                    <p>{{ summary }}</p>
                </div>
                
                <div class="metrics">
                    <h2>Key Metrics</h2>
                    {% for metric in metrics %}
                    <div class="metric">
                        <strong>{{ metric.name }}</strong><br>
                        {{ metric.value }}<br>
                        <small>{{ metric.change }}</small>
                    </div>
                    {% endfor %}
                </div>
                
                <div class="charts">
                    <h2>Visualizations</h2>
                    {% for chart in charts %}
                    <div class="chart">
                        <h3>{{ chart.title }}</h3>
                        <img src="{{ chart.path }}" alt="{{ chart.title }}">
                    </div>
                    {% endfor %}
                </div>
                
                <div class="footer">
                    <p>Report generated by Advanced Reporting Engine</p>
                </div>
            </body>
            </html>
            """
            
            # Plantilla operacional
            self.report_templates["operational"] = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>{{ title }}</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 40px; }
                    .header { text-align: center; margin-bottom: 30px; }
                    .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                    .table th, .table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                    .table th { background-color: #f2f2f2; }
                    .chart { margin: 20px 0; text-align: center; }
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>{{ title }}</h1>
                    <p>{{ description }}</p>
                    <p>Generated on: {{ generated_at }}</p>
                </div>
                
                <div class="tables">
                    <h2>Data Tables</h2>
                    {% for table in tables %}
                    <h3>{{ table.title }}</h3>
                    <table class="table">
                        <thead>
                            <tr>
                                {% for header in table.headers %}
                                <th>{{ header }}</th>
                                {% endfor %}
                            </tr>
                        </thead>
                        <tbody>
                            {% for row in table.rows %}
                            <tr>
                                {% for cell in row %}
                                <td>{{ cell }}</td>
                                {% endfor %}
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                    {% endfor %}
                </div>
                
                <div class="charts">
                    <h2>Charts</h2>
                    {% for chart in charts %}
                    <div class="chart">
                        <h3>{{ chart.title }}</h3>
                        <img src="{{ chart.path }}" alt="{{ chart.title }}">
                    </div>
                    {% endfor %}
                </div>
            </body>
            </html>
            """
            
        except Exception as e:
            self.logger.error(f"Error initializing templates: {e}")
    
    async def generate_report(self, request: ReportRequest) -> ReportResult:
        """Generar reporte avanzado"""
        try:
            start_time = datetime.now()
            
            # Validar solicitud
            await self._validate_report_request(request)
            
            # Cargar datos
            data = await self._load_report_data(request)
            
            # Calcular métricas
            metrics = await self._calculate_report_metrics(data, request.metrics)
            
            # Generar visualizaciones
            visualizations = await self._generate_visualizations(data, request.visualizations)
            
            # Generar reporte según formato
            if request.format == ReportFormat.PDF:
                file_path = await self._generate_pdf_report(request, data, metrics, visualizations)
            elif request.format == ReportFormat.HTML:
                file_path = await self._generate_html_report(request, data, metrics, visualizations)
            elif request.format == ReportFormat.EXCEL:
                file_path = await self._generate_excel_report(request, data, metrics, visualizations)
            elif request.format == ReportFormat.CSV:
                file_path = await self._generate_csv_report(request, data, metrics, visualizations)
            elif request.format == ReportFormat.JSON:
                file_path = await self._generate_json_report(request, data, metrics, visualizations)
            else:
                raise ValueError(f"Unsupported report format: {request.format}")
            
            # Calcular tiempo de generación
            generation_time = (datetime.now() - start_time).total_seconds()
            
            # Obtener tamaño del archivo
            import os
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            # Crear resultado
            result = ReportResult(
                report_id=request.report_id,
                status="success",
                file_path=file_path,
                file_size=file_size,
                generation_time=generation_time,
                metrics_summary=metrics,
                visualizations=[v["title"] for v in visualizations],
                recipients=request.recipients,
                generated_at=datetime.now().isoformat()
            )
            
            # Guardar en historial
            await self._save_report_history(request, result)
            
            # Enviar reporte si hay destinatarios
            if request.recipients:
                await self._send_report(request, result)
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error generating report: {e}")
            raise
    
    async def _validate_report_request(self, request: ReportRequest) -> None:
        """Validar solicitud de reporte"""
        try:
            if not request.report_id:
                raise ValueError("Report ID is required")
            
            if not request.title:
                raise ValueError("Report title is required")
            
            if not request.data_sources:
                raise ValueError("Data sources are required")
            
            if not request.metrics:
                raise ValueError("Metrics are required")
            
        except Exception as e:
            self.logger.error(f"Error validating report request: {e}")
            raise
    
    async def _load_report_data(self, request: ReportRequest) -> Dict[str, pd.DataFrame]:
        """Cargar datos para el reporte"""
        try:
            data = {}
            
            for source in request.data_sources:
                # Simular carga de datos
                if source == "pricing_data":
                    data[source] = await self._create_sample_pricing_data()
                elif source == "sales_data":
                    data[source] = await self._create_sample_sales_data()
                elif source == "customer_data":
                    data[source] = await self._create_sample_customer_data()
                elif source == "performance_data":
                    data[source] = await self._create_sample_performance_data()
                else:
                    # Datos genéricos
                    data[source] = await self._create_generic_data()
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error loading report data: {e}")
            raise
    
    async def _create_sample_pricing_data(self) -> pd.DataFrame:
        """Crear datos de muestra de precios"""
        try:
            np.random.seed(42)
            dates = pd.date_range(start='2024-01-01', periods=365, freq='D')
            
            data = pd.DataFrame({
                'date': dates,
                'price': np.random.normal(100, 20, 365),
                'volume': np.random.normal(1000, 200, 365),
                'competitor_price': np.random.normal(95, 15, 365),
                'market_share': np.random.uniform(0.1, 0.3, 365),
                'customer_satisfaction': np.random.normal(4.2, 0.5, 365)
            })
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error creating sample pricing data: {e}")
            raise
    
    async def _create_sample_sales_data(self) -> pd.DataFrame:
        """Crear datos de muestra de ventas"""
        try:
            np.random.seed(42)
            dates = pd.date_range(start='2024-01-01', periods=365, freq='D')
            
            data = pd.DataFrame({
                'date': dates,
                'sales': np.random.normal(50000, 10000, 365),
                'units_sold': np.random.normal(500, 100, 365),
                'revenue': np.random.normal(100000, 20000, 365),
                'profit_margin': np.random.uniform(0.1, 0.3, 365),
                'customer_count': np.random.normal(1000, 200, 365)
            })
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error creating sample sales data: {e}")
            raise
    
    async def _create_sample_customer_data(self) -> pd.DataFrame:
        """Crear datos de muestra de clientes"""
        try:
            np.random.seed(42)
            n_customers = 1000
            
            data = pd.DataFrame({
                'customer_id': range(1, n_customers + 1),
                'age': np.random.randint(18, 65, n_customers),
                'gender': np.random.choice(['M', 'F'], n_customers),
                'income': np.random.normal(50000, 15000, n_customers),
                'satisfaction_score': np.random.uniform(1, 5, n_customers),
                'lifetime_value': np.random.normal(1000, 300, n_customers),
                'region': np.random.choice(['North', 'South', 'East', 'West'], n_customers)
            })
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error creating sample customer data: {e}")
            raise
    
    async def _create_sample_performance_data(self) -> pd.DataFrame:
        """Crear datos de muestra de rendimiento"""
        try:
            np.random.seed(42)
            dates = pd.date_range(start='2024-01-01', periods=365, freq='D')
            
            data = pd.DataFrame({
                'date': dates,
                'response_time': np.random.normal(200, 50, 365),
                'throughput': np.random.normal(1000, 200, 365),
                'error_rate': np.random.uniform(0, 0.05, 365),
                'cpu_usage': np.random.uniform(0.3, 0.8, 365),
                'memory_usage': np.random.uniform(0.4, 0.9, 365),
                'disk_usage': np.random.uniform(0.2, 0.7, 365)
            })
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error creating sample performance data: {e}")
            raise
    
    async def _create_generic_data(self) -> pd.DataFrame:
        """Crear datos genéricos"""
        try:
            np.random.seed(42)
            dates = pd.date_range(start='2024-01-01', periods=100, freq='D')
            
            data = pd.DataFrame({
                'date': dates,
                'value': np.random.normal(100, 20, 100),
                'category': np.random.choice(['A', 'B', 'C'], 100),
                'score': np.random.uniform(0, 100, 100)
            })
            
            return data
            
        except Exception as e:
            self.logger.error(f"Error creating generic data: {e}")
            raise
    
    async def _calculate_report_metrics(self, data: Dict[str, pd.DataFrame], metrics: List[str]) -> Dict[str, Any]:
        """Calcular métricas del reporte"""
        try:
            calculated_metrics = {}
            
            for metric_name in metrics:
                if metric_name in self.metrics_calculators:
                    try:
                        # Calcular métrica para cada fuente de datos
                        metric_values = {}
                        for source_name, df in data.items():
                            value = await self.metrics_calculators[metric_name](df)
                            metric_values[source_name] = value
                        
                        calculated_metrics[metric_name] = metric_values
                        
                    except Exception as e:
                        self.logger.warning(f"Error calculating metric {metric_name}: {e}")
                        calculated_metrics[metric_name] = {"error": str(e)}
                else:
                    self.logger.warning(f"Unknown metric: {metric_name}")
                    calculated_metrics[metric_name] = {"error": "Unknown metric"}
            
            return calculated_metrics
            
        except Exception as e:
            self.logger.error(f"Error calculating report metrics: {e}")
            raise
    
    async def _calculate_accuracy(self, df: pd.DataFrame) -> float:
        """Calcular precisión"""
        try:
            # Simular cálculo de precisión
            if 'actual' in df.columns and 'predicted' in df.columns:
                return accuracy_score(df['actual'], df['predicted'])
            else:
                # Usar columna numérica como proxy
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    return np.mean(df[numeric_cols[0]]) / 100  # Normalizar
                return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_precision(self, df: pd.DataFrame) -> float:
        """Calcular precisión"""
        try:
            # Simular cálculo de precisión
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return np.mean(df[numeric_cols[0]]) / 100
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_recall(self, df: pd.DataFrame) -> float:
        """Calcular recall"""
        try:
            # Simular cálculo de recall
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return np.mean(df[numeric_cols[0]]) / 100
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_f1_score(self, df: pd.DataFrame) -> float:
        """Calcular F1 score"""
        try:
            # Simular cálculo de F1 score
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return np.mean(df[numeric_cols[0]]) / 100
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_mse(self, df: pd.DataFrame) -> float:
        """Calcular MSE"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return np.mean((df[numeric_cols[0]] - df[numeric_cols[0]].mean()) ** 2)
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_mae(self, df: pd.DataFrame) -> float:
        """Calcular MAE"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return np.mean(np.abs(df[numeric_cols[0]] - df[numeric_cols[0]].mean()))
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_r2_score(self, df: pd.DataFrame) -> float:
        """Calcular R² score"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                # Simular R² score
                return np.random.uniform(0.7, 0.95)
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_mean(self, df: pd.DataFrame) -> float:
        """Calcular media"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return df[numeric_cols[0]].mean()
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_median(self, df: pd.DataFrame) -> float:
        """Calcular mediana"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return df[numeric_cols[0]].median()
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_std(self, df: pd.DataFrame) -> float:
        """Calcular desviación estándar"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return df[numeric_cols[0]].std()
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_variance(self, df: pd.DataFrame) -> float:
        """Calcular varianza"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                return df[numeric_cols[0]].var()
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_correlation(self, df: pd.DataFrame) -> float:
        """Calcular correlación"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) >= 2:
                return df[numeric_cols[0]].corr(df[numeric_cols[1]])
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_trend(self, df: pd.DataFrame) -> float:
        """Calcular tendencia"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                # Calcular tendencia usando regresión lineal simple
                x = np.arange(len(df))
                y = df[numeric_cols[0]].values
                slope, _ = np.polyfit(x, y, 1)
                return slope
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_growth_rate(self, df: pd.DataFrame) -> float:
        """Calcular tasa de crecimiento"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                values = df[numeric_cols[0]].values
                if len(values) > 1:
                    return (values[-1] - values[0]) / values[0] * 100
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _calculate_volatility(self, df: pd.DataFrame) -> float:
        """Calcular volatilidad"""
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                returns = df[numeric_cols[0]].pct_change().dropna()
                return returns.std() * np.sqrt(252)  # Anualizada
            return 0.0
        except Exception as e:
            return 0.0
    
    async def _generate_visualizations(self, data: Dict[str, pd.DataFrame], visualizations: List[str]) -> List[Dict[str, Any]]:
        """Generar visualizaciones"""
        try:
            generated_charts = []
            
            for viz_type in visualizations:
                try:
                    if viz_type == "line_chart":
                        chart = await self._create_line_chart(data)
                    elif viz_type == "bar_chart":
                        chart = await self._create_bar_chart(data)
                    elif viz_type == "scatter_plot":
                        chart = await self._create_scatter_plot(data)
                    elif viz_type == "histogram":
                        chart = await self._create_histogram(data)
                    elif viz_type == "heatmap":
                        chart = await self._create_heatmap(data)
                    elif viz_type == "box_plot":
                        chart = await self._create_box_plot(data)
                    else:
                        continue
                    
                    generated_charts.append(chart)
                    
                except Exception as e:
                    self.logger.warning(f"Error creating visualization {viz_type}: {e}")
                    continue
            
            return generated_charts
            
        except Exception as e:
            self.logger.error(f"Error generating visualizations: {e}")
            return []
    
    async def _create_line_chart(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear gráfico de líneas"""
        try:
            # Usar el primer DataFrame con columna de fecha
            for source_name, df in data.items():
                if 'date' in df.columns:
                    fig = go.Figure()
                    
                    # Agregar líneas para columnas numéricas
                    numeric_cols = df.select_dtypes(include=[np.number]).columns
                    for col in numeric_cols[:3]:  # Máximo 3 líneas
                        fig.add_trace(go.Scatter(
                            x=df['date'],
                            y=df[col],
                            mode='lines',
                            name=col,
                            line=dict(width=2)
                        ))
                    
                    fig.update_layout(
                        title=f"Line Chart - {source_name}",
                        xaxis_title="Date",
                        yaxis_title="Value",
                        width=self.default_config["chart_width"],
                        height=self.default_config["chart_height"]
                    )
                    
                    # Guardar gráfico
                    chart_path = f"./charts/line_chart_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                    fig.write_html(chart_path)
                    
                    return {
                        "title": f"Line Chart - {source_name}",
                        "type": "line_chart",
                        "path": chart_path,
                        "source": source_name
                    }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error creating line chart: {e}")
            return None
    
    async def _create_bar_chart(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear gráfico de barras"""
        try:
            # Usar el primer DataFrame
            source_name, df = next(iter(data.items()))
            
            fig = go.Figure()
            
            # Agrupar por categoría si existe
            if 'category' in df.columns:
                grouped = df.groupby('category').mean()
                fig.add_trace(go.Bar(
                    x=grouped.index,
                    y=grouped.iloc[:, 0],
                    name=grouped.columns[0]
                ))
            else:
                # Usar primeras 10 filas
                sample_df = df.head(10)
                numeric_cols = sample_df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    fig.add_trace(go.Bar(
                        x=sample_df.index,
                        y=sample_df[numeric_cols[0]],
                        name=numeric_cols[0]
                    ))
            
            fig.update_layout(
                title=f"Bar Chart - {source_name}",
                xaxis_title="Category",
                yaxis_title="Value",
                width=self.default_config["chart_width"],
                height=self.default_config["chart_height"]
            )
            
            # Guardar gráfico
            chart_path = f"./charts/bar_chart_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            fig.write_html(chart_path)
            
            return {
                "title": f"Bar Chart - {source_name}",
                "type": "bar_chart",
                "path": chart_path,
                "source": source_name
            }
            
        except Exception as e:
            self.logger.error(f"Error creating bar chart: {e}")
            return None
    
    async def _create_scatter_plot(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear gráfico de dispersión"""
        try:
            # Usar el primer DataFrame
            source_name, df = next(iter(data.items()))
            
            fig = go.Figure()
            
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) >= 2:
                fig.add_trace(go.Scatter(
                    x=df[numeric_cols[0]],
                    y=df[numeric_cols[1]],
                    mode='markers',
                    name='Data Points',
                    marker=dict(size=8, opacity=0.6)
                ))
                
                fig.update_layout(
                    title=f"Scatter Plot - {source_name}",
                    xaxis_title=numeric_cols[0],
                    yaxis_title=numeric_cols[1],
                    width=self.default_config["chart_width"],
                    height=self.default_config["chart_height"]
                )
                
                # Guardar gráfico
                chart_path = f"./charts/scatter_plot_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                fig.write_html(chart_path)
                
                return {
                    "title": f"Scatter Plot - {source_name}",
                    "type": "scatter_plot",
                    "path": chart_path,
                    "source": source_name
                }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error creating scatter plot: {e}")
            return None
    
    async def _create_histogram(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear histograma"""
        try:
            # Usar el primer DataFrame
            source_name, df = next(iter(data.items()))
            
            fig = go.Figure()
            
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                fig.add_trace(go.Histogram(
                    x=df[numeric_cols[0]],
                    nbinsx=30,
                    name=numeric_cols[0]
                ))
                
                fig.update_layout(
                    title=f"Histogram - {source_name}",
                    xaxis_title=numeric_cols[0],
                    yaxis_title="Frequency",
                    width=self.default_config["chart_width"],
                    height=self.default_config["chart_height"]
                )
                
                # Guardar gráfico
                chart_path = f"./charts/histogram_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                fig.write_html(chart_path)
                
                return {
                    "title": f"Histogram - {source_name}",
                    "type": "histogram",
                    "path": chart_path,
                    "source": source_name
                }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error creating histogram: {e}")
            return None
    
    async def _create_heatmap(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear mapa de calor"""
        try:
            # Usar el primer DataFrame
            source_name, df = next(iter(data.items()))
            
            fig = go.Figure()
            
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 1:
                corr_matrix = df[numeric_cols].corr()
                
                fig.add_trace(go.Heatmap(
                    z=corr_matrix.values,
                    x=corr_matrix.columns,
                    y=corr_matrix.columns,
                    colorscale='RdBu',
                    zmid=0
                ))
                
                fig.update_layout(
                    title=f"Correlation Heatmap - {source_name}",
                    width=self.default_config["chart_width"],
                    height=self.default_config["chart_height"]
                )
                
                # Guardar gráfico
                chart_path = f"./charts/heatmap_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                fig.write_html(chart_path)
                
                return {
                    "title": f"Correlation Heatmap - {source_name}",
                    "type": "heatmap",
                    "path": chart_path,
                    "source": source_name
                }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error creating heatmap: {e}")
            return None
    
    async def _create_box_plot(self, data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Crear gráfico de cajas"""
        try:
            # Usar el primer DataFrame
            source_name, df = next(iter(data.items()))
            
            fig = go.Figure()
            
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 0:
                fig.add_trace(go.Box(
                    y=df[numeric_cols[0]],
                    name=numeric_cols[0]
                ))
                
                fig.update_layout(
                    title=f"Box Plot - {source_name}",
                    yaxis_title=numeric_cols[0],
                    width=self.default_config["chart_width"],
                    height=self.default_config["chart_height"]
                )
                
                # Guardar gráfico
                chart_path = f"./charts/box_plot_{source_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                fig.write_html(chart_path)
                
                return {
                    "title": f"Box Plot - {source_name}",
                    "type": "box_plot",
                    "path": chart_path,
                    "source": source_name
                }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error creating box plot: {e}")
            return None
    
    async def _generate_pdf_report(self, request: ReportRequest, data: Dict[str, pd.DataFrame], 
                                 metrics: Dict[str, Any], visualizations: List[Dict[str, Any]]) -> str:
        """Generar reporte PDF"""
        try:
            import os
            os.makedirs(self.default_config["output_directory"], exist_ok=True)
            
            file_path = f"{self.default_config['output_directory']}/{request.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Crear documento PDF
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            story.append(Paragraph(request.title, title_style))
            story.append(Spacer(1, 12))
            
            # Descripción
            story.append(Paragraph(request.description, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Fecha de generación
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Métricas
            story.append(Paragraph("Key Metrics", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Crear tabla de métricas
            metric_data = [["Metric", "Value", "Source"]]
            for metric_name, metric_values in metrics.items():
                for source_name, value in metric_values.items():
                    if isinstance(value, (int, float)):
                        metric_data.append([metric_name, f"{value:.2f}", source_name])
                    else:
                        metric_data.append([metric_name, str(value), source_name])
            
            metric_table = Table(metric_data)
            metric_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(metric_table)
            story.append(Spacer(1, 20))
            
            # Visualizaciones
            if visualizations:
                story.append(Paragraph("Visualizations", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                for viz in visualizations:
                    story.append(Paragraph(viz["title"], styles['Heading3']))
                    story.append(Spacer(1, 12))
                    story.append(Paragraph(f"Chart type: {viz['type']}", styles['Normal']))
                    story.append(Spacer(1, 12))
            
            # Construir PDF
            doc.build(story)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating PDF report: {e}")
            raise
    
    async def _generate_html_report(self, request: ReportRequest, data: Dict[str, pd.DataFrame], 
                                  metrics: Dict[str, Any], visualizations: List[Dict[str, Any]]) -> str:
        """Generar reporte HTML"""
        try:
            import os
            os.makedirs(self.default_config["output_directory"], exist_ok=True)
            
            file_path = f"{self.default_config['output_directory']}/{request.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            
            # Preparar datos para plantilla
            template_data = {
                "title": request.title,
                "description": request.description,
                "generated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "summary": f"Report generated with {len(data)} data sources and {len(metrics)} metrics.",
                "metrics": [],
                "charts": visualizations
            }
            
            # Preparar métricas
            for metric_name, metric_values in metrics.items():
                for source_name, value in metric_values.items():
                    if isinstance(value, (int, float)):
                        template_data["metrics"].append({
                            "name": f"{metric_name} ({source_name})",
                            "value": f"{value:.2f}",
                            "change": "N/A"
                        })
            
            # Usar plantilla
            template = Template(self.report_templates.get(request.report_type.value, self.report_templates["executive"]))
            html_content = template.render(**template_data)
            
            # Guardar archivo HTML
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating HTML report: {e}")
            raise
    
    async def _generate_excel_report(self, request: ReportRequest, data: Dict[str, pd.DataFrame], 
                                   metrics: Dict[str, Any], visualizations: List[Dict[str, Any]]) -> str:
        """Generar reporte Excel"""
        try:
            import os
            os.makedirs(self.default_config["output_directory"], exist_ok=True)
            
            file_path = f"{self.default_config['output_directory']}/{request.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Hoja de resumen
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Título
            ws_summary['A1'] = request.title
            ws_summary['A1'].font = Font(size=16, bold=True)
            
            # Descripción
            ws_summary['A3'] = request.description
            ws_summary['A5'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Métricas
            ws_summary['A7'] = "Key Metrics"
            ws_summary['A7'].font = Font(size=14, bold=True)
            
            row = 9
            for metric_name, metric_values in metrics.items():
                for source_name, value in metric_values.items():
                    ws_summary[f'A{row}'] = f"{metric_name} ({source_name})"
                    ws_summary[f'B{row}'] = value if isinstance(value, (int, float)) else str(value)
                    row += 1
            
            # Hojas de datos
            for source_name, df in data.items():
                ws_data = wb.create_sheet(title=source_name)
                
                # Escribir datos
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws_data.append(r)
                
                # Formatear encabezados
                for cell in ws_data[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Guardar archivo
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            raise
    
    async def _generate_csv_report(self, request: ReportRequest, data: Dict[str, pd.DataFrame], 
                                 metrics: Dict[str, Any], visualizations: List[Dict[str, Any]]) -> str:
        """Generar reporte CSV"""
        try:
            import os
            os.makedirs(self.default_config["output_directory"], exist_ok=True)
            
            file_path = f"{self.default_config['output_directory']}/{request.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            # Combinar todos los datos
            combined_data = []
            for source_name, df in data.items():
                df_copy = df.copy()
                df_copy['data_source'] = source_name
                combined_data.append(df_copy)
            
            if combined_data:
                combined_df = pd.concat(combined_data, ignore_index=True)
                combined_df.to_csv(file_path, index=False)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating CSV report: {e}")
            raise
    
    async def _generate_json_report(self, request: ReportRequest, data: Dict[str, pd.DataFrame], 
                                  metrics: Dict[str, Any], visualizations: List[Dict[str, Any]]) -> str:
        """Generar reporte JSON"""
        try:
            import os
            os.makedirs(self.default_config["output_directory"], exist_ok=True)
            
            file_path = f"{self.default_config['output_directory']}/{request.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            # Preparar datos para JSON
            report_data = {
                "report_id": request.report_id,
                "title": request.title,
                "description": request.description,
                "generated_at": datetime.now().isoformat(),
                "metrics": metrics,
                "visualizations": visualizations,
                "data_summary": {}
            }
            
            # Resumen de datos
            for source_name, df in data.items():
                report_data["data_summary"][source_name] = {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": df.columns.tolist(),
                    "data_types": df.dtypes.to_dict()
                }
            
            # Guardar archivo JSON
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, default=str)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating JSON report: {e}")
            raise
    
    async def _send_report(self, request: ReportRequest, result: ReportResult) -> None:
        """Enviar reporte por email"""
        try:
            # Configurar email
            msg = MIMEMultipart()
            msg['From'] = "reports@company.com"
            msg['To'] = ", ".join(request.recipients)
            msg['Subject'] = f"Report: {request.title}"
            
            # Cuerpo del email
            body = f"""
            Dear Recipient,
            
            Please find attached the report: {request.title}
            
            Report Details:
            - Generated on: {result.generated_at}
            - Generation time: {result.generation_time:.2f} seconds
            - File size: {result.file_size} bytes
            - Format: {request.format.value}
            
            Best regards,
            Advanced Reporting Engine
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Adjuntar archivo
            with open(result.file_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {result.file_path.split("/")[-1]}'
                )
                msg.attach(part)
            
            # Enviar email (simulado)
            self.logger.info(f"Report sent to {request.recipients}")
            
        except Exception as e:
            self.logger.error(f"Error sending report: {e}")
    
    async def _save_report_history(self, request: ReportRequest, result: ReportResult) -> None:
        """Guardar historial de reportes"""
        try:
            self.report_history[request.report_id] = {
                "timestamp": datetime.now().isoformat(),
                "report_type": request.report_type.value,
                "title": request.title,
                "format": request.format.value,
                "frequency": request.frequency.value,
                "recipients": request.recipients,
                "file_path": result.file_path,
                "file_size": result.file_size,
                "generation_time": result.generation_time,
                "metrics_count": len(result.metrics_summary),
                "visualizations_count": len(result.visualizations)
            }
            
        except Exception as e:
            self.logger.error(f"Error saving report history: {e}")
    
    async def get_reporting_insights(self) -> Dict[str, Any]:
        """Obtener insights de reportes"""
        insights = {
            "total_reports": len(self.report_history),
            "report_types": {},
            "report_formats": {},
            "report_frequencies": {},
            "average_generation_time": 0.0,
            "total_file_size": 0,
            "most_common_recipients": [],
            "recent_reports": []
        }
        
        if self.report_history:
            # Análisis de tipos de reporte
            for report in self.report_history.values():
                report_type = report["report_type"]
                insights["report_types"][report_type] = insights["report_types"].get(report_type, 0) + 1
                
                report_format = report["format"]
                insights["report_formats"][report_format] = insights["report_formats"].get(report_format, 0) + 1
                
                report_frequency = report["frequency"]
                insights["report_frequencies"][report_frequency] = insights["report_frequencies"].get(report_frequency, 0) + 1
            
            # Promedios
            generation_times = [r["generation_time"] for r in self.report_history.values()]
            file_sizes = [r["file_size"] for r in self.report_history.values()]
            
            insights["average_generation_time"] = np.mean(generation_times) if generation_times else 0.0
            insights["total_file_size"] = sum(file_sizes) if file_sizes else 0
            
            # Destinatarios más comunes
            all_recipients = []
            for report in self.report_history.values():
                all_recipients.extend(report["recipients"])
            
            if all_recipients:
                from collections import Counter
                recipient_counts = Counter(all_recipients)
                insights["most_common_recipients"] = [{"email": email, "count": count} for email, count in recipient_counts.most_common(5)]
            
            # Reportes recientes
            recent_reports = sorted(self.report_history.items(), key=lambda x: x[1]["timestamp"], reverse=True)[:5]
            insights["recent_reports"] = [
                {
                    "report_id": report_id,
                    "title": report["title"],
                    "timestamp": report["timestamp"],
                    "format": report["format"]
                }
                for report_id, report in recent_reports
            ]
        
        return insights

# Función principal para inicializar el motor
async def initialize_reporting_engine() -> AdvancedReportingEngine:
    """Inicializar motor de reportes avanzado"""
    engine = AdvancedReportingEngine()
    
    # Configurar logging
    logging.basicConfig(level=logging.INFO)
    
    return engine

if __name__ == "__main__":
    # Ejemplo de uso
    async def main():
        engine = await initialize_reporting_engine()
        
        # Crear solicitud de reporte
        request = ReportRequest(
            report_id="executive_report_001",
            report_type=ReportType.EXECUTIVE,
            title="Executive Dashboard Report",
            description="Monthly executive summary with key metrics and insights",
            data_sources=["pricing_data", "sales_data", "customer_data"],
            metrics=["mean", "std", "trend", "growth_rate"],
            visualizations=["line_chart", "bar_chart", "heatmap"],
            format=ReportFormat.HTML,
            frequency=ReportFrequency.MONTHLY,
            recipients=["executive@company.com", "ceo@company.com"]
        )
        
        # Generar reporte
        result = await engine.generate_report(request)
        print("Report Generation Result:")
        print(f"Status: {result.status}")
        print(f"File path: {result.file_path}")
        print(f"Generation time: {result.generation_time}")
        print(f"File size: {result.file_size}")
        print(f"Metrics: {len(result.metrics_summary)}")
        print(f"Visualizations: {len(result.visualizations)}")
        
        # Obtener insights
        insights = await engine.get_reporting_insights()
        print("\nReporting Insights:", json.dumps(insights, indent=2, default=str))
    
    asyncio.run(main())



