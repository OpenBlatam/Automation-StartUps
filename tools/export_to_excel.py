#!/usr/bin/env python3
"""
Exporta datos de creativos a Excel con formato avanzado, gráficos y análisis
"""
import csv
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl no está instalado. Instala con: pip install openpyxl")
    print("   Exportando CSV básico en su lugar...")

def load_creatives():
    """Carga creativos desde CSV Master"""
    script_dir = Path(__file__).parent
    root_dir = script_dir.parent
    csv_path = root_dir / 'docs' / 'LINKEDIN_ADS_CREATIVES_MASTER.csv'
    
    if not csv_path.exists():
        return None
    
    creatives = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            creatives.append(row)
    
    return creatives

def analyze_data(creatives):
    """Analiza datos para resúmenes"""
    by_format = defaultdict(int)
    by_angle = defaultdict(int)
    by_product = defaultdict(int)
    
    for creative in creatives:
        by_format[creative.get('formato', 'unknown')] += 1
        by_angle[creative.get('angulo', 'unknown')] += 1
        by_product[creative.get('producto', 'unknown')] += 1
    
    return {
        'by_format': dict(by_format),
        'by_angle': dict(by_angle),
        'by_product': dict(by_product),
        'total': len(creatives)
    }

def create_excel_export(creatives, analysis, output_path):
    """Crea Excel con formato avanzado"""
    wb = Workbook()
    
    # Hoja 1: Datos Completos
    ws_data = wb.active
    ws_data.title = "Creativos"
    
    if creatives:
        headers = list(creatives[0].keys())
        ws_data.append(headers)
        
        # Formato header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Datos
        for row_idx, creative in enumerate(creatives, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws_data.cell(row=row_idx, column=col_idx)
                cell.value = creative.get(header, '')
                cell.border = border_style
                if col_idx == 1:  # Primera columna en negrita
                    cell.font = Font(bold=True)
        
        # Ajustar ancho de columnas
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws_data.column_dimensions[col_letter].width = max(len(header) + 2, 15)
    
    # Hoja 2: Resumen por Formato
    ws_format = wb.create_sheet("Resumen Formato")
    ws_format.append(["Formato", "Cantidad", "Porcentaje"])
    
    format_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for row in ws_format.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = format_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    total = analysis['total']
    for formato, count in sorted(analysis['by_format'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total * 100) if total > 0 else 0
        ws_format.append([formato, count, f"{pct:.1f}%"])
    
    # Ajustar columnas
    for col in ['A', 'B', 'C']:
        ws_format.column_dimensions[col].width = 20
    
    # Hoja 3: Resumen por Ángulo
    ws_angle = wb.create_sheet("Resumen Ángulo")
    ws_angle.append(["Ángulo", "Cantidad", "Porcentaje"])
    
    for row in ws_angle.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = format_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    for angle, count in sorted(analysis['by_angle'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total * 100) if total > 0 else 0
        ws_angle.append([angle, count, f"{pct:.1f}%"])
    
    for col in ['A', 'B', 'C']:
        ws_angle.column_dimensions[col].width = 20
    
    # Hoja 4: Resumen por Producto
    ws_product = wb.create_sheet("Resumen Producto")
    ws_product.append(["Producto", "Cantidad", "Porcentaje"])
    
    for row in ws_product.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = format_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    for product, count in sorted(analysis['by_product'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total * 100) if total > 0 else 0
        ws_product.append([product, count, f"{pct:.1f}%"])
    
    for col in ['A', 'B', 'C']:
        ws_product.column_dimensions[col].width = 20
    
    # Hoja 5: Dashboard (Métricas Principales)
    ws_dashboard = wb.create_sheet("Dashboard")
    
    # Título
    ws_dashboard['A1'] = "📊 Dashboard de Creativos"
    ws_dashboard['A1'].font = Font(size=16, bold=True)
    ws_dashboard.merge_cells('A1:C1')
    
    # Métricas
    row = 3
    ws_dashboard.cell(row, 1, "Total Creativos:").font = Font(bold=True)
    ws_dashboard.cell(row, 2, total)
    
    row += 1
    ws_dashboard.cell(row, 1, "Formatos Únicos:").font = Font(bold=True)
    ws_dashboard.cell(row, 2, len(analysis['by_format']))
    
    row += 1
    ws_dashboard.cell(row, 1, "Ángulos Únicos:").font = Font(bold=True)
    ws_dashboard.cell(row, 2, len(analysis['by_angle']))
    
    row += 1
    ws_dashboard.cell(row, 1, "Productos Únicos:").font = Font(bold=True)
    ws_dashboard.cell(row, 2, len(analysis['by_product']))
    
    row += 3
    ws_dashboard.cell(row, 1, "Fecha Exportación:").font = Font(bold=True)
    ws_dashboard.cell(row, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    # Ajustar columnas
    ws_dashboard.column_dimensions['A'].width = 25
    ws_dashboard.column_dimensions['B'].width = 15
    
    # Guardar
    wb.save(output_path)
    return True

def create_csv_fallback(creatives, analysis, output_path):
    """Crea CSV simple si openpyxl no está disponible"""
    with open(output_path.with_suffix('.csv'), 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        
        # Resumen
        writer.writerow(["Resumen de Creativos"])
        writer.writerow(["Total", analysis['total']])
        writer.writerow([])
        
        # Por formato
        writer.writerow(["Distribución por Formato"])
        writer.writerow(["Formato", "Cantidad", "Porcentaje"])
        total = analysis['total']
        for formato, count in sorted(analysis['by_format'].items(), key=lambda x: x[1], reverse=True):
            pct = (count / total * 100) if total > 0 else 0
            writer.writerow([formato, count, f"{pct:.1f}%"])
        
        writer.writerow([])
        
        # Por ángulo
        writer.writerow(["Distribución por Ángulo"])
        writer.writerow(["Ángulo", "Cantidad", "Porcentaje"])
        for angle, count in sorted(analysis['by_angle'].items(), key=lambda x: x[1], reverse=True):
            pct = (count / total * 100) if total > 0 else 0
            writer.writerow([angle, count, f"{pct:.1f}%"])
    
    return True

def main():
    print("=" * 80)
    print("📊 Exportador a Excel")
    print("=" * 80)
    print()
    
    creatives = load_creatives()
    if not creatives:
        print("❌ No se encontró CSV Master")
        return
    
    print(f"✅ Cargados {len(creatives)} creativos")
    print()
    
    # Analizar datos
    print("📊 Analizando datos...")
    analysis = analyze_data(creatives)
    
    # Generar output path
    script_dir = Path(__file__).parent
    root_dir = script_dir.parent
    exports_dir = root_dir / 'exports'
    exports_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = exports_dir / f'creativos_export_{timestamp}.xlsx'
    
    # Exportar
    print(f"📝 Exportando a Excel...")
    
    if OPENPYXL_AVAILABLE:
        success = create_excel_export(creatives, analysis, output_path)
        if success:
            print(f"✅ Excel exportado: {output_path}")
            print()
            print("📋 Hojas incluidas:")
            print("   1. Creativos - Datos completos")
            print("   2. Resumen Formato - Distribución por formato")
            print("   3. Resumen Ángulo - Distribución por ángulo")
            print("   4. Resumen Producto - Distribución por producto")
            print("   5. Dashboard - Métricas principales")
        else:
            print("❌ Error al exportar")
    else:
        csv_path = output_path.with_suffix('.csv')
        success = create_csv_fallback(creatives, analysis, csv_path)
        if success:
            print(f"✅ CSV exportado: {csv_path}")
            print()
            print("💡 Para Excel con formato, instala: pip install openpyxl")
    
    print()

if __name__ == '__main__':
    main()

