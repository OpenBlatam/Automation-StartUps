"""
Exportación de Análisis de Mercado a Múltiples Formatos

Exporta análisis, insights y reportes a:
- Excel (con múltiples hojas)
- PDF (con gráficos)
- CSV
- JSON estructurado
"""

from __future__ import annotations

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
import io

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger(__name__)


class MarketExporter:
    """Exportador de análisis de mercado a múltiples formatos."""
    
    def __init__(self):
        """Inicializa el exportador."""
        self.logger = logging.getLogger(__name__)
    
    def export_to_excel(
        self,
        market_analysis: Dict[str, Any],
        insights: List[Dict[str, Any]],
        predictions: Optional[List[Dict[str, Any]]] = None,
        output_path: str = "/tmp/market_analysis.xlsx"
    ) -> str:
        """
        Exporta análisis a Excel con múltiples hojas.
        
        Args:
            market_analysis: Análisis completo
            insights: Lista de insights
            predictions: Predicciones ML (opcional)
            output_path: Ruta de salida
            
        Returns:
            Ruta del archivo generado
        """
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas and openpyxl required for Excel export")
        
        logger.info(f"Exporting to Excel: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            summary_data = {
                "Métrica": [
                    "Industria",
                    "Fecha de Análisis",
                    "Tendencias Analizadas",
                    "Insights Generados",
                    "Oportunidades",
                    "Riesgos",
                    "Predicciones ML"
                ],
                "Valor": [
                    market_analysis.get("industry", "N/A"),
                    market_analysis.get("analysis_date", datetime.utcnow().isoformat()),
                    len(market_analysis.get("trends", [])),
                    len(insights),
                    len(market_analysis.get("opportunities", [])),
                    len(market_analysis.get("risk_factors", [])),
                    len(predictions) if predictions else 0
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="Resumen", index=False)
            
            # Hoja 2: Tendencias
            if market_analysis.get("trends"):
                trends_data = []
                for trend in market_analysis["trends"]:
                    trends_data.append({
                        "Nombre": trend.get("trend_name", "N/A"),
                        "Categoría": trend.get("category", "N/A"),
                        "Valor Actual": trend.get("current_value", 0),
                        "Valor Anterior": trend.get("previous_value", 0),
                        "Cambio %": trend.get("change_percentage", 0),
                        "Dirección": trend.get("trend_direction", "N/A"),
                        "Confianza": trend.get("confidence", 0)
                    })
                df_trends = pd.DataFrame(trends_data)
                df_trends.to_excel(writer, sheet_name="Tendencias", index=False)
            
            # Hoja 3: Insights
            if insights:
                insights_data = []
                for insight in insights:
                    insights_data.append({
                        "Título": insight.get("title", "N/A"),
                        "Categoría": insight.get("category", "N/A"),
                        "Prioridad": insight.get("priority", "N/A"),
                        "Descripción": insight.get("description", "N/A"),
                        "Impacto Esperado": insight.get("expected_impact", "N/A"),
                        "Timeframe": insight.get("timeframe", "N/A"),
                        "Confianza": insight.get("confidence_score", 0),
                        "Pasos Accionables": "; ".join(insight.get("actionable_steps", []))
                    })
                df_insights = pd.DataFrame(insights_data)
                df_insights.to_excel(writer, sheet_name="Insights", index=False)
            
            # Hoja 4: Oportunidades
            if market_analysis.get("opportunities"):
                opps_data = []
                for opp in market_analysis["opportunities"]:
                    opps_data.append({
                        "Título": opp.get("title", "N/A"),
                        "Descripción": opp.get("description", "N/A"),
                        "Categoría": opp.get("category", "N/A"),
                        "Confianza": opp.get("confidence", 0)
                    })
                df_opps = pd.DataFrame(opps_data)
                df_opps.to_excel(writer, sheet_name="Oportunidades", index=False)
            
            # Hoja 5: Riesgos
            if market_analysis.get("risk_factors"):
                risks_data = []
                for risk in market_analysis["risk_factors"]:
                    risks_data.append({
                        "Título": risk.get("title", "N/A"),
                        "Descripción": risk.get("description", "N/A"),
                        "Categoría": risk.get("category", "N/A"),
                        "Confianza": risk.get("confidence", 0)
                    })
                df_risks = pd.DataFrame(risks_data)
                df_risks.to_excel(writer, sheet_name="Riesgos", index=False)
            
            # Hoja 6: Predicciones ML
            if predictions:
                preds_data = []
                for pred in predictions:
                    preds_data.append({
                        "Métrica": pred.get("metric_name", "N/A"),
                        "Valor Actual": pred.get("current_value", 0),
                        "Valor Predicho": pred.get("predicted_value", 0),
                        "Cambio %": pred.get("change_percentage", 0),
                        "Dirección": pred.get("trend_direction", "N/A"),
                        "Confianza": pred.get("confidence", 0),
                        "Timeframe (días)": pred.get("timeframe_days", 0)
                    })
                df_preds = pd.DataFrame(preds_data)
                df_preds.to_excel(writer, sheet_name="Predicciones ML", index=False)
        
        # Aplicar formato
        self._format_excel(output_path)
        
        logger.info(f"Excel export completed: {output_path}")
        return output_path
    
    def _format_excel(self, file_path: str):
        """Aplica formato a archivo Excel."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Formatear encabezados
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
        except Exception as e:
            logger.warning(f"Error formatting Excel: {e}")
    
    def export_to_pdf(
        self,
        market_analysis: Dict[str, Any],
        insights: List[Dict[str, Any]],
        executive_summary: Optional[str] = None,
        output_path: str = "/tmp/market_analysis.pdf"
    ) -> str:
        """
        Exporta análisis a PDF.
        
        Args:
            market_analysis: Análisis completo
            insights: Lista de insights
            executive_summary: Resumen ejecutivo (opcional)
            output_path: Ruta de salida
            
        Returns:
            Ruta del archivo generado
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab required for PDF export")
        
        logger.info(f"Exporting to PDF: {output_path}")
        
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30
        )
        story.append(Paragraph("Market Research Report", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Resumen ejecutivo
        if executive_summary:
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            for para in executive_summary.split('\n\n'):
                if para.strip():
                    story.append(Paragraph(para.strip(), styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))
            story.append(PageBreak())
        
        # Información general
        story.append(Paragraph("General Information", styles['Heading2']))
        info_data = [
            ["Industry:", market_analysis.get("industry", "N/A")],
            ["Analysis Date:", market_analysis.get("analysis_date", "N/A")],
            ["Timeframe:", f"{market_analysis.get('timeframe_months', 0)} months"],
            ["Trends Analyzed:", str(len(market_analysis.get("trends", [])))],
            ["Insights Generated:", str(len(insights))],
        ]
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Insights de alta prioridad
        high_priority = [i for i in insights if i.get("priority") == "high"]
        if high_priority:
            story.append(Paragraph("High Priority Insights", styles['Heading2']))
            for insight in high_priority[:5]:  # Top 5
                story.append(Paragraph(f"<b>{insight.get('title', 'N/A')}</b>", styles['Heading3']))
                story.append(Paragraph(insight.get("description", ""), styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
            story.append(PageBreak())
        
        # Todos los insights
        story.append(Paragraph("All Insights", styles['Heading2']))
        for i, insight in enumerate(insights, 1):
            story.append(Paragraph(f"{i}. {insight.get('title', 'N/A')}", styles['Heading3']))
            story.append(Paragraph(f"Priority: {insight.get('priority', 'N/A').upper()}", styles['Normal']))
            story.append(Paragraph(insight.get("description", ""), styles['Normal']))
            if insight.get("actionable_steps"):
                story.append(Paragraph("Actionable Steps:", styles['Heading4']))
                for step in insight.get("actionable_steps", []):
                    story.append(Paragraph(f"• {step}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Construir PDF
        doc.build(story)
        
        logger.info(f"PDF export completed: {output_path}")
        return output_path
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        output_path: str = "/tmp/market_data.csv"
    ) -> str:
        """Exporta datos a CSV."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas required for CSV export")
        
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8')
        logger.info(f"CSV export completed: {output_path}")
        return output_path
    
    def export_to_json(
        self,
        market_analysis: Dict[str, Any],
        insights: List[Dict[str, Any]],
        predictions: Optional[List[Dict[str, Any]]] = None,
        output_path: str = "/tmp/market_analysis.json"
    ) -> str:
        """Exporta análisis completo a JSON."""
        export_data = {
            "export_date": datetime.utcnow().isoformat(),
            "market_analysis": market_analysis,
            "insights": insights,
            "predictions": predictions or []
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON export completed: {output_path}")
        return output_path






