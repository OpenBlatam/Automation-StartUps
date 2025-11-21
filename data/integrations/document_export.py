"""
Exportador Avanzado de Documentos
==================================

Exporta documentos procesados a múltiples formatos.
"""

from typing import Dict, Any, List, Optional
from pathlib import Path
import logging
import json
import csv
from datetime import datetime
from io import BytesIO

logger = logging.getLogger(__name__)


class DocumentExporter:
    """Exportador avanzado de documentos"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def export_to_excel(
        self,
        documents: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """Exporta a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl es requerido. Instala con: pip install openpyxl"
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Documentos Procesados"
        
        if not documents:
            wb.save(output_path)
            return output_path
        
        # Obtener todas las columnas posibles
        all_columns = set()
        for doc in documents:
            all_columns.update(doc.keys())
            if "extracted_fields" in doc:
                all_columns.update(doc["extracted_fields"].keys())
        
        columns = [
            "document_id", "original_filename", "document_type",
            "classification_confidence", "ocr_confidence",
            "processed_at", "archive_path"
        ]
        
        # Agregar campos extraídos
        for col in sorted(all_columns):
            if col not in columns:
                columns.append(col)
        
        # Encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for row_idx, doc in enumerate(documents, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = doc.get(col_name)
                
                # Si no está en nivel superior, buscar en extracted_fields
                if value is None and "extracted_fields" in doc:
                    value = doc["extracted_fields"].get(col_name)
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        wb.save(output_path)
        self.logger.info(f"Exportado a Excel: {output_path}")
        return output_path
    
    def export_to_xml(
        self,
        documents: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """Exporta a XML"""
        try:
            import xml.etree.ElementTree as ET
        except ImportError:
            raise ImportError("xml.etree.ElementTree no disponible")
        
        root = ET.Element("documents")
        root.set("export_date", datetime.now().isoformat())
        root.set("count", str(len(documents)))
        
        for doc in documents:
            doc_elem = ET.SubElement(root, "document")
            
            for key, value in doc.items():
                if key == "extracted_fields":
                    fields_elem = ET.SubElement(doc_elem, "extracted_fields")
                    for field_name, field_value in value.items():
                        field_elem = ET.SubElement(fields_elem, "field")
                        field_elem.set("name", field_name)
                        field_elem.text = str(field_value)
                elif isinstance(value, (dict, list)):
                    # Serializar estructuras complejas como JSON
                    elem = ET.SubElement(doc_elem, key)
                    elem.text = json.dumps(value)
                else:
                    elem = ET.SubElement(doc_elem, key)
                    elem.text = str(value) if value is not None else ""
        
        tree = ET.ElementTree(root)
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
        
        self.logger.info(f"Exportado a XML: {output_path}")
        return output_path
    
    def export_to_jsonl(
        self,
        documents: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """Exporta a JSONL (una línea JSON por documento)"""
        with open(output_path, 'w', encoding='utf-8') as f:
            for doc in documents:
                json.dump(doc, f, ensure_ascii=False)
                f.write('\n')
        
        self.logger.info(f"Exportado a JSONL: {output_path}")
        return output_path
    
    def export_to_pdf_report(
        self,
        documents: List[Dict[str, Any]],
        output_path: str,
        title: str = "Reporte de Documentos"
    ) -> str:
        """Exporta reporte a PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
        except ImportError:
            raise ImportError(
                "reportlab es requerido. Instala con: pip install reportlab"
            )
        
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title_para = Paragraph(title, styles['Title'])
        story.append(title_para)
        story.append(Spacer(1, 0.2 * inch))
        
        # Resumen
        summary = f"""
        <b>Total de documentos:</b> {len(documents)}<br/>
        <b>Fecha de exportación:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        story.append(Paragraph(summary, styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))
        
        # Tabla de documentos
        if documents:
            data = [["ID", "Tipo", "Confianza", "Fecha"]]
            
            for doc in documents[:50]:  # Limitar a 50 para PDF
                data.append([
                    doc.get("document_id", "")[:20],
                    doc.get("document_type", ""),
                    f"{doc.get('classification_confidence', 0):.2%}",
                    doc.get("processed_at", "")[:10] if doc.get("processed_at") else ""
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        doc.build(story)
        self.logger.info(f"Exportado a PDF: {output_path}")
        return output_path
    
    def export_to_html_dashboard(
        self,
        documents: List[Dict[str, Any]],
        output_path: str,
        title: str = "Dashboard de Documentos"
    ) -> str:
        """Exporta dashboard HTML interactivo"""
        html = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .stat-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stat-value {{
            font-size: 2em;
            font-weight: bold;
            color: #667eea;
        }}
        .stat-label {{
            color: #666;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background: #667eea;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #eee;
        }}
        tr:hover {{
            background: #f5f5f5;
        }}
        .badge {{
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85em;
        }}
        .badge-invoice {{ background: #e3f2fd; color: #1976d2; }}
        .badge-contract {{ background: #f3e5f5; color: #7b1fa2; }}
        .badge-form {{ background: #e8f5e9; color: #388e3c; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{title}</h1>
        <p>Generado el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
    
    <div class="stats">
        <div class="stat-card">
            <div class="stat-value">{len(documents)}</div>
            <div class="stat-label">Total Documentos</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{len(set(d.get('document_type') for d in documents))}</div>
            <div class="stat-label">Tipos Únicos</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{sum(d.get('ocr_confidence', 0) for d in documents) / len(documents) * 100:.1f}%</div>
            <div class="stat-label">Confianza Promedio</div>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>ID</th>
                <th>Archivo</th>
                <th>Tipo</th>
                <th>Confianza</th>
                <th>Fecha</th>
            </tr>
        </thead>
        <tbody>
"""
        
        for doc in documents:
            doc_type = doc.get('document_type', '')
            confidence = doc.get('classification_confidence', 0) * 100
            
            html += f"""
            <tr>
                <td>{doc.get('document_id', '')[:15]}...</td>
                <td>{doc.get('original_filename', '')}</td>
                <td><span class="badge badge-{doc_type}">{doc_type}</span></td>
                <td>{confidence:.1f}%</td>
                <td>{doc.get('processed_at', '')[:10] if doc.get('processed_at') else ''}</td>
            </tr>
"""
        
        html += """
        </tbody>
    </table>
    
    <script>
        // Filtrado básico
        document.addEventListener('DOMContentLoaded', function() {
            console.log('Dashboard cargado');
        });
    </script>
</body>
</html>
"""
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
        
        self.logger.info(f"Exportado a HTML: {output_path}")
        return output_path

